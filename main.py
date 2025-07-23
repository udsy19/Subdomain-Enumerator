#!/usr/bin/env python3

import asyncio
import aiohttp
import aiodns
import ssl
import time
import os
import json
import sys
import hashlib
import random
from typing import Set, List, Dict, Tuple, Optional
from collections import defaultdict, Counter
from dataclasses import dataclass
import re
import ipaddress
from itertools import product, combinations
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@dataclass
class SubdomainResult:
    subdomain: str
    source: str
    http_status: int
    ip_addresses: List[str]
    technologies: List[str]
    confidence_score: float
    discovered_at: float
    response_time: Optional[float] = None
    title: Optional[str] = None
    server: Optional[str] = None


class IntelligentDNSResolver:
    """Advanced DNS resolver with multiple nameservers and intelligent failover"""
    
    def __init__(self):
        # Public DNS resolvers with different characteristics
        self.resolvers = [
            # Google DNS
            ['8.8.8.8', '8.8.4.4'],
            # Cloudflare DNS
            ['1.1.1.1', '1.0.0.1'],
            # Quad9 DNS
            ['9.9.9.9', '149.112.112.112'],
            # OpenDNS
            ['208.67.222.222', '208.67.220.220'],
            # Level3 DNS
            ['4.2.2.1', '4.2.2.2'],
            # Comodo DNS
            ['8.26.56.26', '8.20.247.20'],
            # DNS.WATCH
            ['84.200.69.80', '84.200.70.40'],
            # AdGuard DNS
            ['94.140.14.14', '94.140.15.15']
        ]
        
        self.resolver_pool = []
        self.resolver_stats = defaultdict(lambda: {'success': 0, 'failure': 0, 'avg_time': 0})
        self._setup_resolvers()
    
    def _setup_resolvers(self):
        """Initialize DNS resolver pool"""
        for resolver_pair in self.resolvers:
            for resolver_ip in resolver_pair:
                dns_resolver = aiodns.DNSResolver(nameservers=[resolver_ip], timeout=3, tries=2)
                self.resolver_pool.append((resolver_ip, dns_resolver))
    
    async def resolve_with_intelligence(self, subdomain: str) -> Tuple[bool, List[str], str]:
        """Intelligent DNS resolution with failover and performance tracking"""
        best_resolver = self._get_best_resolver()
        
        for resolver_ip, resolver in [best_resolver] + self.resolver_pool[:3]:
            start_time = time.time()
            try:
                result = await resolver.query(subdomain, 'A')
                response_time = time.time() - start_time
                
                # Update resolver statistics
                self.resolver_stats[resolver_ip]['success'] += 1
                self.resolver_stats[resolver_ip]['avg_time'] = (
                    self.resolver_stats[resolver_ip]['avg_time'] + response_time
                ) / 2
                
                ip_addresses = [str(r.host) for r in result]
                return True, ip_addresses, resolver_ip
                
            except Exception:
                self.resolver_stats[resolver_ip]['failure'] += 1
                continue
        
        return False, [], "none"
    
    def _get_best_resolver(self) -> Tuple[str, aiodns.DNSResolver]:
        """Select the best performing DNS resolver"""
        best_score = float('-inf')
        best_resolver = self.resolver_pool[0]
        
        for resolver_ip, resolver in self.resolver_pool:
            stats = self.resolver_stats[resolver_ip]
            total_queries = stats['success'] + stats['failure']
            
            if total_queries == 0:
                score = 0  # No data yet
            else:
                success_rate = stats['success'] / total_queries
                avg_time = stats['avg_time'] if stats['avg_time'] > 0 else 1.0
                score = success_rate / avg_time  # Higher success rate, lower time = better
            
            if score > best_score:
                best_score = score
                best_resolver = (resolver_ip, resolver)
        
        return best_resolver


class MLSubdomainPredictor:
    """Machine Learning-based subdomain prediction using pattern analysis"""
    
    def __init__(self):
        self.patterns = defaultdict(int)
        self.ngrams = defaultdict(int)
        self.common_prefixes = Counter()
        self.common_suffixes = Counter()
        self.length_distribution = Counter()
    
    def analyze_patterns(self, known_subdomains: List[str], domain: str):
        """Analyze patterns from known subdomains"""
        for subdomain in known_subdomains:
            if subdomain.endswith(f'.{domain}'):
                prefix = subdomain[:-len(f'.{domain}')]
                self._extract_features(prefix)
    
    def _extract_features(self, prefix: str):
        """Extract linguistic and structural features"""
        # Length distribution
        self.length_distribution[len(prefix)] += 1
        
        # N-gram analysis
        for n in range(2, min(5, len(prefix) + 1)):
            for i in range(len(prefix) - n + 1):
                ngram = prefix[i:i+n]
                self.ngrams[ngram] += 1
        
        # Common patterns
        parts = re.split(r'[-_.]', prefix)
        for part in parts:
            if len(part) >= 2:
                self.common_prefixes[part[:3]] += 1
                self.common_suffixes[part[-3:]] += 1
        
        # Structural patterns
        if '-' in prefix:
            self.patterns['has_dash'] += 1
        if '_' in prefix:
            self.patterns['has_underscore'] += 1
        if prefix.isdigit():
            self.patterns['is_numeric'] += 1
        if any(c.isdigit() for c in prefix):
            self.patterns['has_numbers'] += 1
    
    def generate_predictions(self, domain: str, count: int = 1000) -> List[str]:
        """Generate predicted subdomains based on learned patterns"""
        predictions = set()
        
        # Use most common n-grams to build new subdomains
        top_ngrams = [ngram for ngram, _ in self.ngrams.most_common(50)]
        top_prefixes = [prefix for prefix, _ in self.common_prefixes.most_common(20)]
        top_suffixes = [suffix for suffix, _ in self.common_suffixes.most_common(20)]
        
        # Generate combinations
        for prefix in top_prefixes:
            for suffix in top_suffixes:
                predictions.add(f"{prefix}{suffix}.{domain}")
                predictions.add(f"{prefix}-{suffix}.{domain}")
                predictions.add(f"{prefix}_{suffix}.{domain}")
        
        # Generate based on common patterns
        numbers = ['1', '2', '3', '01', '02', '03', '10', '11', '12']
        for prefix in top_prefixes:
            for num in numbers:
                predictions.add(f"{prefix}{num}.{domain}")
                predictions.add(f"{prefix}-{num}.{domain}")
        
        # Generate variations of known patterns
        environments = ['dev', 'test', 'prod', 'stage', 'qa', 'demo', 'beta']
        for env in environments:
            for prefix in top_prefixes[:10]:
                predictions.add(f"{env}-{prefix}.{domain}")
                predictions.add(f"{prefix}-{env}.{domain}")
        
        return list(predictions)[:count]


class AdvancedCTMiner:
    """Advanced Certificate Transparency mining with multiple sources"""
    
    def __init__(self, session: aiohttp.ClientSession):
        self.session = session
        self.ct_sources = [
            {'name': 'crt.sh', 'url': 'https://crt.sh/?q=%25.{domain}&output=json', 'weight': 1.0},
            {'name': 'certspotter', 'url': 'https://api.certspotter.com/v1/issuances?domain={domain}&include_subdomains=true&expand=dns_names', 'weight': 0.8},
            {'name': 'entrust', 'url': 'https://ctsearch.entrust.com/api/v1/certificates?fields=subjectDN&domain={domain}&includeExpired=true', 'weight': 0.6}
        ]
    
    async def comprehensive_mining(self, domain: str) -> List[SubdomainResult]:
        """Mine Certificate Transparency logs from multiple sources"""
        all_results = []
        
        tasks = []
        for source in self.ct_sources:
            task = self._mine_ct_source(source, domain)
            tasks.append(task)
        
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        for source, result in zip(self.ct_sources, results):
            if isinstance(result, list):
                print(f"✅ CT {source['name']}: {len(result)} certificates processed")
                all_results.extend(result)
            else:
                print(f"⚠️  CT {source['name']}: {result}")
        
        # Deduplicate and score results
        unique_subdomains = {}
        for result in all_results:
            if result.subdomain not in unique_subdomains:
                unique_subdomains[result.subdomain] = result
            else:
                # Merge confidence scores
                existing = unique_subdomains[result.subdomain]
                existing.confidence_score = max(existing.confidence_score, result.confidence_score)
        
        return list(unique_subdomains.values())
    
    async def _mine_ct_source(self, source: Dict, domain: str) -> List[SubdomainResult]:
        """Mine a specific CT source"""
        results = []
        
        try:
            url = source['url'].format(domain=domain)
            async with self.session.get(url, timeout=aiohttp.ClientTimeout(total=30)) as response:
                if response.status == 200:
                    data = await response.json()
                    results = self._parse_ct_data(data, source, domain)
                else:
                    raise Exception(f"HTTP {response.status}")
        except Exception as e:
            return []
        
        return results
    
    def _parse_ct_data(self, data: List[dict], source: Dict, domain: str) -> List[SubdomainResult]:
        """Parse Certificate Transparency data"""
        results = []
        
        for entry in data:
            subdomains = set()
            
            # Extract subdomains from different CT log formats
            if 'name_value' in entry:  # crt.sh format
                names = entry['name_value'].split('\n')
                subdomains.update(names)
            elif 'dns_names' in entry:  # certspotter format
                subdomains.update(entry['dns_names'])
            elif 'subjectDN' in entry:  # entrust format
                cn_match = re.search(r'CN=([^,]+)', entry['subjectDN'])
                if cn_match:
                    subdomains.add(cn_match.group(1))
            
            for subdomain in subdomains:
                if subdomain and self._is_valid_subdomain(subdomain, domain):
                    result = SubdomainResult(
                        subdomain=subdomain.lower(),
                        source=f"CT_{source['name']}",
                        http_status=0,
                        ip_addresses=[],
                        technologies=[],
                        confidence_score=source['weight'],
                        discovered_at=time.time()
                    )
                    results.append(result)
        
        return results
    
    def _is_valid_subdomain(self, subdomain: str, domain: str) -> bool:
        """Validate subdomain format"""
        if not subdomain or not subdomain.endswith(f'.{domain}'):
            return False
        if len(subdomain) > 253 or '..' in subdomain:
            return False
        if subdomain.startswith('*'):  # Skip wildcard certificates
            return False
        return bool(re.match(r'^[a-zA-Z0-9.*-]+$', subdomain))


class NetworkInfraAnalyzer:
    """Network infrastructure analysis for discovering related subdomains"""
    
    def __init__(self, session: aiohttp.ClientSession, dns_resolver: IntelligentDNSResolver):
        self.session = session
        self.dns_resolver = dns_resolver
    
    async def analyze_infrastructure(self, known_results: List[SubdomainResult], domain: str) -> List[SubdomainResult]:
        """Analyze network infrastructure to discover related subdomains"""
        new_results = []
        
        # Extract unique IP addresses
        ip_addresses = set()
        for result in known_results:
            ip_addresses.update(result.ip_addresses)
        
        print(f"🔍 Analyzing {len(ip_addresses)} unique IP addresses...")
        
        # Perform reverse DNS lookups
        reverse_dns_results = await self._reverse_dns_analysis(ip_addresses, domain)
        new_results.extend(reverse_dns_results)
        
        # Analyze IP ranges and subnets
        subnet_results = await self._subnet_analysis(ip_addresses, domain)
        new_results.extend(subnet_results)
        
        # Analyze ASN and hosting providers
        asn_results = await self._asn_analysis(ip_addresses, domain)
        new_results.extend(asn_results)
        
        return new_results
    
    async def _reverse_dns_analysis(self, ip_addresses: Set[str], domain: str) -> List[SubdomainResult]:
        """Perform reverse DNS lookups on discovered IP addresses"""
        results = []
        
        for ip in ip_addresses:
            try:
                # Reverse DNS lookup
                resolver = self.dns_resolver.resolver_pool[0][1]
                ptr_records = await resolver.query(ip, 'PTR')
                
                for record in ptr_records:
                    hostname = str(record.host).rstrip('.')
                    if hostname.endswith(f'.{domain}') and hostname != domain:
                        result = SubdomainResult(
                            subdomain=hostname,
                            source="Reverse_DNS",
                            http_status=0,
                            ip_addresses=[ip],
                            technologies=[],
                            confidence_score=0.8,
                            discovered_at=time.time()
                        )
                        results.append(result)
            except Exception:
                continue
        
        return results
    
    async def _subnet_analysis(self, ip_addresses: Set[str], domain: str) -> List[SubdomainResult]:
        """Analyze IP subnets for additional hosts"""
        results = []
        
        # Group IPs by /24 subnet
        subnets = defaultdict(list)
        for ip in ip_addresses:
            try:
                network = ipaddress.IPv4Network(f"{ip}/24", strict=False)
                subnets[str(network)].append(ip)
            except Exception:
                continue
        
        # For subnets with multiple known IPs, scan the range
        for subnet, ips in subnets.items():
            if len(ips) >= 2:  # Only scan subnets with multiple known IPs
                network = ipaddress.IPv4Network(subnet)
                
                # Scan common IP offsets in the subnet
                scan_ips = []
                for offset in [1, 2, 3, 10, 11, 12, 50, 51, 52, 100, 101, 102]:
                    scan_ip = str(network.network_address + offset)
                    if scan_ip not in ip_addresses:
                        scan_ips.append(scan_ip)
                
                # Perform reverse DNS on scan IPs
                subnet_results = await self._reverse_dns_analysis(set(scan_ips), domain)
                results.extend(subnet_results)
        
        return results
    
    async def _asn_analysis(self, ip_addresses: Set[str], domain: str) -> List[SubdomainResult]:
        """Analyze Autonomous System Numbers for hosting patterns"""
        # This would typically involve ASN databases
        # For now, return empty list as ASN analysis requires external APIs
        return []


class UltraRobustEnumerator:
    """Ultra-robust subdomain enumerator with advanced techniques"""
    
    def __init__(self):
        self.dns_resolver = IntelligentDNSResolver()
        self.ml_predictor = MLSubdomainPredictor()
        self.results = {}  # subdomain -> SubdomainResult
        
        # Advanced configuration
        self.config = {
            'max_concurrent_dns': 500,
            'max_concurrent_http': 100,
            'dns_timeout': 5,
            'http_timeout': 10,
            'confidence_threshold': 0.3,
            'max_recursion_depth': 3
        }
    
    def display_advanced_banner(self):
        """Display advanced banner"""
        banner = """
╔══════════════════════════════════════════════════════════════╗
║             ULTRA-ROBUST SUBDOMAIN ENUMERATOR                ║
║           Advanced AI-Powered Reconnaissance v3.0            ║
╚══════════════════════════════════════════════════════════════╝

🚀 Advanced Features:
• Multi-Resolver Intelligent DNS with Failover
• Machine Learning Subdomain Prediction
• Advanced Certificate Transparency Mining (3+ sources)
• Network Infrastructure Analysis & Reverse DNS
• Recursive Subdomain Discovery
• Technology Detection & Fingerprinting
• Performance-Optimized with Real-time Statistics

        """
        print(banner)
    
    def get_advanced_config(self) -> Dict:
        """Get advanced configuration from user"""
        print("⚙️  ADVANCED CONFIGURATION")
        print("=" * 50)
        
        # Domain input
        while True:
            domain = input("🎯 Target domain: ").strip().lower()
            if domain and '.' in domain and not domain.startswith(('http://', 'https://')):
                break
            print("❌ Please enter a valid domain name")
        
        # Enumeration mode
        print("\n🔬 Enumeration Mode:")
        print("1. Standard - Balanced speed and coverage")
        print("2. Aggressive - Maximum coverage, slower")
        print("3. Stealth - Minimal footprint, longer timeouts")
        print("4. Lightning - Speed focused, basic techniques")
        
        while True:
            try:
                mode = int(input("Select mode (1-4): "))
                if 1 <= mode <= 4:
                    break
                print("❌ Please enter 1, 2, 3, or 4")
            except ValueError:
                print("❌ Please enter a valid number")
        
        # Wordlist size
        print("\n📚 Wordlist Configuration:")
        print("1. Compact (10,000 subdomains)")
        print("2. Standard (50,000 subdomains)")
        print("3. Extensive (110,000 subdomains)")
        print("4. Custom + ML Predictions")
        
        while True:
            try:
                wordlist_choice = int(input("Select wordlist (1-4): "))
                if 1 <= wordlist_choice <= 4:
                    break
                print("❌ Please enter 1, 2, 3, or 4")
            except ValueError:
                print("❌ Please enter a valid number")
        
        # Map choices to configuration
        mode_configs = {
            1: {'threads': 200, 'timeout': 5, 'http_workers': 50},    # Standard
            2: {'threads': 400, 'timeout': 8, 'http_workers': 100},  # Aggressive
            3: {'threads': 50, 'timeout': 15, 'http_workers': 20},   # Stealth
            4: {'threads': 500, 'timeout': 3, 'http_workers': 200}   # Lightning
        }
        
        wordlist_sizes = {1: 10000, 2: 50000, 3: 110000, 4: 25000}
        
        config = mode_configs[mode]
        config.update({
            'domain': domain,
            'wordlist_size': wordlist_sizes[wordlist_choice],
            'use_ml_predictions': wordlist_choice == 4,
            'mode': ['Standard', 'Aggressive', 'Stealth', 'Lightning'][mode-1]
        })
        
        return config
    
    async def ultra_enumerate(self, domain_or_config, mode: int = 1, wordlist: int = 2) -> Dict[str, SubdomainResult]:
        """Ultra-robust enumeration with all advanced techniques"""
        
        # Handle both config dict and individual parameters
        if isinstance(domain_or_config, dict):
            config = domain_or_config
            domain = config['domain']
        else:
            domain = domain_or_config
            # Map choices to configuration
            mode_configs = {
                1: {'threads': 200, 'timeout': 5, 'http_workers': 50},    # Standard
                2: {'threads': 400, 'timeout': 8, 'http_workers': 100},  # Aggressive
                3: {'threads': 50, 'timeout': 15, 'http_workers': 20},   # Stealth
                4: {'threads': 500, 'timeout': 3, 'http_workers': 200}   # Lightning
            }
            
            wordlist_sizes = {1: 10000, 2: 50000, 3: 110000, 4: 25000}
            
            config = mode_configs[mode]
            config.update({
                'domain': domain,
                'wordlist_size': wordlist_sizes[wordlist],
                'use_ml_predictions': wordlist == 4,
                'mode': ['Standard', 'Aggressive', 'Stealth', 'Lightning'][mode-1]
            })
        
        print(f"\n🎯 ULTRA-ROBUST ENUMERATION: {domain.upper()}")
        print(f"Mode: {config['mode']} | Wordlist: {config['wordlist_size']:,} | Threads: {config['threads']}")
        print("=" * 80)
        
        start_time = time.time()
        
        # Update internal configuration
        self.config['max_concurrent_dns'] = config['threads']
        self.config['max_concurrent_http'] = config['http_workers']
        self.config['dns_timeout'] = config['timeout']
        
        # Create HTTP session
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        connector = aiohttp.TCPConnector(
            ssl=ssl_context, 
            limit=config['http_workers'] * 2,
            limit_per_host=20,
            ttl_dns_cache=300
        )
        timeout = aiohttp.ClientTimeout(total=config['timeout'] * 2)
        
        async with aiohttp.ClientSession(connector=connector, timeout=timeout) as session:
            # Phase 1: Advanced Certificate Transparency Mining
            await self._phase_ct_mining(session, domain)
            
            # Phase 2: Intelligent DNS Brute Force
            await self._phase_intelligent_dns_bruteforce(domain, config)
            
            # Phase 3: Machine Learning Predictions (if enabled)
            if config['use_ml_predictions']:
                await self._phase_ml_predictions(domain)
            
            # Phase 4: Network Infrastructure Analysis
            await self._phase_infrastructure_analysis(session, domain)
            
            # Phase 5: Recursive Discovery
            await self._phase_recursive_discovery(session, domain)
            
            # Phase 6: Comprehensive HTTP Analysis
            await self._phase_http_analysis(session)
        
        elapsed = time.time() - start_time
        
        print("=" * 80)
        print(f"✅ Ultra-robust enumeration completed in {elapsed:.2f} seconds")
        print(f"🎯 Total unique subdomains discovered: {len(self.results)}")
        self._print_discovery_stats()
        
        return self.results
    
    async def _phase_ct_mining(self, session: aiohttp.ClientSession, domain: str):
        """Phase 1: Advanced Certificate Transparency Mining"""
        print("\n📜 PHASE 1: Advanced Certificate Transparency Mining")
        print("-" * 60)
        
        ct_miner = AdvancedCTMiner(session)
        ct_results = await ct_miner.comprehensive_mining(domain)
        
        # Add to main results
        for result in ct_results:
            self.results[result.subdomain] = result
        
        print(f"📊 CT Mining: {len(ct_results)} unique subdomains from multiple sources")
    
    async def _phase_intelligent_dns_bruteforce(self, domain: str, config: Dict):
        """Phase 2: Intelligent DNS Brute Force"""
        print(f"\n🎯 PHASE 2: Intelligent DNS Brute Force ({config['wordlist_size']:,} candidates)")
        print("-" * 60)
        
        # Load wordlist
        wordlist = self._load_comprehensive_wordlist(config['wordlist_size'])
        candidates = [f"{word}.{domain}" for word in wordlist]
        
        print(f"🔄 Testing {len(candidates):,} candidates with {len(self.dns_resolver.resolver_pool)} DNS resolvers...")
        
        # Process in optimized batches
        batch_size = 5000
        total_found = 0
        
        semaphore = asyncio.Semaphore(self.config['max_concurrent_dns'])
        
        for i in range(0, len(candidates), batch_size):
            batch = candidates[i:i + batch_size]
            batch_num = i // batch_size + 1
            total_batches = (len(candidates) + batch_size - 1) // batch_size
            
            print(f"🔄 Processing batch {batch_num}/{total_batches} ({len(batch):,} candidates) - Progress: {(batch_num/total_batches)*100:.1f}%")
            
            # Progress tracking
            batch_start_time = time.time()
            tasks = [self._resolve_with_intelligence(semaphore, candidate) for candidate in batch]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            batch_time = time.time() - batch_start_time
            
            batch_found = sum(1 for result in results if isinstance(result, SubdomainResult))
            total_found += batch_found
            
            # Add valid results
            for result in results:
                if isinstance(result, SubdomainResult):
                    self.results[result.subdomain] = result
            
            # Enhanced progress reporting
            rate = len(batch) / batch_time if batch_time > 0 else 0
            estimated_remaining = ((total_batches - batch_num) * batch_time) if batch_time > 0 else 0
            
            if batch_found > 0:
                print(f"✅ Batch {batch_num}: {batch_found} valid subdomains found | Rate: {rate:.0f} queries/sec | ETA: {estimated_remaining:.0f}s")
            else:
                print(f"⚪ Batch {batch_num}: No new subdomains | Rate: {rate:.0f} queries/sec | ETA: {estimated_remaining:.0f}s")
        
        print(f"📊 Intelligent DNS: {total_found} subdomains with {self._get_resolver_stats()}")
    
    async def _resolve_with_intelligence(self, semaphore: asyncio.Semaphore, subdomain: str) -> Optional[SubdomainResult]:
        """Resolve with intelligent DNS and create result object"""
        async with semaphore:
            success, ip_addresses, resolver_used = await self.dns_resolver.resolve_with_intelligence(subdomain)
            
            if success:
                return SubdomainResult(
                    subdomain=subdomain,
                    source=f"DNS_Intelligence_{resolver_used}",
                    http_status=0,
                    ip_addresses=ip_addresses,
                    technologies=[],
                    confidence_score=0.9,
                    discovered_at=time.time()
                )
            return None
    
    async def _phase_ml_predictions(self, domain: str):
        """Phase 3: Machine Learning Subdomain Predictions"""
        print("\n🤖 PHASE 3: Machine Learning Subdomain Predictions")
        print("-" * 60)
        
        # Train the model on existing results
        known_subdomains = list(self.results.keys())
        if len(known_subdomains) >= 10:  # Need minimum data to train
            self.ml_predictor.analyze_patterns(known_subdomains, domain)
            
            # Generate predictions
            print("🔬 Generating ML-based subdomain predictions...")
            predictions = self.ml_predictor.generate_predictions(domain, 2000)
            print(f"🧠 Generated {len(predictions)} ML-based predictions from pattern analysis")
            
            # Test predictions with DNS
            print(f"🔄 Testing {len(predictions):,} ML predictions with DNS resolution...")
            prediction_start_time = time.time()
            
            semaphore = asyncio.Semaphore(self.config['max_concurrent_dns'])
            tasks = [self._resolve_with_intelligence(semaphore, pred) for pred in predictions]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            prediction_time = time.time() - prediction_start_time
            
            ml_found = 0
            for result in results:
                if isinstance(result, SubdomainResult) and result.subdomain not in self.results:
                    result.source = "ML_Prediction"
                    result.confidence_score = 0.7
                    self.results[result.subdomain] = result
                    ml_found += 1
            
            rate = len(predictions) / prediction_time if prediction_time > 0 else 0
            print(f"📊 ML Predictions: {ml_found} new subdomains discovered in {prediction_time:.1f}s | Rate: {rate:.0f} predictions/sec")
        else:
            print("⚠️  Insufficient data for ML training, skipping...")
    
    async def _phase_infrastructure_analysis(self, session: aiohttp.ClientSession, domain: str):
        """Phase 4: Network Infrastructure Analysis"""
        print("\n🌐 PHASE 4: Network Infrastructure Analysis")
        print("-" * 60)
        
        analyzer = NetworkInfraAnalyzer(session, self.dns_resolver)
        current_results = list(self.results.values())
        
        # Extract unique IPs for analysis
        unique_ips = set()
        for result in current_results:
            unique_ips.update(result.ip_addresses)
        
        print(f"🔍 Analyzing network infrastructure from {len(unique_ips)} unique IP addresses...")
        
        infra_start_time = time.time()
        infra_results = await analyzer.analyze_infrastructure(current_results, domain)
        infra_time = time.time() - infra_start_time
        
        # Add new discoveries
        new_count = 0
        for result in infra_results:
            if result.subdomain not in self.results:
                self.results[result.subdomain] = result
                new_count += 1
        
        print(f"📊 Infrastructure Analysis: {new_count} new subdomains discovered in {infra_time:.1f}s")
    
    async def _phase_recursive_discovery(self, session: aiohttp.ClientSession, domain: str):
        """Phase 5: Recursive Subdomain Discovery"""
        print("\n🔄 PHASE 5: Recursive Subdomain Discovery")
        print("-" * 60)
        
        # Find subdomains that might have their own subdomains
        recursive_candidates = []
        for result in self.results.values():
            subdomain = result.subdomain
            # Skip direct subdomains, look for nested ones
            if subdomain.count('.') == domain.count('.') + 1:
                recursive_candidates.append(subdomain)
        
        if recursive_candidates:
            # Test common nested subdomain patterns
            nested_patterns = ['www', 'api', 'app', 'admin', 'secure', 'mail', 'ftp']
            total_tests = len(recursive_candidates) * len(nested_patterns)
            
            print(f"🔍 Testing {len(recursive_candidates)} subdomains for nested patterns...")
            print(f"🔄 Generating {total_tests:,} recursive combinations ({len(nested_patterns)} patterns × {len(recursive_candidates)} candidates)")
            
            recursive_start_time = time.time()
            tasks = []
            
            for candidate in recursive_candidates:
                for pattern in nested_patterns:
                    nested_subdomain = f"{pattern}.{candidate}"
                    task = self._resolve_with_intelligence(
                        asyncio.Semaphore(50), 
                        nested_subdomain
                    )
                    tasks.append(task)
            
            if tasks:
                results = await asyncio.gather(*tasks, return_exceptions=True)
                recursive_time = time.time() - recursive_start_time
                
                recursive_found = 0
                for result in results:
                    if isinstance(result, SubdomainResult) and result.subdomain not in self.results:
                        result.source = "Recursive_Discovery"
                        result.confidence_score = 0.8
                        self.results[result.subdomain] = result
                        recursive_found += 1
                
                rate = len(tasks) / recursive_time if recursive_time > 0 else 0
                print(f"📊 Recursive Discovery: {recursive_found} nested subdomains found in {recursive_time:.1f}s | Rate: {rate:.0f} tests/sec")
        else:
            print("⚠️  No suitable candidates for recursive discovery")
    
    async def _phase_http_analysis(self, session: aiohttp.ClientSession):
        """Phase 6: Comprehensive HTTP Analysis"""
        print(f"\n🔍 PHASE 6: Comprehensive HTTP Analysis ({len(self.results)} subdomains)")
        print("-" * 60)
        
        if not self.results:
            print("⚠️  No subdomains to analyze")
            return
        
        # Progress tracking
        analysis_start_time = time.time()
        total_subdomains = len(self.results)
        completed = 0
        
        print(f"🌐 Starting HTTP analysis for {total_subdomains:,} subdomains...")
        
        semaphore = asyncio.Semaphore(self.config['max_concurrent_http'])
        
        # Create tasks with progress callback
        async def analyze_with_progress(result):
            nonlocal completed
            await self._analyze_http_comprehensive(semaphore, session, result)
            completed += 1
            
            # Progress reporting every 10% or every 50 completions
            if completed % max(1, total_subdomains // 10) == 0 or completed % 50 == 0:
                progress_pct = (completed / total_subdomains) * 100
                elapsed = time.time() - analysis_start_time
                rate = completed / elapsed if elapsed > 0 else 0
                eta = (total_subdomains - completed) / rate if rate > 0 else 0
                
                print(f"🔄 HTTP Analysis Progress: {completed:,}/{total_subdomains:,} ({progress_pct:.1f}%) | Rate: {rate:.1f}/sec | ETA: {eta:.0f}s")
        
        tasks = [analyze_with_progress(result) for result in self.results.values()]
        await asyncio.gather(*tasks, return_exceptions=True)
        
        analysis_time = time.time() - analysis_start_time
        
        # Statistics
        status_counts = Counter(result.http_status for result in self.results.values())
        live_services = sum(1 for result in self.results.values() if result.http_status == 200)
        
        print(f"✅ HTTP Analysis completed in {analysis_time:.1f}s")
        print("📊 HTTP Status Distribution:")
        for status, count in sorted(status_counts.items()):
            status_name = {
                0: "No Response",
                200: "OK",
                301: "Moved Permanently", 
                302: "Found",
                403: "Forbidden",
                404: "Not Found",
                500: "Internal Error"
            }.get(status, f"HTTP {status}")
            print(f"   {status_name} ({status}): {count} subdomains")
        
        print(f"🌐 Live Services Detected: {live_services} subdomains responding with HTTP 200")
    
    async def _analyze_http_comprehensive(self, semaphore: asyncio.Semaphore, session: aiohttp.ClientSession, result: SubdomainResult):
        """Comprehensive HTTP analysis with technology detection"""
        async with semaphore:
            for protocol in ['https', 'http']:
                try:
                    url = f"{protocol}://{result.subdomain}"
                    start_time = time.time()
                    
                    async with session.get(url, allow_redirects=False) as response:
                        result.http_status = response.status
                        result.response_time = time.time() - start_time
                        
                        # Extract headers for technology detection
                        server = response.headers.get('Server', '')
                        if server:
                            result.server = server
                            result.technologies.append(server)
                        
                        # Extract other technology indicators
                        powered_by = response.headers.get('X-Powered-By', '')
                        if powered_by:
                            result.technologies.append(powered_by)
                        
                        # Try to get page title
                        if response.status == 200:
                            try:
                                text = await response.text()
                                title_match = re.search(r'<title>([^<]+)</title>', text, re.IGNORECASE)
                                if title_match:
                                    result.title = title_match.group(1).strip()
                            except:
                                pass
                        
                        return  # Successfully analyzed
                        
                except Exception:
                    continue
            
            # If both protocols failed
            result.http_status = 0
    
    def _load_comprehensive_wordlist(self, size: int) -> List[str]:
        """Load comprehensive wordlist"""
        wordlist_paths = [
            'wordlists/subdomains-top1million-110000.txt',
            'wordlists/top-1000.txt',
            'wordlists/dns-records.txt',
            'wordlists/cloud-services.txt'
        ]
        
        for path in wordlist_paths:
            try:
                with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                    wordlist = [line.strip() for line in f if line.strip() and not line.startswith('#')]
                    return wordlist[:size]
            except FileNotFoundError:
                continue
        
        # Fallback comprehensive wordlist
        return [
            'www', 'mail', 'email', 'webmail', 'admin', 'administrator', 'root', 'test', 'testing',
            'dev', 'development', 'stage', 'staging', 'prod', 'production', 'demo', 'beta', 'alpha',
            'api', 'apis', 'app', 'apps', 'web', 'www1', 'www2', 'secure', 'security', 'ssl',
            'vpn', 'ftp', 'sftp', 'ssh', 'remote', 'access', 'portal', 'gateway', 'proxy',
        ] * (size // 30)  # Repeat to reach desired size
    
    def _get_resolver_stats(self) -> str:
        """Get DNS resolver performance statistics"""
        stats = []
        for resolver_ip, data in self.dns_resolver.resolver_stats.items():
            total = data['success'] + data['failure']
            if total > 0:
                success_rate = (data['success'] / total) * 100
                stats.append(f"{resolver_ip}: {success_rate:.1f}%")
        
        return f"Resolver performance: {', '.join(stats[:3])}"
    
    def _print_discovery_stats(self):
        """Print detailed discovery statistics"""
        source_stats = Counter(result.source for result in self.results.values())
        
        print("\n📊 Discovery Sources:")
        for source, count in source_stats.most_common():
            print(f"   • {source}: {count} subdomains")
        
        # Confidence distribution
        high_confidence = sum(1 for r in self.results.values() if r.confidence_score >= 0.8)
        medium_confidence = sum(1 for r in self.results.values() if 0.5 <= r.confidence_score < 0.8)
        low_confidence = sum(1 for r in self.results.values() if r.confidence_score < 0.5)
        
        print(f"\n🎯 Confidence Distribution:")
        print(f"   • High (≥80%): {high_confidence} subdomains")
        print(f"   • Medium (50-79%): {medium_confidence} subdomains")  
        print(f"   • Low (<50%): {low_confidence} subdomains")
    
    def save_advanced_excel(self, results: Dict[str, SubdomainResult], domain: str):
        """Save results to advanced Excel file with comprehensive data"""
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"output/{domain}_ultra_robust_{timestamp}.xlsx"
        
        os.makedirs('output', exist_ok=True)
        
        # Create workbook with multiple sheets
        wb = openpyxl.Workbook()
        
        # Main results sheet
        ws_main = wb.active
        ws_main.title = "Subdomain Discovery"
        
        # Headers for main sheet
        headers = [
            'Subdomain', 'Source', 'HTTP_Status', 'Status_Explanation', 'IP_Addresses', 
            'Response_Time', 'Technologies', 'Page_Title', 'Server', 'Confidence_Score'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        status_explanations = {
            0: "No Response - Server unreachable",
            200: "OK - Server responding normally",
            301: "Moved Permanently - Redirected",
            302: "Found - Temporary redirect",
            400: "Bad Request - Invalid request",
            401: "Unauthorized - Authentication required",
            403: "Forbidden - Access denied",
            404: "Not Found - Resource not found",
            500: "Internal Server Error - Server malfunction"
        }
        
        confidence_colors = {
            'high': "E6FFE6",    # Light green
            'medium': "FFF2CC",  # Light yellow
            'low': "FFE6E6"      # Light red
        }
        
        row = 2
        for result in sorted(results.values(), key=lambda x: x.confidence_score, reverse=True):
            # Subdomain
            ws_main.cell(row=row, column=1, value=result.subdomain)
            
            # Source
            ws_main.cell(row=row, column=2, value=result.source)
            
            # HTTP Status with color coding
            status_cell = ws_main.cell(row=row, column=3, value=result.http_status)
            if result.http_status == 200:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif result.http_status in [301, 302]:
                status_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            elif result.http_status >= 400:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Status explanation
            explanation = status_explanations.get(result.http_status, f"HTTP {result.http_status}")
            ws_main.cell(row=row, column=4, value=explanation)
            
            # IP Addresses
            ws_main.cell(row=row, column=5, value=", ".join(result.ip_addresses))
            
            # Response Time
            if result.response_time:
                ws_main.cell(row=row, column=6, value=f"{result.response_time:.3f}s")
            
            # Technologies
            ws_main.cell(row=row, column=7, value=", ".join(result.technologies))
            
            # Page Title
            if result.title:
                ws_main.cell(row=row, column=8, value=result.title[:100])  # Truncate long titles
            
            # Server
            if result.server:
                ws_main.cell(row=row, column=9, value=result.server)
            
            # Confidence Score with color coding
            conf_cell = ws_main.cell(row=row, column=10, value=f"{result.confidence_score:.2f}")
            if result.confidence_score >= 0.8:
                conf_cell.fill = PatternFill(start_color=confidence_colors['high'], end_color=confidence_colors['high'], fill_type="solid")
            elif result.confidence_score >= 0.5:
                conf_cell.fill = PatternFill(start_color=confidence_colors['medium'], end_color=confidence_colors['medium'], fill_type="solid")
            else:
                conf_cell.fill = PatternFill(start_color=confidence_colors['low'], end_color=confidence_colors['low'], fill_type="solid")
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws_main[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_main.column_dimensions[column].width = adjusted_width
        
        # Add statistics sheet
        ws_stats = wb.create_sheet("Statistics")
        self._create_statistics_sheet(ws_stats, results)
        
        # Save workbook
        wb.save(filename)
        
        print(f"💾 Advanced Excel report saved: {filename}")
        print(f"📊 Contains {len(results)} subdomains with comprehensive analysis")
        return filename
    
    def _create_statistics_sheet(self, ws, results: Dict[str, SubdomainResult]):
        """Create statistics sheet in Excel workbook"""
        # Source distribution
        source_stats = Counter(result.source for result in results.values())
        
        ws.cell(row=1, column=1, value="Discovery Source Statistics").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value="Source").font = Font(bold=True)
        ws.cell(row=2, column=2, value="Count").font = Font(bold=True)
        
        row = 3
        for source, count in source_stats.most_common():
            ws.cell(row=row, column=1, value=source)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # HTTP Status distribution
        status_stats = Counter(result.http_status for result in results.values())
        
        ws.cell(row=row + 2, column=1, value="HTTP Status Distribution").font = Font(bold=True, size=14)
        ws.cell(row=row + 3, column=1, value="Status Code").font = Font(bold=True)
        ws.cell(row=row + 3, column=2, value="Count").font = Font(bold=True)
        
        row += 4
        for status, count in sorted(status_stats.items()):
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=count)
            row += 1


async def main():
    """Main function for CLI enumeration"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Ultra-Robust Subdomain Enumerator')
    parser.add_argument('domain', help='Target domain to enumerate')
    parser.add_argument('--mode', type=int, default=1, choices=[1, 2, 3, 4], 
                       help='Enumeration mode (1=Standard, 2=Aggressive, 3=Stealth, 4=Lightning)')
    parser.add_argument('--wordlist', type=int, default=2, choices=[1, 2, 3, 4],
                       help='Wordlist size (1=Compact, 2=Standard, 3=Extensive, 4=Custom+ML)')
    
    args = parser.parse_args()
    
    if not args.domain:
        print("❌ Domain is required")
        sys.exit(1)
    
    try:
        enumerator = UltraRobustEnumerator()
        
        # Display banner
        enumerator.display_advanced_banner()
        
        # Run ultra-robust enumeration with the same workflow as main_tui
        results = await enumerator.ultra_enumerate(args.domain, args.mode, args.wordlist)
        
        # Save advanced Excel report
        filename = enumerator.save_advanced_excel(results, args.domain)
        
        # Final summary
        print(f"\n{'='*80}")
        print("🎯 ULTRA-ROBUST ENUMERATION COMPLETE")
        print(f"{'='*80}")
        
        if results:
            high_confidence = sum(1 for r in results.values() if r.confidence_score >= 0.8)
            live_subdomains = sum(1 for r in results.values() if r.http_status == 200)
            
            print(f"✅ Total subdomains discovered: {len(results)}")
            print(f"🎯 High confidence results: {high_confidence}")
            print(f"🌐 Live HTTP services: {live_subdomains}")
            print(f"📄 Advanced report: {filename}")
            print(f"\n🚀 Features used:")
            print(f"   • Multi-resolver intelligent DNS")
            print(f"   • Advanced Certificate Transparency mining")
            print(f"   • Network infrastructure analysis")
            print(f"   • Technology detection & fingerprinting")
            print(f"   • Performance optimization & statistics")
        else:
            print("⚠️  No subdomains discovered with current configuration")
        
        print(f"\n💡 Ultra-Robust Subdomain Enumerator v3.0 - Mission Complete!")
        
    except KeyboardInterrupt:
        print("\n⚠️  Enumeration interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Critical error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    # Check advanced dependencies
    try:
        import openpyxl
        import aiodns
        import aiohttp
        import ipaddress
    except ImportError as e:
        print(f"❌ Missing advanced dependency: {e}")
        print("📦 Install with: pip install openpyxl aiodns aiohttp ipaddress")
        exit(1)
    
    asyncio.run(main())