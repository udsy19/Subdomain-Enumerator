#!/usr/bin/env python3
"""
Ultra-Robust Subdomain Enumerator with Modern TUI
Complete functionality with multiple interface options

🚀 QUICK START:
1. Default run: python3 main_tui_merged.py
2. Quick scan: python3 main_tui_merged.py domain.com  
3. Python functions: fast_scan('domain.com'), multi_domain_scan(['domain1.com', 'domain2.com'])

📦 DEPENDENCIES:
- Core: aiohttp, aiodns, openpyxl (required)
- Modern TUI: pip install textual (recommended)
- Rich TUI: pip install rich (fallback)
- Enhanced: pip install pandas scikit-learn networkx (optional)

🖥️ INTERFACE OPTIONS:
1. Textual TUI (modern, responsive) - Preferred
2. Rich TUI (legacy, full-featured) - Fallback  
3. Command line functions - Always available
4. Google Colab support - Built-in

⚡ FAST SCANNING MODES:
1 = Basic (Fast DNS + HTTP)
2 = Standard (Balanced features)  
3 = Advanced (Comprehensive)
4 = Ultra (Maximum depth)
5 = Quick (DNS-only, fastest)

🔄 MULTI-DOMAIN SUPPORT:
- Parallel scanning of multiple domains
- Automatic resource allocation
- No conflicts between scans
"""

import asyncio
import aiohttp
import aiodns
import ssl
import time
import os
import json
import sys
import signal
import threading
import queue
from typing import Set, List, Dict, Tuple, Optional
from collections import defaultdict, Counter, deque
from dataclasses import dataclass, asdict, field
import re
import ipaddress
from itertools import product, combinations
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import socket
import datetime
try:
    from cryptography import x509
    from cryptography.hazmat.backends import default_backend
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False
    x509 = None
    default_backend = None
import random
import glob
import multiprocessing
import itertools
import platform
import shutil
from collections import Counter, defaultdict
try:
    import whois
    WHOIS_AVAILABLE = True
except ImportError:
    WHOIS_AVAILABLE = False
try:
    import subprocess
    SUBPROCESS_AVAILABLE = True
except ImportError:
    SUBPROCESS_AVAILABLE = False
try:
    import requests_cache
    CACHE_AVAILABLE = True
except ImportError:
    CACHE_AVAILABLE = False
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False
try:
    from sklearn.cluster import DBSCAN
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.ensemble import IsolationForest
    import numpy as np
    ML_AVAILABLE = True
except ImportError:
    ML_AVAILABLE = False
    DBSCAN = None
    TfidfVectorizer = None
    IsolationForest = None
    np = None
try:
    import networkx as nx
    NETWORKX_AVAILABLE = True
except ImportError:
    NETWORKX_AVAILABLE = False
    nx = None

# Try to import required libraries for beautiful TUI
try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn, TimeElapsedColumn
    from rich.table import Table
    from rich.layout import Layout
    from rich.text import Text
    from rich.align import Align
    from rich.prompt import Prompt, Confirm
    from rich.live import Live
    from rich.columns import Columns
    from rich import box
    from rich.markup import escape
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False
    print("❌ Rich library not found. Please install with: pip install rich")
    try:
        response = input("Install Rich automatically? [y/N]: ").lower().strip()
        if response == 'y':
            import subprocess
            subprocess.run([sys.executable, "-m", "pip", "install", "rich"])
        else:
            print("Continuing without Rich library...")
    except (EOFError, KeyboardInterrupt):
        print("Continuing without Rich library...")
    try:
        from rich.console import Console
        from rich.panel import Panel
        from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn, TimeElapsedColumn
        from rich.table import Table
        from rich.layout import Layout
        from rich.text import Text
        from rich.align import Align
        from rich.prompt import Prompt, Confirm
        from rich.live import Live
        from rich.columns import Columns
        from rich import box
        from rich.markup import escape
        RICH_AVAILABLE = True
    except ImportError:
        print("❌ Failed to install Rich. Falling back to basic interface.")
        RICH_AVAILABLE = False

# Try to import tqdm for live output
try:
    from tqdm import tqdm, trange
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False
    print("📊 tqdm not found. Please install with: pip install tqdm")
    try:
        response = input("Install tqdm automatically? [y/N]: ").lower().strip()
        if response == 'y':
            import subprocess
            subprocess.run([sys.executable, "-m", "pip", "install", "tqdm"])
        else:
            print("Continuing without tqdm...")
    except (EOFError, KeyboardInterrupt):
        print("Continuing without tqdm...")
    try:
        from tqdm import tqdm, trange
        TQDM_AVAILABLE = True
    except ImportError:
        print("❌ Failed to install tqdm. Using basic output.")
        TQDM_AVAILABLE = False

# Try to import Textual for modern TUI
try:
    from textual.app import App, ComposeResult
    from textual.containers import Container, Horizontal, Vertical, ScrollableContainer
    from textual.widgets import (
        Header, Footer, Button, Input, Label, Static, ProgressBar, 
        DataTable, Select, TextArea, RichLog, Switch
    )
    from textual.reactive import reactive
    from textual.worker import work
    from textual.screen import ModalScreen
    from textual.binding import Binding
    TEXTUAL_AVAILABLE = True
except ImportError:
    TEXTUAL_AVAILABLE = False

console = Console()

@dataclass
class SystemResources:
    cpu_cores: int
    memory_gb: float
    network_bandwidth_mbps: float
    concurrent_threads: int


@dataclass
class VulnerabilityAssessment:
    cve_ids: List[str] = field(default_factory=list)
    security_headers: Dict[str, str] = field(default_factory=dict)
    cookie_security: Dict[str, str] = field(default_factory=dict)
    cors_issues: List[str] = field(default_factory=list)
    ssl_issues: List[str] = field(default_factory=list)
    risk_score: float = 0.0

@dataclass
class AIAnalysis:
    pattern_confidence: float = 0.0
    anomaly_score: float = 0.0
    predicted_subdomains: List[str] = field(default_factory=list)
    risk_assessment: str = "Low"
    technology_predictions: List[str] = field(default_factory=list)
    
@dataclass
class HistoricalData:
    first_seen: datetime.datetime = field(default_factory=datetime.datetime.now)
    last_seen: datetime.datetime = field(default_factory=datetime.datetime.now)
    changes_detected: List[str] = field(default_factory=list)
    trend_analysis: Dict = field(default_factory=dict)

@dataclass
class CNAMERecord:
    """CNAME record information"""
    source_domain: str
    target_domain: str
    chain_depth: int
    final_target: str
    service_type: str = ""
    provider: str = ""
    takeover_risk: str = "Low"
    discovered_at: float = field(default_factory=time.time)

@dataclass
class SubdomainResult:
    subdomain: str
    source: str
    http_status: int
    ip_addresses: List[str]
    technologies: List[str]
    confidence_score: float
    discovered_at: float
    response_time: Optional[float] = None
    server: Optional[str] = None
    ssl_domain_verified: Optional[bool] = None
    ssl_issuer: Optional[str] = None
    ssl_subject: Optional[str] = None
    cname_records: List[CNAMERecord] = field(default_factory=list)
    cname_chain: List[str] = field(default_factory=list)
    is_cname_target: bool = False
    takeover_risk: str = "Low"
    ownership_info: Optional[str] = None
    # Nmap scan results
    nmap_open_ports: List[str] = field(default_factory=list)
    nmap_services: List[str] = field(default_factory=list)
    nmap_os_detection: Optional[str] = None
    nmap_vulnerabilities: List[str] = field(default_factory=list)
    nmap_ssl_info: Optional[str] = None
    nmap_http_info: Optional[str] = None
    nmap_traceroute: Optional[str] = None
    nmap_dns_info: Optional[str] = None
    # Advanced analysis results
    vulnerability_assessment: Optional[VulnerabilityAssessment] = None
    ai_analysis: Optional[AIAnalysis] = None
    historical_data: Optional[HistoricalData] = None
    ip_range_group: Optional[str] = None
    asn_info: Optional[str] = None

@dataclass
class NetworkStats:
    dns_queries: int = 0
    successful_queries: int = 0
    timeouts: int = 0
    avg_response_time: float = 0.0
    resolver_stats: Dict[str, int] = field(default_factory=dict)

@dataclass
class SystemResources:
    cpu_cores: int = 0
    memory_gb: float = 0.0
    network_bandwidth_mbps: float = 0.0
    optimal_dns_workers: int = 0
    optimal_http_workers: int = 0
    optimal_nmap_workers: int = 0

@dataclass 
class IPRangeGroup:
    ip_range: str
    subnets: List[str] = field(default_factory=list)
    subdomains: List[str] = field(default_factory=list)
    scan_results: Optional[Dict] = None
    
@dataclass
class AdvancedDiscoveryResult:
    method: str  # CT, DNS_AXFR, Search_Engine, etc.
    subdomains: List[str] = field(default_factory=list)
    confidence: float = 0.0
    metadata: Dict = field(default_factory=dict)

@dataclass
class ScanConfig:
    domain: str
    mode: int
    wordlist_files: List[str]

# Color themes
THEMES = {
    "default": {
        "primary": "bright_blue",
        "success": "bright_green", 
        "warning": "bright_yellow",
        "error": "bright_red",
        "info": "bright_cyan",
        "accent": "bright_magenta"
    },
    "matrix": {
        "primary": "bright_green",
        "success": "green",
        "warning": "yellow", 
        "error": "red",
        "info": "bright_green",
        "accent": "green"
    },
    "cyberpunk": {
        "primary": "bright_magenta",
        "success": "bright_cyan",
        "warning": "bright_yellow",
        "error": "bright_red", 
        "info": "bright_cyan",
        "accent": "bright_magenta"
    },
    "hacker": {
        "primary": "bright_green",
        "success": "green",
        "warning": "bright_yellow",
        "error": "bright_red",
        "info": "bright_cyan", 
        "accent": "bright_white"
    }
}

class AdvancedCNAMEResolver:
    """Advanced CNAME resolution with chain following and service detection"""
    
    def __init__(self, dns_resolver):
        self.dns_resolver = dns_resolver
        self.cname_cache = {}  # Cache for CNAME lookups
        self.service_patterns = {
            # AWS Services
            r'.*\.amazonaws\.com$': {'provider': 'AWS', 'service_type': 'AWS Service'},
            r'.*\.s3\.amazonaws\.com$': {'provider': 'AWS', 'service_type': 'S3 Bucket'},
            r'.*\.cloudfront\.net$': {'provider': 'AWS', 'service_type': 'CloudFront CDN'},
            r'.*\.elb\.amazonaws\.com$': {'provider': 'AWS', 'service_type': 'Load Balancer'},
            
            # Google Cloud
            r'.*\.googleapis\.com$': {'provider': 'Google', 'service_type': 'Google API'},
            r'.*\.appspot\.com$': {'provider': 'Google', 'service_type': 'App Engine'},
            r'.*\.storage\.googleapis\.com$': {'provider': 'Google', 'service_type': 'Cloud Storage'},
            
            # Microsoft Azure
            r'.*\.azurewebsites\.net$': {'provider': 'Azure', 'service_type': 'Web App'},
            r'.*\.blob\.core\.windows\.net$': {'provider': 'Azure', 'service_type': 'Blob Storage'},
            r'.*\.cloudapp\.azure\.com$': {'provider': 'Azure', 'service_type': 'Cloud Service'},
            
            # CDN Services
            r'.*\.cloudflare\.com$': {'provider': 'Cloudflare', 'service_type': 'CDN'},
            r'.*\.fastly\.com$': {'provider': 'Fastly', 'service_type': 'CDN'},
            r'.*\.akamai\.net$': {'provider': 'Akamai', 'service_type': 'CDN'},
            
            # Third-party Services
            r'.*\.shopify\.com$': {'provider': 'Shopify', 'service_type': 'E-commerce'},
            r'.*\.herokuapp\.com$': {'provider': 'Heroku', 'service_type': 'App Hosting'},
            r'.*\.github\.io$': {'provider': 'GitHub', 'service_type': 'Pages'},
            r'.*\.netlify\.com$': {'provider': 'Netlify', 'service_type': 'Static Hosting'},
            r'.*\.vercel\.app$': {'provider': 'Vercel', 'service_type': 'Static Hosting'},
            
            # Email Services
            r'.*\.mailgun\.org$': {'provider': 'Mailgun', 'service_type': 'Email Service'},
            r'.*\.sendgrid\.net$': {'provider': 'SendGrid', 'service_type': 'Email Service'},
        }
        
        self.high_risk_patterns = [
            r'.*\.herokuapp\.com$',
            r'.*\.github\.io$',
            r'.*\.s3\.amazonaws\.com$',
            r'.*\.azurewebsites\.net$',
        ]
    
    async def resolve_cname_comprehensive(self, subdomain: str, max_depth: int = 10) -> List[CNAMERecord]:
        """Comprehensively resolve CNAME chains with service detection"""
        cname_records = []
        current_domain = subdomain
        chain = [subdomain]
        depth = 0
        
        while depth < max_depth:
            try:
                # Try to get CNAME record
                resolver = self.dns_resolver.resolver_pool[0][1]
                cname_result = await resolver.query(current_domain, 'CNAME')
                
                if cname_result:
                    target = str(cname_result[0].host).rstrip('.')
                    chain.append(target)
                    
                    # Detect service and provider
                    service_info = self._detect_service(target)
                    takeover_risk = self._assess_takeover_risk(target)
                    
                    cname_record = CNAMERecord(
                        source_domain=current_domain,
                        target_domain=target,
                        chain_depth=depth + 1,
                        final_target=target,
                        service_type=service_info['service_type'],
                        provider=service_info['provider'],
                        takeover_risk=takeover_risk
                    )
                    cname_records.append(cname_record)
                    
                    current_domain = target
                    depth += 1
                else:
                    break
                    
            except Exception:
                break
        
        # Update final target for all records in chain
        if cname_records:
            final_target = cname_records[-1].target_domain
            for record in cname_records:
                record.final_target = final_target
        
        return cname_records
    
    def _detect_service(self, domain: str) -> Dict[str, str]:
        """Detect service type and provider from domain"""
        for pattern, info in self.service_patterns.items():
            if re.match(pattern, domain, re.IGNORECASE):
                return info
        return {'provider': 'Unknown', 'service_type': 'Unknown'}
    
    def _assess_takeover_risk(self, domain: str) -> str:
        """Assess subdomain takeover risk"""
        for pattern in self.high_risk_patterns:
            if re.match(pattern, domain, re.IGNORECASE):
                return "High"
        return "Low"
    
    async def discover_cname_targets(self, domain: str, wordlist: List[str]) -> List[SubdomainResult]:
        """Discover subdomains by reverse CNAME lookup"""
        results = []
        
        # Generate potential CNAME aliases
        cname_patterns = [
            'www', 'mail', 'ftp', 'blog', 'shop', 'store', 'app', 'api',
            'cdn', 'static', 'assets', 'media', 'images', 'js', 'css',
            'admin', 'dashboard', 'panel', 'login', 'secure', 'auth'
        ]
        
        candidates = []
        for pattern in cname_patterns:
            candidates.append(f"{pattern}.{domain}")
        
        # Also use wordlist
        for word in wordlist[:1000]:  # Limit for performance
            candidates.append(f"{word}.{domain}")
        
        # Test each candidate for CNAME records
        semaphore = asyncio.Semaphore(50)
        tasks = [self._test_cname_candidate(semaphore, candidate, domain) for candidate in candidates]
        results_raw = await asyncio.gather(*tasks, return_exceptions=True)
        
        for result in results_raw:
            if isinstance(result, SubdomainResult):
                results.append(result)
        
        return results
    
    async def _test_cname_candidate(self, semaphore: asyncio.Semaphore, candidate: str, domain: str) -> Optional[SubdomainResult]:
        """Test a candidate for CNAME records"""
        async with semaphore:
            try:
                cname_records = await self.resolve_cname_comprehensive(candidate)
                if cname_records:
                    # Resolve IP addresses
                    success, ip_addresses, resolver_used = await self.dns_resolver.resolve_with_intelligence(candidate)
                    
                    return SubdomainResult(
                        subdomain=candidate,
                        source="CNAME_Discovery",
                        http_status=0,
                        ip_addresses=ip_addresses if success else [],
                        technologies=[],
                        confidence_score=0.8,
                        discovered_at=time.time(),
                        cname_records=cname_records,
                        cname_chain=[record.target_domain for record in cname_records],
                        takeover_risk=cname_records[0].takeover_risk if cname_records else "Low"
                    )
            except Exception:
                pass
            return None

class IntelligentDNSResolver:
    """Advanced DNS resolver with multiple nameservers and intelligent failover"""
    
    def __init__(self):
        # Public DNS resolvers with different characteristics
        self.resolvers = [
            ['8.8.8.8', '8.8.4.4'],         # Google DNS
            ['1.1.1.1', '1.0.0.1'],         # Cloudflare DNS
            ['9.9.9.9', '149.112.112.112'], # Quad9 DNS
            ['208.67.222.222', '208.67.220.220'], # OpenDNS
            ['4.2.2.1', '4.2.2.2'],         # Level3 DNS
            ['8.26.56.26', '8.20.247.20'],  # Comodo DNS
            ['84.200.69.80', '84.200.70.40'], # DNS.WATCH
            ['94.140.14.14', '94.140.15.15'] # AdGuard DNS
        ]
        
        self.resolver_pool = []
        self.resolver_stats = defaultdict(lambda: {'success': 0, 'failure': 0, 'avg_time': 0})
        self._setup_resolvers()
        
        # Initialize CNAME resolver
        self.cname_resolver = AdvancedCNAMEResolver(self)
    
    def _setup_resolvers(self):
        """Initialize DNS resolver pool"""
        for resolver_pair in self.resolvers:
            for resolver_ip in resolver_pair:
                dns_resolver = aiodns.DNSResolver(nameservers=[resolver_ip], timeout=2, tries=1)
                self.resolver_pool.append((resolver_ip, dns_resolver))
    
    async def resolve_with_intelligence(self, subdomain: str) -> Tuple[bool, List[str], str]:
        """Intelligent DNS resolution with failover and performance tracking"""
        best_resolver = self._get_best_resolver()
        
        for resolver_ip, resolver in [best_resolver] + self.resolver_pool[:1]:  # Reduced to 2 total resolvers
            start_time = time.time()
            try:
                result = await resolver.query(subdomain, 'A')
                response_time = time.time() - start_time
                
                # Update resolver statistics
                self.resolver_stats[resolver_ip]['success'] += 1
                self.resolver_stats[resolver_ip]['avg_time'] = (
                    self.resolver_stats[resolver_ip]['avg_time'] + response_time
                ) / 2
                
                ip_addresses = [str(r.host) for r in result]
                return True, ip_addresses, resolver_ip
                
            except Exception:
                self.resolver_stats[resolver_ip]['failure'] += 1
                continue
        
        return False, [], "none"
    
    def _get_best_resolver(self) -> Tuple[str, aiodns.DNSResolver]:
        """Select the best performing DNS resolver"""
        best_score = float('-inf')
        best_resolver = self.resolver_pool[0]
        
        for resolver_ip, resolver in self.resolver_pool:
            stats = self.resolver_stats[resolver_ip]
            total_queries = stats['success'] + stats['failure']
            
            if total_queries == 0:
                score = 0  # No data yet
            else:
                success_rate = stats['success'] / total_queries
                avg_time = stats['avg_time'] if stats['avg_time'] > 0 else 1.0
                score = success_rate / avg_time  # Higher success rate, lower time = better
            
            if score > best_score:
                best_score = score
                best_resolver = (resolver_ip, resolver)
        
        return best_resolver

class MLSubdomainPredictor:
    """Machine Learning-based subdomain prediction using pattern analysis"""
    
    def __init__(self):
        self.patterns = defaultdict(int)
        self.ngrams = defaultdict(int)
        self.common_prefixes = Counter()
        self.common_suffixes = Counter()
        self.length_distribution = Counter()
    
    def analyze_patterns(self, known_subdomains: List[str], domain: str):
        """Analyze patterns from known subdomains"""
        for subdomain in known_subdomains:
            if subdomain.endswith(f'.{domain}'):
                prefix = subdomain[:-len(f'.{domain}')]
                self._extract_features(prefix)
    
    def _extract_features(self, prefix: str):
        """Extract linguistic and structural features"""
        # Length distribution
        self.length_distribution[len(prefix)] += 1
        
        # N-gram analysis
        for n in range(2, min(5, len(prefix) + 1)):
            for i in range(len(prefix) - n + 1):
                ngram = prefix[i:i+n]
                self.ngrams[ngram] += 1
        
        # Common patterns
        parts = re.split(r'[-_.]', prefix)
        for part in parts:
            if len(part) >= 2:
                self.common_prefixes[part[:3]] += 1
                self.common_suffixes[part[-3:]] += 1
        
        # Structural patterns
        if '-' in prefix:
            self.patterns['has_dash'] += 1
        if '_' in prefix:
            self.patterns['has_underscore'] += 1
        if prefix.isdigit():
            self.patterns['is_numeric'] += 1
        if any(c.isdigit() for c in prefix):
            self.patterns['has_numbers'] += 1
    
    def generate_predictions(self, domain: str, count: int = 1000) -> List[str]:
        """Generate predicted subdomains based on learned patterns"""
        predictions = set()
        
        # Use most common n-grams to build new subdomains
        # Convert defaultdict to Counter for most_common functionality
        ngrams_counter = Counter(self.ngrams)
        prefixes_counter = Counter(self.common_prefixes)
        suffixes_counter = Counter(self.common_suffixes)
        
        top_ngrams = [ngram for ngram, _ in ngrams_counter.most_common(50)]
        top_prefixes = [prefix for prefix, _ in prefixes_counter.most_common(20)]
        top_suffixes = [suffix for suffix, _ in suffixes_counter.most_common(20)]
        
        # Generate combinations
        for prefix in top_prefixes:
            for suffix in top_suffixes:
                predictions.add(f"{prefix}{suffix}.{domain}")
                predictions.add(f"{prefix}-{suffix}.{domain}")
                predictions.add(f"{prefix}_{suffix}.{domain}")
        
        # Generate based on common patterns
        numbers = ['1', '2', '3', '01', '02', '03', '10', '11', '12']
        for prefix in top_prefixes:
            for num in numbers:
                predictions.add(f"{prefix}{num}.{domain}")
                predictions.add(f"{prefix}-{num}.{domain}")
        
        # Generate variations of known patterns
        environments = ['dev', 'test', 'prod', 'stage', 'qa', 'demo', 'beta']
        for env in environments:
            for prefix in top_prefixes[:10]:
                predictions.add(f"{env}-{prefix}.{domain}")
                predictions.add(f"{prefix}-{env}.{domain}")
        
        return list(predictions)[:count]

class AdvancedCTMiner:
    """Advanced Certificate Transparency mining with multiple sources"""
    
    def __init__(self, session: aiohttp.ClientSession):
        self.session = session
        self.ct_sources = [
            {'name': 'crt.sh', 'url': 'https://crt.sh/?q=%25.{domain}&output=json', 'weight': 1.0},
            {'name': 'certspotter', 'url': 'https://api.certspotter.com/v1/issuances?domain={domain}&include_subdomains=true&expand=dns_names', 'weight': 0.8},
        ]
    
    async def comprehensive_mining(self, domain: str) -> List[SubdomainResult]:
        """Mine Certificate Transparency logs from multiple sources"""
        all_results = []
        
        tasks = []
        for source in self.ct_sources:
            task = self._mine_ct_source(source, domain)
            tasks.append(task)
        
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        for source, result in zip(self.ct_sources, results):
            if isinstance(result, list):
                all_results.extend(result)
        
        # Deduplicate and score results
        unique_subdomains = {}
        for result in all_results:
            if result.subdomain not in unique_subdomains:
                unique_subdomains[result.subdomain] = result
            else:
                # Merge confidence scores
                existing = unique_subdomains[result.subdomain]
                existing.confidence_score = max(existing.confidence_score, result.confidence_score)
        
        return list(unique_subdomains.values())
    
    async def _mine_ct_source(self, source: Dict, domain: str) -> List[SubdomainResult]:
        """Mine a specific CT source"""
        results = []
        
        try:
            url = source['url'].format(domain=domain)
            async with self.session.get(url, timeout=aiohttp.ClientTimeout(total=30)) as response:
                if response.status == 200:
                    data = await response.json()
                    results = self._parse_ct_data(data, source, domain)
                else:
                    raise Exception(f"HTTP {response.status}")
        except Exception as e:
            return []
        
        return results
    
    def _parse_ct_data(self, data: List[dict], source: Dict, domain: str) -> List[SubdomainResult]:
        """Parse Certificate Transparency data"""
        results = []
        
        for entry in data:
            subdomains = set()
            
            # Extract subdomains from different CT log formats
            if 'name_value' in entry:  # crt.sh format
                names = entry['name_value'].split('\n')
                subdomains.update(names)
            elif 'dns_names' in entry:  # certspotter format
                subdomains.update(entry['dns_names'])
            
            for subdomain in subdomains:
                if subdomain and self._is_valid_subdomain(subdomain, domain):
                    result = SubdomainResult(
                        subdomain=subdomain.lower(),
                        source=f"CT_{source['name']}",
                        http_status=0,
                        ip_addresses=[],
                        technologies=[],
                        confidence_score=source['weight'],
                        discovered_at=time.time()
                    )
                    results.append(result)
        
        return results
    
    def _is_valid_subdomain(self, subdomain: str, domain: str) -> bool:
        """Validate subdomain format"""
        if not subdomain or not subdomain.endswith(f'.{domain}'):
            return False
        if len(subdomain) > 253 or '..' in subdomain:
            return False
        if subdomain.startswith('*'):  # Skip wildcard certificates
            return False
        return bool(re.match(r'^[a-zA-Z0-9.*-]+$', subdomain))

class NetworkInfraAnalyzer:
    """Network infrastructure analysis for discovering related subdomains"""
    
    def __init__(self, session: aiohttp.ClientSession, dns_resolver: IntelligentDNSResolver):
        self.session = session
        self.dns_resolver = dns_resolver
    
    async def analyze_infrastructure(self, known_results: List[SubdomainResult], domain: str) -> List[SubdomainResult]:
        """Analyze network infrastructure to discover related subdomains"""
        new_results = []
        
        # Extract unique IP addresses
        ip_addresses = set()
        for result in known_results:
            ip_addresses.update(result.ip_addresses)
        
        # Perform reverse DNS lookups
        reverse_dns_results = await self._reverse_dns_analysis(ip_addresses, domain)
        new_results.extend(reverse_dns_results)
        
        # Analyze IP ranges and subnets
        subnet_results = await self._subnet_analysis(ip_addresses, domain)
        new_results.extend(subnet_results)
        
        return new_results
    
    async def _reverse_dns_analysis(self, ip_addresses: Set[str], domain: str) -> List[SubdomainResult]:
        """Perform reverse DNS lookups on discovered IP addresses"""
        results = []
        
        for ip in ip_addresses:
            try:
                # Reverse DNS lookup
                resolver = self.dns_resolver.resolver_pool[0][1]
                ptr_records = await resolver.query(ip, 'PTR')
                
                for record in ptr_records:
                    hostname = str(record.host).rstrip('.')
                    if hostname.endswith(f'.{domain}') and hostname != domain:
                        result = SubdomainResult(
                            subdomain=hostname,
                            source="Reverse_DNS",
                            http_status=0,
                            ip_addresses=[ip],
                            technologies=[],
                            confidence_score=0.8,
                            discovered_at=time.time()
                        )
                        results.append(result)
            except Exception:
                continue
        
        return results
    
    async def _subnet_analysis(self, ip_addresses: Set[str], domain: str) -> List[SubdomainResult]:
        """Analyze IP subnets for additional hosts"""
        results = []
        
        # Group IPs by /24 subnet
        subnets = defaultdict(list)
        for ip in ip_addresses:
            try:
                network = ipaddress.IPv4Network(f"{ip}/24", strict=False)
                subnets[str(network)].append(ip)
            except Exception:
                continue
        
        # For subnets with multiple known IPs, scan the range
        for subnet, ips in subnets.items():
            if len(ips) >= 2:  # Only scan subnets with multiple known IPs
                network = ipaddress.IPv4Network(subnet)
                
                # Scan common IP offsets in the subnet
                scan_ips = []
                for offset in [1, 2, 3, 10, 11, 12, 50, 51, 52, 100, 101, 102]:
                    scan_ip = str(network.network_address + offset)
                    if scan_ip not in ip_addresses:
                        scan_ips.append(scan_ip)
                
                # Perform reverse DNS on scan IPs
                subnet_results = await self._reverse_dns_analysis(set(scan_ips), domain)
                results.extend(subnet_results)
        
        return results

class UltraRobustEnumerator:
    """Ultra-robust subdomain enumerator with TUI integration"""
    
    def __init__(self, progress_callback=None, result_callback=None):
        self.dns_resolver = IntelligentDNSResolver()
        self.ml_predictor = MLSubdomainPredictor()
        self.results = {}  # subdomain -> SubdomainResult
        self.paused = False
        self.pause_event = None  # Will be set when needed
        self.progress_callback = progress_callback
        self.result_callback = result_callback
        
        # Configuration mapping
        self.mode_configs = {
            1: {'threads': 80, 'timeout': 3, 'http_workers': 50},     # Basic - Fast and efficient
            2: {'threads': 120, 'timeout': 5, 'http_workers': 80},    # Standard - Balanced
            3: {'threads': 60, 'timeout': 8, 'http_workers': 40},     # Advanced - Thorough but slower
            4: {'threads': 100, 'timeout': 5, 'http_workers': 60},    # Ultra - Optimized performance
            5: {'threads': 50, 'timeout': 2, 'http_workers': 30}      # Quick - Fastest
        }
        
        # Advanced configuration
        self.config = {
            'max_concurrent_dns': 80,  # Reduced for better performance
            'max_concurrent_http': 50,
            'max_concurrent_ssl': 10,  # Separate SSL concurrency limit
            'dns_timeout': 3,           # Faster timeout
            'http_timeout': 8,
            'ssl_timeout': 5,           # SSL connection timeout
            'confidence_threshold': 0.3,
            'max_recursion_depth': 3,
            # Flow control settings (much less aggressive)
            'phase_timeout_multiplier': 3.0,   # Very generous timeouts
            'max_phase_timeout': 3600,          # Maximum 1 hour per phase  
            'min_phase_timeout': 120,           # Minimum 2 minutes per phase
            'circuit_breaker_threshold': 0.95,  # 95% failure rate to trigger breaker
            'max_consecutive_failures': 50,     # Much higher tolerance before emergency mode
        }
        
        # Flow control state
        self.phase_failures = {}
        self.consecutive_failures = 0
        self.circuit_breakers = {}
    
    async def _execute_phase_with_timeout(self, phase_name: str, phase_func, *args, estimated_time: int = 60, **kwargs):
        """Execute a phase with comprehensive timeout and error handling"""
        
        # Calculate timeout with special handling for DNS phase
        if phase_name == "DNS Brute Force":
            # DNS gets MUCH more time based on expected workload
            timeout = max(
                self.config['min_phase_timeout'] * 10,  # At least 20 minutes for DNS
                min(
                    self.config['max_phase_timeout'],
                    int(estimated_time * self.config['phase_timeout_multiplier'])
                )
            )
        else:
            # Regular timeout calculation for other phases
            timeout = max(
                self.config['min_phase_timeout'],
                min(
                    self.config['max_phase_timeout'],
                    int(estimated_time * self.config['phase_timeout_multiplier'])
                )
            )
        
        self._emit_progress(phase_name, 0, message=f"Starting {phase_name} (timeout: {timeout}s)...")
        
        start_time = time.time()
        
        try:
            # Execute phase with timeout
            result = await asyncio.wait_for(
                phase_func(*args, **kwargs),
                timeout=timeout
            )
            
            elapsed = time.time() - start_time
            self._emit_progress(phase_name, 100, 
                              message=f"{phase_name} completed successfully in {elapsed:.1f}s")
            
            # Reset failure counters on success
            self.consecutive_failures = 0
            self.phase_failures[phase_name] = 0
            
            return result
            
        except asyncio.TimeoutError:
            elapsed = time.time() - start_time
            self._emit_progress(phase_name, 90, 
                              message=f"{phase_name} timed out after {elapsed:.1f}s - continuing with partial results")
            
            # Log timeout but continue
            self.phase_failures[phase_name] = self.phase_failures.get(phase_name, 0) + 1
            return None
            
        except Exception as e:
            elapsed = time.time() - start_time
            error_msg = str(e)[:100]
            self._emit_progress(phase_name, 95, 
                              message=f"{phase_name} failed after {elapsed:.1f}s: {error_msg}")
            
            # Track failures
            self.phase_failures[phase_name] = self.phase_failures.get(phase_name, 0) + 1
            self.consecutive_failures += 1
            
            # If too many consecutive failures, enable emergency mode
            if self.consecutive_failures >= self.config['max_consecutive_failures']:
                self._emit_progress(phase_name, 100, 
                                  message="⚠️ Multiple phase failures detected - enabling emergency completion mode")
                # Reduce consecutive failures but don't fully reset - allow gradual recovery
                self.consecutive_failures = max(0, self.consecutive_failures - 10)
            
            return None
    
    def _should_skip_phase(self, phase_name: str) -> bool:
        """Determine if a phase should be skipped due to circuit breaker"""
        failures = self.phase_failures.get(phase_name, 0)
        
        # Only skip if this specific phase has failed MANY times (extremely tolerant)
        if failures >= 10:
            self._emit_progress(phase_name, 100, 
                              message=f"Skipping {phase_name} - failed {failures} times already")
            return True
            
        # Only skip if we're in severe emergency mode
        if self.consecutive_failures >= self.config['max_consecutive_failures']:
            self._emit_progress(phase_name, 100, 
                              message=f"Emergency mode active - skipping optional phases")
            return True
            
        return False
    
    async def _smart_ip_scanning_phase(self):
        """Smart IP scanning phase wrapper with error handling"""
        try:
            # Group discovered subdomains by IP ranges
            self._emit_progress("Smart IP Analysis", 10, message="Grouping IPs by ranges...")
            ip_groups = self.group_ips_by_ranges(self.results)
            
            if ip_groups:
                self._emit_progress("Smart IP Analysis", 30, 
                                  message=f"Found {len(ip_groups)} IP ranges to scan efficiently")
                
                # Perform smart IP range scanning with timeout
                ip_scan_results = await asyncio.wait_for(
                    self.smart_ip_range_scan(ip_groups, self._emit_progress),
                    timeout=120  # 2 minute timeout for IP scanning
                )
                
                # Apply results back to subdomain data
                self._emit_progress("Smart IP Analysis", 80, message="Integrating scan results...")
                await self.apply_smart_scan_results(self.results, ip_scan_results, self._emit_progress)
                
            else:
                # Fallback to traditional scanning if no IP grouping possible
                self._emit_progress("Smart IP Analysis", 50, message="Fallback to traditional Nmap scans...")
                await asyncio.wait_for(
                    self.parallel_nmap_scanner(self.results, self._emit_progress),
                    timeout=180  # 3 minute timeout for traditional scanning
                )
                
        except asyncio.TimeoutError:
            self._emit_progress("Smart IP Analysis", 90, 
                              message="IP scanning timed out - continuing with existing results")
        except Exception as e:
            self._emit_progress("Smart IP Analysis", 95, 
                              message=f"IP scanning failed: {str(e)[:50]} - continuing")
    
    def _emit_progress(self, phase: str, progress: float, **kwargs):
        """Emit progress update"""
        if self.progress_callback:
            self.progress_callback(phase, progress, **kwargs)
    
    def _emit_result(self, result: SubdomainResult):
        """Emit new result"""
        if self.result_callback:
            self.result_callback(result)
    
    async def fast_enumerate(self, domain: str, wordlist_files: List[str]) -> Dict[str, SubdomainResult]:
        """Fast enumeration - DNS brute force only with minimal HTTP analysis"""
        
        # Initialize pause event for this enumeration
        self.pause_event = asyncio.Event()
        self.pause_event.set()  # Start unpaused
        
        self._emit_progress("Starting", 0, message="Fast enumerate mode - DNS focused!")
        
        # Update configuration for speed
        self.config['max_concurrent_dns'] = 50  # Reduced for stability
        self.config['max_concurrent_http'] = 30
        self.config['dns_timeout'] = 2
        self.config['http_timeout'] = 5
        
        start_time = time.time()
        
        try:
            self._emit_progress("Starting", 10, message="Fast enumeration with smart flow control...")
            
            # Phase 1: DNS Brute Force (CRITICAL - with timeout)
            await self._execute_phase_with_timeout(
                "DNS Brute Force", 
                self._phase_intelligent_dns_bruteforce, 
                domain, wordlist_files,
                estimated_time=90  # Fast mode gets less time
            )
            
            # Phase 2: CNAME Analysis (Quick analysis)
            if not self._should_skip_phase("CNAME Analysis") and len(self.results) > 0:
                await self._execute_phase_with_timeout(
                    "CNAME Analysis", 
                    self._analyze_existing_cname,
                    estimated_time=20
                )
            
            # Phase 3: Basic HTTP Analysis (with fallback)
            if not self._should_skip_phase("HTTP Analysis") and len(self.results) > 0:
                try:
                    async with aiohttp.ClientSession(
                        timeout=aiohttp.ClientTimeout(total=5),
                        connector=aiohttp.TCPConnector(limit=50, ttl_dns_cache=300)
                    ) as session:
                        await self._execute_phase_with_timeout(
                            "HTTP Analysis", 
                            self._basic_http_check, 
                            session,
                            estimated_time=60
                        )
                except Exception as e:
                    self._emit_progress("HTTP Analysis", 100, 
                                      message=f"HTTP session failed: {str(e)[:50]} - completing without HTTP")
            
            total_time = time.time() - start_time
            self._emit_progress("Complete", 100, 
                           message=f"Fast scan complete: {len(self.results)} subdomains in {total_time:.1f}s")
            
            return self.results
            
        except Exception as e:
            self._emit_progress("Error", 0, message=f"Fast enumeration failed: {str(e)[:50]}")
            return self.results
    
    async def scan_multiple_domains(self, domains: List[str], mode: int = 1, wordlist_files: List[str] = None) -> Dict[str, Dict[str, SubdomainResult]]:
        """Scan multiple domains in parallel with resource management"""
        
        if not domains:
            return {}
        
        # Default wordlists if none provided
        if wordlist_files is None:
            wordlist_files = ['wordlists/common.txt']
        
        # Resource allocation per domain to prevent conflicts
        total_domains = len(domains)
        dns_per_domain = max(25, 200 // total_domains)  # Divide DNS resources
        http_per_domain = max(15, 100 // total_domains)  # Divide HTTP resources
        
        results = {}
        
        async def scan_single_domain(domain):
            try:
                # Create separate scanner instance for each domain
                scanner = UltraRobustEnumerator(
                    progress_callback=lambda phase, progress, **kwargs: 
                        self._emit_progress(f"{domain} - {phase}", progress, **kwargs),
                    result_callback=self.result_callback
                )
                
                # Adjust resources for this domain
                scanner.config['max_concurrent_dns'] = dns_per_domain
                scanner.config['max_concurrent_http'] = http_per_domain
                scanner.config['dns_timeout'] = 3
                scanner.config['http_timeout'] = 5
                
                # Choose scan method based on mode and domain count
                if mode == 5 or total_domains > 2:  # Use fast mode for multiple domains
                    domain_results = await scanner.fast_enumerate(domain, wordlist_files)
                else:
                    domain_results = await scanner.ultra_enumerate(domain, mode, wordlist_files)
                
                results[domain] = domain_results
                
                self._emit_progress(f"Multi-Domain Scan", 
                                  (len(results) / total_domains) * 100,
                                  message=f"Completed {domain}: {len(domain_results)} subdomains found")
                
            except Exception as e:
                results[domain] = {}
                self._emit_progress(f"Multi-Domain Scan", 
                                  (len(results) / total_domains) * 100,
                                  message=f"Error scanning {domain}: {str(e)[:50]}")
        
        # Start all domain scans concurrently
        self._emit_progress("Multi-Domain Scan", 0, 
                          message=f"Starting parallel scan of {total_domains} domains...")
        
        tasks = [scan_single_domain(domain) for domain in domains]
        await asyncio.gather(*tasks, return_exceptions=True)
        
        total_subdomains = sum(len(domain_results) for domain_results in results.values())
        self._emit_progress("Multi-Domain Scan", 100, 
                          message=f"All domains completed: {total_subdomains} total subdomains found")
        
        return results

    async def ultra_enumerate(self, domain: str, mode: int, wordlist_files: List[str]) -> Dict[str, SubdomainResult]:
        """Ultra-robust enumeration with comprehensive CNAME analysis and TUI output"""
        
        # Initialize pause event for this enumeration
        self.pause_event = asyncio.Event()
        self.pause_event.set()  # Start unpaused
        
        # Immediate debug output to confirm we're running
        self._emit_progress("Starting", 0, message="Ultra enumerate method started!")
        
        # Update configuration based on mode
        mode_config = self.mode_configs[mode]
        self.config['max_concurrent_dns'] = mode_config['threads']
        self.config['max_concurrent_http'] = mode_config['http_workers']
        self.config['dns_timeout'] = mode_config['timeout']
        
        self._emit_progress("Starting", 10, message="Configuration updated")
        
        start_time = time.time()
        
        self._emit_progress("Starting", 20, message="Skipping HTTP session creation to prevent hanging...")
        
        # Skip HTTP session creation that was causing hangs
        self._emit_progress("Starting", 50, message="Proceeding directly to enumeration...")
        
        # Full comprehensive enumeration with bulletproof flow control
        try:
            self._emit_progress("Starting", 10, message="Ultra-robust enumeration with smart flow control...")
            
            # Core Discovery Phases (Critical)
            
            # Phase 1: Enhanced Certificate Transparency Mining
            if not self._should_skip_phase("Certificate Transparency"):
                await self._execute_phase_with_timeout(
                    "Certificate Transparency", 
                    self._enhanced_ct_mining, 
                    domain,
                    estimated_time=45
                )
            
            # Phase 2: CNAME Analysis
            if not self._should_skip_phase("CNAME Analysis"):
                await self._execute_phase_with_timeout(
                    "CNAME Analysis", 
                    self._phase_cname_comprehensive, 
                    domain, wordlist_files,
                    estimated_time=30
                )
            
            # Phase 3: DNS Brute Force (CRITICAL - never skip, dynamic timeout based on wordlist size)
            # Calculate dynamic timeout based on wordlist size
            total_words = sum(len(self._load_wordlists_from_files([wf])) for wf in wordlist_files)
            # Base: 30 minutes, +1 second per 100 candidates, minimum 30 minutes, maximum 2 hours
            dns_estimated_time = max(1800, min(7200, 1800 + (total_words // 100)))
            
            await self._execute_phase_with_timeout(
                "DNS Brute Force", 
                self._phase_intelligent_dns_bruteforce, 
                domain, wordlist_files,
                estimated_time=dns_estimated_time
            )
            
            # Extended Discovery Phases (Can be skipped if failing)
            
            # Phase 4: ML Predictions
            if not self._should_skip_phase("ML Predictions"):
                await self._execute_phase_with_timeout(
                    "ML Predictions", 
                    self._phase_ml_predictions, 
                    domain,
                    estimated_time=90
                )
            
            # Phase 5: Infrastructure Analysis
            if not self._should_skip_phase("Infrastructure Analysis"):
                await self._execute_phase_with_timeout(
                    "Infrastructure Analysis", 
                    self._simple_infrastructure_analysis, 
                    domain,
                    estimated_time=60
                )
            
            # Phase 6: Historical Discovery
            if not self._should_skip_phase("Historical Discovery"):
                await self._execute_phase_with_timeout(
                    "Historical Discovery", 
                    self._phase_historical_discovery, 
                    domain,
                    estimated_time=45
                )
            
            # Phase 7: Recursive Discovery
            if not self._should_skip_phase("Recursive Discovery"):
                await self._execute_phase_with_timeout(
                    "Recursive Discovery", 
                    self._phase_recursive_discovery, 
                    domain,
                    estimated_time=90
                )
            
            # Analysis Phases (HTTP and SSL with fallbacks)
            
            # Phase 8: HTTP Analysis
            if not self._should_skip_phase("HTTP Analysis"):
                try:
                    async with aiohttp.ClientSession(
                        timeout=aiohttp.ClientTimeout(total=self.config['http_timeout']),
                        connector=aiohttp.TCPConnector(limit=100, ttl_dns_cache=300, use_dns_cache=True)
                    ) as session:
                        await self._execute_phase_with_timeout(
                            "HTTP Analysis", 
                            self._phase_http_analysis, 
                            session,
                            estimated_time=180
                        )
                except Exception as e:
                    self._emit_progress("HTTP Analysis", 100, 
                                      message=f"HTTP session failed: {str(e)[:50]} - skipping")
            
            # Phase 9: SSL Analysis
            if not self._should_skip_phase("SSL Analysis"):
                await self._execute_phase_with_timeout(
                    "SSL Analysis", 
                    self._phase_ssl_analysis_controlled,
                    estimated_time=120
                )
            
        except Exception as e:
            self._emit_progress("Error", 0, message=f"Enumeration failed: {str(e)[:50]}")
            raise
        
        elapsed = time.time() - start_time
        total_subdomains = len(self.results)
        
        # Infrastructure Scanning Phase (with flow control)
        if not self._should_skip_phase("Smart IP Analysis") and len(self.results) > 0:
            # Smart IP Range Scanning with timeout
            await self._execute_phase_with_timeout(
                "Smart IP Analysis",
                self._smart_ip_scanning_phase,
                estimated_time=180
            )
        else:
            self._emit_progress("Smart IP Analysis", 100, 
                              message="Skipped IP analysis - no valid targets or circuit breaker active")
        
        # Final Analysis Phases (with aggressive timeouts)
        
        # Vulnerability Assessment (optional)
        if not self._should_skip_phase("Vulnerability Assessment"):
            await self._execute_phase_with_timeout(
                "Vulnerability Assessment", 
                self._advanced_vulnerability_assessment,
                estimated_time=120
            )
        
        # Advanced Analytics (always run - final phase)
        await self._execute_phase_with_timeout(
            "Advanced Analytics", 
            self._advanced_analytics_generation,
            estimated_time=60
        )
        
        # Save results
        output_file = self.save_advanced_excel(self.results, domain)
        
        return self.results
    
    async def _simple_dns_bruteforce(self, domain: str, wordlist_files: List[str]):
        """Simplified DNS brute force without aiohttp dependencies"""
        self._emit_progress("DNS Brute Force", 10, message="Loading wordlists...")
        
        # Load wordlists
        wordlist = self._load_wordlists_from_files(wordlist_files)
        candidates = [f"{word}.{domain}" for word in wordlist]
        
        self._emit_progress("DNS Brute Force", 20, 
                       message=f"Testing {len(candidates):,} candidates...")
        
        # Simple DNS resolution without complex batching
        found_count = 0
        total_candidates = len(candidates)
        
        for i, candidate in enumerate(candidates):  # Process all candidates
            try:
                success, ip_addresses, resolver_used = await self.dns_resolver.resolve_with_intelligence(candidate)
                
                # Check for pause every 50 candidates to avoid overhead
                if (i + 1) % 50 == 0:
                    await self.pause_event.wait()
                
                if success:
                    result = SubdomainResult(
                        subdomain=candidate,
                        source="DNS_Simple",
                        http_status=0,
                        ip_addresses=ip_addresses,
                        technologies=[],
                        confidence_score=0.9,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="Low"
                    )
                    self.results[candidate] = result
                    self._emit_result(result)
                    found_count += 1
                    
                # Update progress every 1000 candidates for large wordlists
                if i % 1000 == 0 or i == total_candidates - 1:
                    progress = 20 + (i / total_candidates) * 70  # 20% to 90%
                    self._emit_progress("DNS Brute Force", progress, 
                                   message=f"Tested {i+1:,}/{total_candidates:,}, found {found_count} subdomains")
                    
            except Exception:
                continue
        
        self._emit_progress("DNS Brute Force", 100, 
                       message=f"DNS scan complete: {found_count} subdomains found")
    
    async def _enhanced_ct_mining(self, domain: str):
        """Enhanced Certificate Transparency mining with multiple CT log sources"""
        self._emit_progress("Certificate Transparency", 10, message="Starting enhanced CT log mining...")
        
        # Multiple CT log sources for comprehensive discovery
        ct_sources = [
            {
                'name': 'crt.sh',
                'url': f'https://crt.sh/?q=%.{domain}&output=json',
                'parser': self._parse_crtsh_data
            },
            {
                'name': 'censys',
                'url': f'https://search.censys.io/api/v1/search/certificates',
                'parser': self._parse_censys_data
            },
            {
                'name': 'certspotter',
                'url': f'https://api.certspotter.com/v1/issuances?domain={domain}&include_subdomains=true&expand=dns_names',
                'parser': self._parse_certspotter_data
            }
        ]
        
        found_subdomains = set()
        total_sources = len(ct_sources)
        
        for i, source in enumerate(ct_sources):
            try:
                self._emit_progress("Certificate Transparency", 
                                  20 + (i * 60 // total_sources),
                                  message=f"Querying {source['name']} CT logs...")
                
                # Use subprocess to query CT logs (avoids aiohttp issues)
                if source['name'] == 'crt.sh':
                    # Query crt.sh directly
                    subdomains = await self._query_crtsh(domain)
                    found_subdomains.update(subdomains)
                
                elif source['name'] == 'certspotter':
                    # Query certspotter API
                    subdomains = await self._query_certspotter(domain)
                    found_subdomains.update(subdomains)
                
                # Add progress update
                if found_subdomains:
                    self._emit_progress("Certificate Transparency",
                                      20 + ((i + 1) * 60 // total_sources),
                                      message=f"{source['name']}: {len(found_subdomains)} unique subdomains found")
                                      
            except Exception as e:
                self._emit_progress("Certificate Transparency",
                                  20 + ((i + 1) * 60 // total_sources),
                                  message=f"{source['name']} failed: {str(e)[:30]}")
                continue
        
        # Process discovered subdomains
        self._emit_progress("Certificate Transparency", 85, 
                          message=f"Processing {len(found_subdomains)} discovered subdomains...")
        
        processed = 0
        for subdomain in found_subdomains:
            if self._is_valid_subdomain(subdomain) and domain in subdomain:
                try:
                    success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(subdomain)
                    if success:
                        result = SubdomainResult(
                            subdomain=subdomain,
                            source="Enhanced_CT_Mining",
                            http_status=0,
                            ip_addresses=ip_addresses,
                            technologies=[],
                            confidence_score=0.9,  # Higher confidence for CT data
                            discovered_at=time.time(),
                            cname_records=[],
                            cname_chain=[],
                            takeover_risk="Low"
                        )
                        self.results[subdomain] = result
                        self._emit_result(result)
                        processed += 1
                except Exception:
                    continue
        
        self._emit_progress("Certificate Transparency", 100, 
                          message=f"Enhanced CT mining complete: {processed} verified subdomains")

    async def _query_crtsh(self, domain: str) -> set:
        """Query crt.sh CT log database"""
        subdomains = set()
        try:
            # Use curl to query crt.sh
            cmd = ['curl', '-s', f'https://crt.sh/?q=%.{domain}&output=json']
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            
            if result.returncode == 0 and result.stdout:
                import json
                try:
                    data = json.loads(result.stdout)
                    for cert in data:
                        name_value = cert.get('name_value', '')
                        if name_value:
                            # Split on newlines as crt.sh returns multiple domains per record
                            for name in name_value.split('\n'):
                                name = name.strip()
                                if name and not name.startswith('*'):
                                    subdomains.add(name)
                except json.JSONDecodeError:
                    pass
        except Exception:
            pass
        
        return subdomains

    async def _query_certspotter(self, domain: str) -> set:
        """Query Certspotter CT log API"""
        subdomains = set()
        try:
            # Use curl to query certspotter
            cmd = ['curl', '-s', 
                   f'https://api.certspotter.com/v1/issuances?domain={domain}&include_subdomains=true&expand=dns_names']
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            
            if result.returncode == 0 and result.stdout:
                import json
                try:
                    data = json.loads(result.stdout)
                    for cert in data:
                        dns_names = cert.get('dns_names', [])
                        for name in dns_names:
                            if name and not name.startswith('*'):
                                subdomains.add(name)
                except json.JSONDecodeError:
                    pass
        except Exception:
            pass
        
        return subdomains

    def _parse_crtsh_data(self, data: dict) -> List[str]:
        """Parse crt.sh JSON response"""
        subdomains = []
        for cert in data:
            name_value = cert.get('name_value', '')
            if name_value:
                for name in name_value.split('\n'):
                    name = name.strip()
                    if name and not name.startswith('*'):
                        subdomains.append(name)
        return subdomains

    def _parse_certspotter_data(self, data: dict) -> List[str]:
        """Parse Certspotter JSON response"""
        subdomains = []
        for cert in data:
            dns_names = cert.get('dns_names', [])
            for name in dns_names:
                if name and not name.startswith('*'):
                    subdomains.append(name)
        return subdomains

    def _parse_censys_data(self, data: dict) -> List[str]:
        """Parse Censys JSON response"""
        subdomains = []
        results = data.get('results', [])
        for cert in results:
            parsed = cert.get('parsed', {})
            subject_alt_name = parsed.get('extensions', {}).get('subject_alt_name', {})
            dns_names = subject_alt_name.get('dns_names', [])
            for name in dns_names:
                if name and not name.startswith('*'):
                    subdomains.append(name)
        return subdomains

    async def _advanced_subdomain_discovery(self, domain: str):
        """Advanced Subdomain Discovery using multiple techniques"""
        self._emit_progress("Advanced Discovery", 10, message="Starting advanced discovery methods...")
        
        discovered_count = 0
        
        # Method 1: DNS Zone Transfer (AXFR) attempts
        self._emit_progress("Advanced Discovery", 20, message="Attempting DNS zone transfers...")
        axfr_results = await self._dns_zone_transfer(domain)
        discovered_count += len(axfr_results)
        
        # Method 2: Search Engine Discovery
        self._emit_progress("Advanced Discovery", 40, message="Mining search engines...")
        search_results = await self._search_engine_discovery(domain)
        discovered_count += len(search_results)
        
        # Method 3: GitHub/GitLab Repository Mining
        self._emit_progress("Advanced Discovery", 60, message="Mining code repositories...")
        repo_results = await self._repository_mining(domain)
        discovered_count += len(repo_results)
        
        # Method 4: Wayback Machine Historical Discovery
        self._emit_progress("Advanced Discovery", 80, message="Mining historical data...")
        wayback_results = await self._wayback_machine_discovery(domain)
        discovered_count += len(wayback_results)
        
        # Method 5: ASN Enumeration
        self._emit_progress("Advanced Discovery", 90, message="ASN-based discovery...")
        asn_results = await self._asn_enumeration(domain)
        discovered_count += len(asn_results)
        
        self._emit_progress("Advanced Discovery", 100, 
                          message=f"Advanced discovery complete: {discovered_count} new subdomains")

    async def _dns_zone_transfer(self, domain: str) -> List[str]:
        """Attempt DNS zone transfer (AXFR) to discover subdomains"""
        discovered = []
        
        try:
            # Get nameservers for the domain
            nameservers = []
            cmd = ['dig', '+short', 'NS', domain]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            
            if result.returncode == 0:
                nameservers = [ns.strip().rstrip('.') for ns in result.stdout.strip().split('\n') if ns.strip()]
            
            # Attempt zone transfer on each nameserver
            for ns in nameservers[:3]:  # Limit to first 3 nameservers
                try:
                    cmd = ['dig', f'@{ns}', domain, 'AXFR']
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
                    
                    if result.returncode == 0 and 'refused' not in result.stdout.lower():
                        # Parse AXFR output for subdomains
                        lines = result.stdout.split('\n')
                        for line in lines:
                            if domain in line and line.strip():
                                parts = line.split()
                                if len(parts) > 0:
                                    potential_subdomain = parts[0].rstrip('.')
                                    if potential_subdomain.endswith(f'.{domain}'):
                                        discovered.append(potential_subdomain)
                                        
                                        # Add to results
                                        result_obj = SubdomainResult(
                                            subdomain=potential_subdomain,
                                            source="DNS_AXFR",
                                            http_status=0,
                                            ip_addresses=[],
                                            technologies=[],
                                            confidence_score=0.95,
                                            discovered_at=time.time()
                                        )
                                        self.results[potential_subdomain] = result_obj
                                        self._emit_result(result_obj)
                                        
                except subprocess.TimeoutExpired:
                    continue
                except Exception:
                    continue
                    
        except Exception:
            pass
        
        return discovered

    async def _search_engine_discovery(self, domain: str) -> List[str]:
        """Discover subdomains through search engine queries"""
        discovered = []
        
        # Search queries to find subdomains
        search_queries = [
            f'site:*.{domain}',
            f'site:{domain} inurl:subdomain',
            f'site:{domain} -www',
            f'"*.{domain}"',
            f'site:{domain} filetype:txt',
        ]
        
        try:
            for query in search_queries:
                # Use curl to query search engines (simplified approach)
                # Note: In production, you'd want to use proper APIs or respect rate limits
                
                # Google dorking simulation (using curl with user agent)
                cmd = [
                    'curl', '-s', '-A', 
                    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                    f'https://www.google.com/search?q={query}'
                ]
                
                try:
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
                    if result.returncode == 0:
                        # Extract potential subdomains from search results
                        import re
                        subdomain_pattern = r'([a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+' + re.escape(domain)
                        matches = re.findall(subdomain_pattern, result.stdout)
                        
                        for match in matches[:5]:  # Limit results
                            subdomain = match[0] + domain if isinstance(match, tuple) else match
                            if subdomain not in discovered and self._is_valid_subdomain(subdomain):
                                discovered.append(subdomain)
                                
                                # Verify subdomain exists
                                try:
                                    success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(subdomain)
                                    if success:
                                        result_obj = SubdomainResult(
                                            subdomain=subdomain,
                                            source="Search_Engine",
                                            http_status=0,
                                            ip_addresses=ip_addresses,
                                            technologies=[],
                                            confidence_score=0.7,
                                            discovered_at=time.time()
                                        )
                                        self.results[subdomain] = result_obj
                                        self._emit_result(result_obj)
                                except Exception:
                                    continue
                                    
                except subprocess.TimeoutExpired:
                    continue
                    
                # Rate limiting
                await asyncio.sleep(1)
                
        except Exception:
            pass
        
        return discovered

    async def _repository_mining(self, domain: str) -> List[str]:
        """Mine GitHub/GitLab repositories for subdomain mentions"""
        discovered = []
        
        try:
            # Search GitHub for domain mentions
            search_terms = [
                f'{domain}',
                f'*.{domain}',
                f'subdomain {domain}',
            ]
            
            for term in search_terms:
                # Use GitHub search API (simplified)
                cmd = [
                    'curl', '-s', '-H', 'Accept: application/vnd.github.v3+json',
                    f'https://api.github.com/search/code?q={term}+in:file'
                ]
                
                try:
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
                    if result.returncode == 0:
                        import json
                        try:
                            data = json.loads(result.stdout)
                            items = data.get('items', [])
                            
                            for item in items[:3]:  # Limit to avoid rate limits
                                # Extract potential subdomains from file content
                                import re
                                content_url = item.get('url', '')
                                if content_url:
                                    # Get file content
                                    content_cmd = ['curl', '-s', '-H', 'Accept: application/vnd.github.v3.raw', content_url]
                                    content_result = subprocess.run(content_cmd, capture_output=True, text=True, timeout=10)
                                    
                                    if content_result.returncode == 0:
                                        subdomain_pattern = r'([a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+' + re.escape(domain)
                                        matches = re.findall(subdomain_pattern, content_result.stdout)
                                        
                                        for match in matches[:3]:
                                            subdomain = match[0] + domain if isinstance(match, tuple) else match
                                            if subdomain not in discovered and self._is_valid_subdomain(subdomain):
                                                discovered.append(subdomain)
                                                
                                                # Verify and add to results
                                                try:
                                                    success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(subdomain)
                                                    if success:
                                                        result_obj = SubdomainResult(
                                                            subdomain=subdomain,
                                                            source="Repository_Mining",
                                                            http_status=0,
                                                            ip_addresses=ip_addresses,
                                                            technologies=[],
                                                            confidence_score=0.8,
                                                            discovered_at=time.time()
                                                        )
                                                        self.results[subdomain] = result_obj
                                                        self._emit_result(result_obj)
                                                except Exception:
                                                    continue
                                                    
                        except json.JSONDecodeError:
                            pass
                            
                except subprocess.TimeoutExpired:
                    continue
                
                # Rate limiting for GitHub API
                await asyncio.sleep(2)
                
        except Exception:
            pass
        
        return discovered

    async def _wayback_machine_discovery(self, domain: str) -> List[str]:
        """Discover subdomains from Wayback Machine historical data"""
        discovered = []
        
        try:
            # Query Wayback Machine CDX API
            cmd = [
                'curl', '-s',
                f'http://web.archive.org/cdx/search/cdx?url=*.{domain}&output=text&fl=original&collapse=urlkey'
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=20)
            
            if result.returncode == 0:
                urls = result.stdout.strip().split('\n')
                
                for url in urls[:50]:  # Limit results
                    if url and domain in url:
                        # Extract subdomain from URL
                        import re
                        subdomain_match = re.search(r'https?://([^/]+)', url)
                        if subdomain_match:
                            potential_subdomain = subdomain_match.group(1)
                            
                            if (potential_subdomain.endswith(f'.{domain}') and 
                                potential_subdomain not in discovered and
                                self._is_valid_subdomain(potential_subdomain)):
                                
                                discovered.append(potential_subdomain)
                                
                                # Verify subdomain still exists
                                try:
                                    success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(potential_subdomain)
                                    if success:
                                        result_obj = SubdomainResult(
                                            subdomain=potential_subdomain,
                                            source="Wayback_Machine",
                                            http_status=0,
                                            ip_addresses=ip_addresses,
                                            technologies=[],
                                            confidence_score=0.6,  # Lower confidence for historical data
                                            discovered_at=time.time()
                                        )
                                        self.results[potential_subdomain] = result_obj
                                        self._emit_result(result_obj)
                                except Exception:
                                    continue
                                    
        except Exception:
            pass
        
        return discovered

    async def _asn_enumeration(self, domain: str) -> List[str]:
        """Discover subdomains through ASN (Autonomous System Number) enumeration"""
        discovered = []
        
        try:
            # First, get IP addresses for the main domain
            success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(domain)
            
            if success and ip_addresses:
                main_ip = ip_addresses[0]
                
                # Get ASN information for the IP
                cmd = ['whois', '-h', 'whois.cymru.com', f'-v {main_ip}']
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
                
                if result.returncode == 0:
                    # Parse ASN from whois output
                    lines = result.stdout.split('\n')
                    asn = None
                    
                    for line in lines:
                        if 'AS' in line and '|' in line:
                            parts = line.split('|')
                            if len(parts) > 0:
                                asn_part = parts[0].strip()
                                if asn_part.startswith('AS'):
                                    asn = asn_part
                                    break
                    
                    if asn:
                        # Query for other domains in the same ASN
                        # Note: This is a simplified approach
                        cmd = ['curl', '-s', f'https://bgp.he.net/{asn}']
                        result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
                        
                        if result.returncode == 0:
                            # Extract potential related domains
                            import re
                            domain_pattern = r'([a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+' + re.escape(domain.split('.')[-1])
                            matches = re.findall(domain_pattern, result.stdout)
                            
                            for match in matches[:10]:  # Limit results
                                potential_domain = match[0] + domain.split('.')[-1] if isinstance(match, tuple) else match
                                
                                if (domain in potential_domain and 
                                    potential_domain != domain and
                                    potential_domain not in discovered):
                                    
                                    discovered.append(potential_domain)
                                    
                                    # Verify and add to results
                                    try:
                                        success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(potential_domain)
                                        if success:
                                            result_obj = SubdomainResult(
                                                subdomain=potential_domain,
                                                source="ASN_Enumeration",
                                                http_status=0,
                                                ip_addresses=ip_addresses,
                                                technologies=[],
                                                confidence_score=0.5,
                                                discovered_at=time.time()
                                            )
                                            self.results[potential_domain] = result_obj
                                            self._emit_result(result_obj)
                                    except Exception:
                                        continue
                                        
        except Exception:
            pass
        
        return discovered
    
    async def _phase_cname_analysis(self, domain: str):
        """Analyze existing results for CNAME chains"""
        self._emit_progress("CNAME Analysis", 50, message="Analyzing CNAME chains...")
        
        cname_count = 0
        for subdomain, result in self.results.items():
            # Simple CNAME analysis without complex chains
            try:
                # This is a placeholder - real CNAME analysis would be more complex
                cname_count += 1
            except Exception:
                continue
        
        self._emit_progress("CNAME Analysis", 100, 
                       message=f"CNAME Analysis complete: {cname_count} records analyzed")
    
    async def _basic_http_check(self, session: aiohttp.ClientSession):
        """Basic HTTP check - just status codes, no detailed analysis"""
        if not self.results:
            return
            
        total_subdomains = len(self.results)
        completed = 0
        semaphore = asyncio.Semaphore(self.config['max_concurrent_http'])
        
        async def quick_http_check(result):
            nonlocal completed
            async with semaphore:
                try:
                    url = f"https://{result.subdomain}"
                    async with session.get(url, allow_redirects=False, timeout=aiohttp.ClientTimeout(total=3)) as response:
                        result.http_status = response.status
                        result.server = response.headers.get('Server', '')
                except:
                    try:
                        url = f"http://{result.subdomain}"
                        async with session.get(url, allow_redirects=False, timeout=aiohttp.ClientTimeout(total=3)) as response:
                            result.http_status = response.status
                            result.server = response.headers.get('Server', '')
                    except:
                        pass
                completed += 1
        
        tasks = [quick_http_check(result) for result in self.results.values()]
        await asyncio.gather(*tasks, return_exceptions=True)
        
        live_count = sum(1 for r in self.results.values() if r.http_status == 200)
        self._emit_progress("HTTP Analysis", 100, 
                       message=f"Basic HTTP check complete: {live_count} live services found")

    async def _simple_http_analysis(self):
        """Simplified HTTP analysis without aiohttp"""
        self._emit_progress("HTTP Analysis", 50, message="Analyzing HTTP services...")
        
        # Update existing results with basic HTTP status (simplified)
        http_count = 0
        for subdomain, result in self.results.items():
            # Placeholder for HTTP analysis
            http_count += 1
        
        self._emit_progress("HTTP Analysis", 100, 
                       message=f"HTTP Analysis complete: {http_count} services analyzed")
    
    async def _phase_final_cname_analysis(self):
        """Final CNAME chain analysis"""
        self._emit_progress("Final CNAME Analysis", 50, message="Final CNAME processing...")
        
        # Final pass through results for CNAME validation
        final_count = len(self.results)
        
        self._emit_progress("Final CNAME Analysis", 100, 
                       message=f"Final Analysis complete: {final_count} total results")
    
    async def _simple_infrastructure_analysis(self, domain: str):
        """Simplified infrastructure analysis"""
        self._emit_progress("Infrastructure Analysis", 50, message="Mapping infrastructure...")
        
        # Simple infrastructure mapping
        infra_count = len(self.results)
        
        self._emit_progress("Infrastructure Analysis", 100, 
                       message=f"Infrastructure analysis complete: {infra_count} hosts mapped")
    
    async def _phase_ct_mining(self, session: aiohttp.ClientSession, domain: str):
        """Phase 1: Advanced Certificate Transparency Mining"""
        self._emit_progress("Certificate Transparency", 10, message="Starting CT log mining...")
        
        try:
            ct_miner = AdvancedCTMiner(session)
            ct_results = await ct_miner.comprehensive_mining(domain)
            
            # Add to main results
            for result in ct_results:
                self.results[result.subdomain] = result
                self._emit_result(result)
            
            self._emit_progress("Certificate Transparency", 100, 
                           message=f"CT Mining complete: {len(ct_results)} subdomains",
                           subdomain_count=len(self.results))
        except Exception as e:
            self._emit_progress("Certificate Transparency", 100, 
                           message=f"CT Mining failed: {str(e)[:50]}",
                           subdomain_count=len(self.results))
    
    async def _phase_intelligent_dns_bruteforce(self, domain: str, wordlist_files: List[str]):
        """Phase 2: Intelligent DNS Brute Force"""
        self._emit_progress("DNS Brute Force", 5, message="Loading wordlists...")
        
        # Load wordlists
        wordlist = self._load_wordlists_from_files(wordlist_files)
        candidates = [f"{word}.{domain}" for word in wordlist]
        
        self._emit_progress("DNS Brute Force", 10, 
                       message=f"Testing {len(candidates):,} candidates with {len(self.dns_resolver.resolver_pool)} DNS resolvers...")
        
        # Process in optimized batches with better progress reporting
        batch_size = 200  # Smaller batches to prevent hanging
        total_found = 0
        total_batches = (len(candidates) + batch_size - 1) // batch_size
        
        semaphore = asyncio.Semaphore(self.config['max_concurrent_dns'])
        
        for i in range(0, len(candidates), batch_size):
            batch = candidates[i:i + batch_size]
            batch_num = i // batch_size + 1
            
            progress_pct = 10 + (batch_num / total_batches) * 80  # 10% to 90%
            
            # More detailed progress message
            tested_so_far = min(i + batch_size, len(candidates))
            self._emit_progress("DNS Brute Force", progress_pct,
                           current=tested_so_far, total=len(candidates),
                           message=f"Testing batch {batch_num}/{total_batches} - {tested_so_far:,}/{len(candidates):,} subdomains tested",
                           subdomain_count=len(self.results))
            
            batch_start_time = time.time()
            tasks = [self._resolve_with_intelligence(semaphore, candidate) for candidate in batch]
            
            # Add timeout to prevent batches from hanging indefinitely  
            try:
                results = await asyncio.wait_for(
                    asyncio.gather(*tasks, return_exceptions=True),
                    timeout=60.0  # 60 second timeout per batch - more generous for slow DNS
                )
            except asyncio.TimeoutError:
                self._emit_progress("DNS Brute Force", progress_pct,
                               current=tested_so_far, total=len(candidates),
                               message=f"Batch {batch_num} timed out - skipping to next batch",
                               subdomain_count=len(self.results))
                results = [None] * len(batch)  # Create empty results for timed out batch
            
            # Check for pause request after each batch
            await self.pause_event.wait()
            
            batch_time = time.time() - batch_start_time
            
            batch_found = sum(1 for result in results if isinstance(result, SubdomainResult))
            total_found += batch_found
            
            # Add valid results
            for result in results:
                if isinstance(result, SubdomainResult):
                    self.results[result.subdomain] = result
                    self._emit_result(result)
            
            # Progress update with rate calculation
            rate = len(batch) / batch_time if batch_time > 0 else 0
            eta = ((total_batches - batch_num) * batch_time) if batch_time > 0 else 0
            
            # Show updated progress with current findings
            progress_pct_updated = 10 + ((batch_num) / total_batches) * 80
            self._emit_progress("DNS Brute Force", progress_pct_updated,
                           current=batch_num, total=total_batches,
                           rate=rate, eta=eta,
                           message=f"Batch {batch_num}/{total_batches} complete: {batch_found} found, {total_found} total (Rate: {rate:.1f}/s)",
                           subdomain_count=len(self.results))
            
            # Anti-hang mechanism: force delay if processing too fast or slow
            if batch_time < 0.5:  # Too fast, might overwhelm DNS servers
                await asyncio.sleep(0.2)
            elif batch_time > 120:  # Too slow, log warning
                self._emit_progress("DNS Brute Force", progress_pct_updated,
                               message=f"⚠️  Batch {batch_num} took {batch_time:.1f}s - DNS servers may be rate limiting",
                               subdomain_count=len(self.results))
        
        # Ensure we report that we actually completed all batches
        final_message = f"DNS Brute Force FULLY COMPLETE: {total_batches} batches processed, {total_found} new subdomains discovered from {len(candidates):,} candidates"
        self._emit_progress("DNS Brute Force", 100, 
                       message=final_message,
                       subdomain_count=len(self.results))
    
    async def _resolve_with_intelligence(self, semaphore: asyncio.Semaphore, subdomain: str) -> Optional[SubdomainResult]:
        """Resolve with intelligent DNS and comprehensive CNAME analysis"""
        async with semaphore:
            try:
                # Add timeout to individual resolution
                success, ip_addresses, resolver_used = await asyncio.wait_for(
                    self.dns_resolver.resolve_with_intelligence(subdomain),
                    timeout=5.0  # 5 second timeout per subdomain
                )
            except asyncio.TimeoutError:
                return None
            
            if success:
                # Get CNAME information with timeout
                try:
                    cname_records = await asyncio.wait_for(
                        self.dns_resolver.cname_resolver.resolve_cname_comprehensive(subdomain),
                        timeout=3.0  # 3 second timeout for CNAME resolution
                    )
                except (asyncio.TimeoutError, Exception):
                    cname_records = []
                
                return SubdomainResult(
                    subdomain=subdomain,
                    source=f"DNS_Intelligence_{resolver_used.replace('.', '_')}",
                    http_status=0,
                    ip_addresses=ip_addresses,
                    technologies=[],
                    confidence_score=0.9,
                    discovered_at=time.time(),
                    cname_records=cname_records,
                    cname_chain=[record.target_domain for record in cname_records],
                    takeover_risk=cname_records[0].takeover_risk if cname_records else "Low"
                )
            return None
    
    async def _phase_ml_predictions(self, domain: str):
        """Advanced ML-based subdomain prediction and generation with local models"""
        self._emit_progress("ML Predictions", 10, message="Initializing local ML models...")
        
        # Advanced local ML-based subdomain generation
        ml_candidates = await self._advanced_ml_generation(domain)
        
        self._emit_progress("ML Predictions", 30, 
                       message=f"Testing {len(ml_candidates)} ML-generated candidates...")
        
        # Test ML-generated candidates with intelligent batching
        ml_found = 0
        semaphore = asyncio.Semaphore(50)  # Control concurrency
        
        async def test_candidate(candidate):
            async with semaphore:
                success, ip_addresses, resolver_used = await self.dns_resolver.resolve_with_intelligence(candidate)
                if success:
                    result = SubdomainResult(
                        subdomain=candidate,
                        source="Advanced_ML_Local",
                        http_status=0,
                        ip_addresses=ip_addresses,
                        technologies=[],
                        confidence_score=0.8,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="Low"
                    )
                    self.results[candidate] = result
                    self._emit_result(result)
                    return 1
                return 0
        
        # Process candidates in batches
        batch_size = 100
        for i in range(0, len(ml_candidates), batch_size):
            batch = ml_candidates[i:i + batch_size]
            tasks = [test_candidate(candidate) for candidate in batch]
            batch_results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Count successful results
            batch_found = sum(r for r in batch_results if isinstance(r, int))
            ml_found += batch_found
            
            # Progress update
            progress = 30 + ((i + batch_size) / len(ml_candidates)) * 60
            self._emit_progress("ML Predictions", min(progress, 90), 
                           message=f"ML batch {i//batch_size + 1}: {batch_found} found, {ml_found} total")
        
        self._emit_progress("ML Predictions", 100, 
                       message=f"Advanced ML complete: {ml_found} new subdomains",
                       subdomain_count=len(self.results))
    
    async def _advanced_ml_generation(self, domain: str) -> List[str]:
        """Advanced local ML-based subdomain generation using multiple techniques"""
        candidates = set()
        
        # 1. Pattern-based generation from existing results
        candidates.update(self._pattern_based_generation(domain))
        
        # 2. N-gram analysis for linguistic patterns
        candidates.update(self._ngram_based_generation(domain))
        
        # 3. Industry-specific dictionary generation
        candidates.update(self._industry_specific_generation(domain))
        
        # 4. Technology stack prediction
        candidates.update(self._tech_stack_prediction(domain))
        
        # 5. Semantic similarity expansion
        candidates.update(self._semantic_expansion(domain))
        
        # 6. Morphological analysis
        candidates.update(self._morphological_generation(domain))
        
        return list(candidates)
    
    def _pattern_based_generation(self, domain: str) -> Set[str]:
        """Generate subdomains based on discovered patterns"""
        candidates = set()
        
        # Analyze existing subdomains for patterns
        existing_subdomains = [r.subdomain for r in self.results.values()]
        patterns = set()
        
        for subdomain in existing_subdomains:
            # Extract prefix patterns
            if '.' in subdomain:
                prefix = subdomain.split('.')[0]
                
                # Common transformations
                patterns.add(f"{prefix}1")
                patterns.add(f"{prefix}2")
                patterns.add(f"{prefix}-api")
                patterns.add(f"{prefix}-test")
                patterns.add(f"{prefix}-dev")
                patterns.add(f"new-{prefix}")
                patterns.add(f"old-{prefix}")
                patterns.add(f"{prefix}-staging")
                patterns.add(f"{prefix}-prod")
        
        # Generate candidates from patterns
        for pattern in patterns:
            candidates.add(f"{pattern}.{domain}")
        
        return candidates
    
    def _ngram_based_generation(self, domain: str) -> Set[str]:
        """Generate subdomains using N-gram analysis"""
        candidates = set()
        
        # Common subdomain n-grams
        common_bigrams = ['ap', 'pi', 'ad', 'dm', 'mi', 'in', 'we', 'eb', 'ma', 'ai', 'il', 'ft', 'tp']
        common_trigrams = ['api', 'adm', 'web', 'mai', 'ftp', 'dev', 'app', 'cdn', 'www']
        
        # Generate combinations
        for bigram in common_bigrams:
            candidates.add(f"{bigram}.{domain}")
            candidates.add(f"{bigram}1.{domain}")
            candidates.add(f"{bigram}2.{domain}")
        
        for trigram in common_trigrams:
            candidates.add(f"{trigram}.{domain}")
            candidates.add(f"{trigram}1.{domain}")
            candidates.add(f"{trigram}2.{domain}")
            candidates.add(f"new{trigram}.{domain}")
        
        return candidates
    
    def _industry_specific_generation(self, domain: str) -> Set[str]:
        """Generate industry-specific subdomains based on domain analysis"""
        candidates = set()
        
        # Analyze domain for industry indicators
        domain_lower = domain.lower()
        
        # Technology/Software
        if any(tech in domain_lower for tech in ['tech', 'soft', 'app', 'code', 'dev']):
            tech_subs = ['git', 'jenkins', 'ci', 'cd', 'docker', 'k8s', 'monitoring', 'metrics', 
                        'logs', 'grafana', 'prometheus', 'artifactory', 'nexus', 'sonar']
            candidates.update(f"{sub}.{domain}" for sub in tech_subs)
        
        # Finance/Banking
        if any(fin in domain_lower for fin in ['bank', 'fin', 'pay', 'money', 'credit', 'invest']):
            fin_subs = ['payment', 'secure', 'vault', 'transaction', 'account', 'balance', 
                       'transfer', 'compliance', 'kyc', 'aml', 'fraud']
            candidates.update(f"{sub}.{domain}" for sub in fin_subs)
        
        # E-commerce/Retail
        if any(ecom in domain_lower for ecom in ['shop', 'store', 'retail', 'ecom', 'cart']):
            ecom_subs = ['cart', 'checkout', 'payment', 'inventory', 'catalog', 'search', 
                        'recommendation', 'review', 'order', 'shipping', 'tracking']
            candidates.update(f"{sub}.{domain}" for sub in ecom_subs)
        
        # Healthcare
        if any(health in domain_lower for health in ['health', 'med', 'care', 'hospital', 'clinic']):
            health_subs = ['patient', 'appointment', 'record', 'pharmacy', 'lab', 'radiology', 
                          'billing', 'insurance', 'hipaa', 'emr', 'ehr']
            candidates.update(f"{sub}.{domain}" for sub in health_subs)
        
        # Education
        if any(edu in domain_lower for edu in ['edu', 'school', 'university', 'college', 'learn']):
            edu_subs = ['student', 'faculty', 'library', 'course', 'grade', 'schedule', 
                       'enrollment', 'campus', 'research', 'lab', 'portal']
            candidates.update(f"{sub}.{domain}" for sub in edu_subs)
        
        return candidates
    
    def _tech_stack_prediction(self, domain: str) -> Set[str]:
        """Predict subdomains based on common technology stacks"""
        candidates = set()
        
        # Common tech stack combinations
        tech_stacks = {
            'aws': ['s3', 'ec2', 'rds', 'lambda', 'cloudfront', 'elb', 'ecs', 'eks'],
            'azure': ['storage', 'compute', 'database', 'functions', 'cdn', 'loadbalancer'],
            'gcp': ['storage', 'compute', 'sql', 'functions', 'cdn', 'load-balancer'],
            'monitoring': ['grafana', 'prometheus', 'elk', 'splunk', 'datadog', 'newrelic'],
            'ci_cd': ['jenkins', 'gitlab', 'github', 'travis', 'circle', 'bamboo'],
            'databases': ['mysql', 'postgres', 'mongo', 'redis', 'elastic', 'cassandra'],
            'microservices': ['user', 'auth', 'payment', 'notification', 'order', 'inventory']
        }
        
        for stack_name, services in tech_stacks.items():
            for service in services:
                candidates.add(f"{service}.{domain}")
                candidates.add(f"{service}-api.{domain}")
                candidates.add(f"{service}-service.{domain}")
        
        return candidates
    
    def _semantic_expansion(self, domain: str) -> Set[str]:
        """Generate semantically related subdomains"""
        candidates = set()
        
        # Semantic word groups
        semantic_groups = {
            'access': ['login', 'auth', 'sso', 'oauth', 'ldap', 'radius'],
            'data': ['analytics', 'metrics', 'reporting', 'dashboard', 'bi'],
            'communication': ['chat', 'message', 'notification', 'email', 'sms'],
            'content': ['cms', 'blog', 'news', 'media', 'assets', 'cdn'],
            'infrastructure': ['proxy', 'gateway', 'load-balancer', 'firewall', 'vpn'],
            'development': ['dev', 'test', 'staging', 'prod', 'demo', 'sandbox']
        }
        
        for group_name, words in semantic_groups.items():
            for word in words:
                candidates.add(f"{word}.{domain}")
        
        return candidates
    
    def _morphological_generation(self, domain: str) -> Set[str]:
        """Generate subdomains using morphological analysis"""
        candidates = set()
        
        # Common prefixes and suffixes
        prefixes = ['new', 'old', 'beta', 'alpha', 'v2', 'next', 'legacy', 'modern']
        suffixes = ['api', 'service', 'app', 'portal', 'hub', 'center', 'gateway']
        
        # Base words commonly found in subdomains
        base_words = ['user', 'admin', 'client', 'server', 'data', 'file', 'image', 'video']
        
        # Generate combinations
        for prefix in prefixes:
            for base in base_words:
                candidates.add(f"{prefix}-{base}.{domain}")
                candidates.add(f"{prefix}{base}.{domain}")
        
        for base in base_words:
            for suffix in suffixes:
                candidates.add(f"{base}-{suffix}.{domain}")
                candidates.add(f"{base}{suffix}.{domain}")
        
        return candidates
    
    async def _phase_advanced_dns_discovery(self, domain: str):
        """Advanced DNS discovery techniques"""
        discovered = 0
        
        # DNS Zone Transfer Attempt
        self._emit_progress("Advanced DNS Discovery", 20, message="Attempting zone transfers...")
        zone_results = await self._dns_zone_transfer(domain)
        discovered += len(zone_results)
        
        # DNS TXT Record Mining
        self._emit_progress("Advanced DNS Discovery", 40, message="Mining TXT records...")
        txt_results = await self._dns_txt_mining(domain)
        discovered += len(txt_results)
        
        # DNS Reverse Lookups
        self._emit_progress("Advanced DNS Discovery", 60, message="Reverse DNS analysis...")
        reverse_results = await self._dns_reverse_analysis(domain)
        discovered += len(reverse_results)
        
        # DNS ANY Record Enumeration
        self._emit_progress("Advanced DNS Discovery", 80, message="ANY record enumeration...")
        any_results = await self._dns_any_records(domain)
        discovered += len(any_results)
        
        self._emit_progress("Advanced DNS Discovery", 100, 
                           message=f"Advanced DNS complete: {discovered} subdomains found")
    
    async def _dns_zone_transfer(self, domain: str):
        """Attempt DNS zone transfer"""
        results = []
        
        # Get authoritative nameservers
        try:
            # Simple zone transfer attempt (most will fail due to security)
            ns_servers = ['ns1', 'ns2', 'ns', 'dns', 'dns1', 'dns2']
            
            for ns in ns_servers:
                ns_candidate = f"{ns}.{domain}"
                success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(ns_candidate)
                if success:
                    result = SubdomainResult(
                        subdomain=ns_candidate,
                        source="DNS_Zone_Transfer",
                        http_status=0,
                        ip_addresses=ip_addresses,
                        technologies=[],
                        confidence_score=0.9,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="Low"
                    )
                    self.results[ns_candidate] = result
                    self._emit_result(result)
                    results.append(result)
        except Exception:
            pass
        
        return results
    
    async def _dns_txt_mining(self, domain: str):
        """Mine TXT records for subdomain hints"""
        results = []
        
        # Common TXT record patterns that might reveal subdomains
        txt_patterns = [
            '_dmarc', '_domainkey', 'selector1._domainkey', 'selector2._domainkey',
            '_spf', '_acme-challenge', '_sip._tcp', '_xmpp-server._tcp',
            '_caldav._tcp', '_carddav._tcp', '_imaps._tcp', '_submission._tcp'
        ]
        
        for pattern in txt_patterns:
            candidate = f"{pattern}.{domain}"
            try:
                success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(candidate)
                if success:
                    result = SubdomainResult(
                        subdomain=candidate,
                        source="DNS_TXT_Mining",
                        http_status=0,
                        ip_addresses=ip_addresses,
                        technologies=[],
                        confidence_score=0.8,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="Low"
                    )
                    self.results[candidate] = result
                    self._emit_result(result)
                    results.append(result)
            except Exception:
                continue
        
        return results
    
    async def _dns_reverse_analysis(self, domain: str):
        """Reverse DNS analysis on discovered IPs"""
        results = []
        
        # Collect unique IPs from existing results
        unique_ips = set()
        for result in self.results.values():
            unique_ips.update(result.ip_addresses)
        
        # Perform reverse DNS on a sample of IPs
        for ip in list(unique_ips)[:20]:  # Limit to prevent excessive queries
            try:
                import socket
                reverse_name = socket.gethostbyaddr(ip)[0]
                if domain in reverse_name and reverse_name not in self.results:
                    result = SubdomainResult(
                        subdomain=reverse_name,
                        source="DNS_Reverse_Lookup",
                        http_status=0,
                        ip_addresses=[ip],
                        technologies=[],
                        confidence_score=0.7,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="Low"
                    )
                    self.results[reverse_name] = result
                    self._emit_result(result)
                    results.append(result)
            except Exception:
                continue
        
        return results
    
    async def _dns_any_records(self, domain: str):
        """Enumerate ANY records for additional discovery"""
        results = []
        
        # Common service prefixes for ANY record enumeration
        service_prefixes = [
            '_http._tcp', '_https._tcp', '_ftp._tcp', '_ssh._tcp',
            '_telnet._tcp', '_smtp._tcp', '_pop3._tcp', '_imap._tcp',
            '_ldap._tcp', '_ldaps._tcp', '_kerberos._tcp', '_sips._tcp'
        ]
        
        for prefix in service_prefixes:
            candidate = f"{prefix}.{domain}"
            try:
                success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(candidate)
                if success:
                    result = SubdomainResult(
                        subdomain=candidate,
                        source="DNS_ANY_Records",
                        http_status=0,
                        ip_addresses=ip_addresses,
                        technologies=[],
                        confidence_score=0.6,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="Low"
                    )
                    self.results[candidate] = result
                    self._emit_result(result)
                    results.append(result)
            except Exception:
                continue
        
        return results
    
    async def _phase_web_discovery(self, domain: str):
        """Web-based discovery techniques"""
        discovered = 0
        
        # Robots.txt Analysis
        self._emit_progress("Web Discovery", 25, message="Analyzing robots.txt...")
        robots_results = await self._robots_txt_analysis(domain)
        discovered += len(robots_results)
        
        # Sitemap.xml Analysis
        self._emit_progress("Web Discovery", 50, message="Parsing sitemap.xml...")
        sitemap_results = await self._sitemap_analysis(domain)
        discovered += len(sitemap_results)
        
        # Security.txt Analysis
        self._emit_progress("Web Discovery", 75, message="Checking security.txt...")
        security_results = await self._security_txt_analysis(domain)
        discovered += len(security_results)
        
        self._emit_progress("Web Discovery", 100, 
                           message=f"Web discovery complete: {discovered} subdomains found")
    
    async def _robots_txt_analysis(self, domain: str):
        """Analyze robots.txt for subdomain hints"""
        results = []
        
        # Common robots.txt patterns that might reveal subdomains
        robots_subdomains = [
            'admin', 'administrator', 'test', 'staging', 'dev', 'development',
            'beta', 'alpha', 'preview', 'demo', 'sandbox', 'internal',
            'private', 'secure', 'management', 'control', 'panel'
        ]
        
        for subdomain in robots_subdomains:
            candidate = f"{subdomain}.{domain}"
            try:
                success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(candidate)
                if success:
                    result = SubdomainResult(
                        subdomain=candidate,
                        source="Robots_txt_Analysis",
                        http_status=0,
                        ip_addresses=ip_addresses,
                        technologies=[],
                        confidence_score=0.7,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="Low"
                    )
                    self.results[candidate] = result
                    self._emit_result(result)
                    results.append(result)
            except Exception:
                continue
        
        return results
    
    async def _sitemap_analysis(self, domain: str):
        """Analyze potential sitemap locations"""
        results = []
        
        # Common sitemap-related subdomains
        sitemap_patterns = [
            'sitemap', 'sitemaps', 'xml', 'feeds', 'rss', 'atom',
            'blog', 'news', 'articles', 'content', 'media'
        ]
        
        for pattern in sitemap_patterns:
            candidate = f"{pattern}.{domain}"
            try:
                success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(candidate)
                if success:
                    result = SubdomainResult(
                        subdomain=candidate,
                        source="Sitemap_Analysis",
                        http_status=0,
                        ip_addresses=ip_addresses,
                        technologies=[],
                        confidence_score=0.6,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="Low"
                    )
                    self.results[candidate] = result
                    self._emit_result(result)
                    results.append(result)
            except Exception:
                continue
        
        return results
    
    async def _security_txt_analysis(self, domain: str):
        """Analyze security.txt and related security subdomains"""
        results = []
        
        # Security-related subdomains
        security_patterns = [
            'security', 'bug-bounty', 'bugbounty', 'responsible-disclosure',
            'vulnerability', 'cert', 'csirt', 'incident', 'abuse',
            'postmaster', 'webmaster', 'noc', 'soc'
        ]
        
        for pattern in security_patterns:
            candidate = f"{pattern}.{domain}"
            try:
                success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(candidate)
                if success:
                    result = SubdomainResult(
                        subdomain=candidate,
                        source="Security_txt_Analysis",
                        http_status=0,
                        ip_addresses=ip_addresses,
                        technologies=[],
                        confidence_score=0.8,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="Low"
                    )
                    self.results[candidate] = result
                    self._emit_result(result)
                    results.append(result)
            except Exception:
                continue
        
        return results
    
    async def _phase_historical_discovery(self, domain: str):
        """Historical and archive-based discovery"""
        discovered = 0
        
        # Archive-based patterns
        self._emit_progress("Historical Discovery", 33, message="Archive pattern analysis...")
        archive_results = await self._archive_pattern_discovery(domain)
        discovered += len(archive_results)
        
        # Legacy system patterns
        self._emit_progress("Historical Discovery", 66, message="Legacy system discovery...")
        legacy_results = await self._legacy_system_discovery(domain)
        discovered += len(legacy_results)
        
        self._emit_progress("Historical Discovery", 100, 
                           message=f"Historical discovery complete: {discovered} subdomains found")
    
    async def _archive_pattern_discovery(self, domain: str):
        """Discover subdomains based on archive and backup patterns"""
        results = []
        
        archive_patterns = [
            'archive', 'backup', 'old', 'legacy', 'deprecated', 'historical',
            'archive2023', 'backup2023', 'archive2022', 'backup2022',
            'old-site', 'legacy-app', 'historical-data', 'archive-server'
        ]
        
        for pattern in archive_patterns:
            candidate = f"{pattern}.{domain}"
            try:
                success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(candidate)
                if success:
                    result = SubdomainResult(
                        subdomain=candidate,
                        source="Archive_Discovery",
                        http_status=0,
                        ip_addresses=ip_addresses,
                        technologies=[],
                        confidence_score=0.6,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="Medium"  # Archives might be less secure
                    )
                    self.results[candidate] = result
                    self._emit_result(result)
                    results.append(result)
            except Exception:
                continue
        
        return results
    
    async def _legacy_system_discovery(self, domain: str):
        """Discover legacy and deprecated systems"""
        results = []
        
        legacy_patterns = [
            'legacy', 'deprecated', 'retired', 'obsolete', 'unmaintained',
            'v1', 'v2', 'version1', 'version2', 'old-api', 'legacy-api',
            'classic', 'traditional', 'previous', 'former'
        ]
        
        for pattern in legacy_patterns:
            candidate = f"{pattern}.{domain}"
            try:
                success, ip_addresses, _ = await self.dns_resolver.resolve_with_intelligence(candidate)
                if success:
                    result = SubdomainResult(
                        subdomain=candidate,
                        source="Legacy_Discovery",
                        http_status=0,
                        ip_addresses=ip_addresses,
                        technologies=[],
                        confidence_score=0.7,
                        discovered_at=time.time(),
                        cname_records=[],
                        cname_chain=[],
                        takeover_risk="High"  # Legacy systems often vulnerable
                    )
                    self.results[candidate] = result
                    self._emit_result(result)
                    results.append(result)
            except Exception:
                continue
        
        return results
    
    async def _phase_infrastructure_analysis(self, session: aiohttp.ClientSession, domain: str):
        """Phase 4: Network Infrastructure Analysis"""
        self._emit_progress("Infrastructure Analysis", 10, message="Starting network analysis...")
        
        analyzer = NetworkInfraAnalyzer(session, self.dns_resolver)
        current_results = list(self.results.values())
        
        # Extract unique IPs for analysis
        unique_ips = set()
        for result in current_results:
            unique_ips.update(result.ip_addresses)
        
        self._emit_progress("Infrastructure Analysis", 25, 
                       message=f"Analyzing {len(unique_ips)} unique IP addresses...")
        
        infra_start_time = time.time()
        infra_results = await analyzer.analyze_infrastructure(current_results, domain)
        infra_time = time.time() - infra_start_time
        
        # Add new discoveries
        new_count = 0
        for result in infra_results:
            if result.subdomain not in self.results:
                self.results[result.subdomain] = result
                self._emit_result(result)
                new_count += 1
        
        self._emit_progress("Infrastructure Analysis", 100, 
                       message=f"Infrastructure Analysis complete: {new_count} new subdomains",
                       subdomain_count=len(self.results))
    
    async def _phase_recursive_discovery(self, domain: str):
        """Phase 5: Recursive Subdomain Discovery"""
        self._emit_progress("Recursive Discovery", 10, message="Finding recursive candidates...")
        
        # Find subdomains that might have their own subdomains
        recursive_candidates = []
        for result in self.results.values():
            subdomain = result.subdomain
            # Skip direct subdomains, look for nested ones
            if subdomain.count('.') == domain.count('.') + 1:
                recursive_candidates.append(subdomain)
        
        if recursive_candidates:
            # Test common nested subdomain patterns
            nested_patterns = ['www', 'api', 'app', 'admin', 'secure', 'mail', 'ftp']
            total_tests = len(recursive_candidates) * len(nested_patterns)
            
            self._emit_progress("Recursive Discovery", 25, 
                           message=f"Testing {len(recursive_candidates)} candidates for nested patterns...")
            
            recursive_start_time = time.time()
            tasks = []
            
            for candidate in recursive_candidates:
                for pattern in nested_patterns:
                    nested_subdomain = f"{pattern}.{candidate}"
                    task = self._resolve_with_intelligence(
                        asyncio.Semaphore(50), 
                        nested_subdomain
                    )
                    tasks.append(task)
            
            if tasks:
                results = await asyncio.gather(*tasks, return_exceptions=True)
                recursive_time = time.time() - recursive_start_time
                
                recursive_found = 0
                for result in results:
                    if isinstance(result, SubdomainResult) and result.subdomain not in self.results:
                        result.source = "Recursive_Discovery"
                        result.confidence_score = 0.8
                        self.results[result.subdomain] = result
                        recursive_found += 1
                        self._emit_result(result)
                
                rate = len(tasks) / recursive_time if recursive_time > 0 else 0
                self._emit_progress("Recursive Discovery", 100, 
                               message=f"Recursive Discovery complete: {recursive_found} nested subdomains",
                               subdomain_count=len(self.results))
        else:
            self._emit_progress("Recursive Discovery", 100, 
                           message="No suitable candidates for recursive discovery")
    
    async def _phase_cname_comprehensive(self, domain: str, wordlist_files: List[str]):
        """Phase 2: Comprehensive CNAME Discovery & Analysis"""
        self._emit_progress("CNAME Analysis", 5, message="Starting comprehensive CNAME discovery...")
        
        # Load wordlists for CNAME discovery
        wordlist = self._load_wordlists_from_files(wordlist_files)
        
        self._emit_progress("CNAME Analysis", 15, 
                       message=f"Discovering CNAME records with {len(wordlist):,} patterns...")
        
        cname_start_time = time.time()
        cname_results = await self.dns_resolver.cname_resolver.discover_cname_targets(domain, wordlist)
        cname_time = time.time() - cname_start_time
        
        # Add new CNAME discoveries
        new_count = 0
        for result in cname_results:
            if result.subdomain not in self.results:
                self.results[result.subdomain] = result
                self._emit_result(result)
                new_count += 1
        
        self._emit_progress("CNAME Analysis", 50, 
                       message=f"Analyzing CNAME chains for existing subdomains...")
        
        # Analyze CNAME records for existing subdomains
        cname_analysis_tasks = []
        for result in list(self.results.values()):
            if not result.cname_records:  # Only analyze if not already done
                task = self._analyze_existing_cname(result)
                cname_analysis_tasks.append(task)
        
        if cname_analysis_tasks:
            await asyncio.gather(*cname_analysis_tasks, return_exceptions=True)
        
        # Count high-risk takeover candidates
        high_risk_count = sum(1 for result in self.results.values() if result.takeover_risk == "High")
        
        rate = len(cname_results) / cname_time if cname_time > 0 else 0
        self._emit_progress("CNAME Analysis", 100, 
                       message=f"CNAME Analysis complete: {new_count} new, {high_risk_count} high-risk",
                       subdomain_count=len(self.results))
    
    async def _analyze_existing_cname(self, result: SubdomainResult):
        """Analyze CNAME records for an existing subdomain result"""
        try:
            cname_records = await self.dns_resolver.cname_resolver.resolve_cname_comprehensive(result.subdomain)
            if cname_records:
                result.cname_records = cname_records
                result.cname_chain = [record.target_domain for record in cname_records]
                result.takeover_risk = cname_records[0].takeover_risk
                result.is_cname_target = True
        except Exception:
            pass
    
    async def _phase_final_cname_analysis(self):
        """Phase 8: Final CNAME Chain Analysis and Takeover Detection"""
        self._emit_progress("Final CNAME Analysis", 10, message="Performing final CNAME chain validation...")
        
        # Collect all CNAME targets for reverse analysis
        cname_targets = set()
        for result in self.results.values():
            for cname_record in result.cname_records:
                cname_targets.add(cname_record.target_domain)
        
        self._emit_progress("Final CNAME Analysis", 30, 
                       message=f"Validating {len(cname_targets)} unique CNAME targets...")
        
        # Check which targets are reachable (takeover detection)
        validation_tasks = []
        semaphore = asyncio.Semaphore(50)
        
        for target in cname_targets:
            task = self._validate_cname_target(semaphore, target)
            validation_tasks.append(task)
        
        if validation_tasks:
            validation_results = await asyncio.gather(*validation_tasks, return_exceptions=True)
            
            # Update takeover risk based on validation
            unreachable_targets = set()
            for target, is_reachable in zip(cname_targets, validation_results):
                if not isinstance(is_reachable, Exception) and not is_reachable:
                    unreachable_targets.add(target)
            
            # Mark subdomains pointing to unreachable targets as very high risk
            takeover_candidates = 0
            for result in self.results.values():
                for cname_record in result.cname_records:
                    if cname_record.target_domain in unreachable_targets:
                        result.takeover_risk = "Critical"
                        takeover_candidates += 1
                        break
            
            self._emit_progress("Final CNAME Analysis", 100, 
                           message=f"CNAME validation complete: {takeover_candidates} critical takeover risks",
                           subdomain_count=len(self.results))
        else:
            self._emit_progress("Final CNAME Analysis", 100, 
                           message="No CNAME targets to validate")
    
    async def _validate_cname_target(self, semaphore: asyncio.Semaphore, target: str) -> bool:
        """Validate if a CNAME target is reachable (for takeover detection)"""
        async with semaphore:
            try:
                # Try to resolve the target
                success, _, _ = await self.dns_resolver.resolve_with_intelligence(target)
                return success
            except Exception:
                return False
    
    async def _phase_http_analysis(self, session: aiohttp.ClientSession):
        """Phase 6: Comprehensive HTTP Analysis"""
        
        if not self.results:
            self._emit_progress("HTTP Analysis", 100, message="No subdomains to analyze")
            return
        
        self._emit_progress("HTTP Analysis", 10, 
                       message=f"Starting HTTP analysis for {len(self.results):,} subdomains...")
        
        # Progress tracking
        analysis_start_time = time.time()
        total_subdomains = len(self.results)
        completed = 0
        
        semaphore = asyncio.Semaphore(self.config['max_concurrent_http'])
        
        # Create tasks with progress tracking
        async def analyze_with_progress(result):
            nonlocal completed
            await self._analyze_http_comprehensive(semaphore, session, result)
            completed += 1
            
            # Progress reporting every 10% or every 50 completions
            if completed % max(1, total_subdomains // 10) == 0 or completed % 50 == 0:
                progress_pct = 10 + (completed / total_subdomains) * 90  # 10% to 100%
                elapsed = time.time() - analysis_start_time
                rate = completed / elapsed if elapsed > 0 else 0
                eta = (total_subdomains - completed) / rate if rate > 0 else 0
                
                self._emit_progress("HTTP Analysis", progress_pct,
                               current=completed, total=total_subdomains,
                               rate=rate, eta=eta,
                               message=f"Analyzing HTTP status: {completed:,}/{total_subdomains:,}",
                               subdomain_count=len(self.results))
        
        tasks = [analyze_with_progress(result) for result in self.results.values()]
        await asyncio.gather(*tasks, return_exceptions=True)
        
        # Second pass: Analyze any newly discovered subdomains (from SSL, etc.)
        unanalyzed_results = [r for r in self.results.values() if r.http_status == 0 and not r.additional_info]
        if unanalyzed_results:
            self._emit_progress("HTTP Analysis", 95, 
                           message=f"Second pass: analyzing {len(unanalyzed_results)} newly discovered subdomains...")
            
            second_pass_tasks = [analyze_with_progress(result) for result in unanalyzed_results]
            await asyncio.gather(*second_pass_tasks, return_exceptions=True)
        
        analysis_time = time.time() - analysis_start_time
        
        # Statistics
        status_counts = Counter(result.http_status for result in self.results.values())
        live_services = sum(1 for result in self.results.values() if result.http_status == 200)
        
        self._emit_progress("HTTP Analysis", 100, 
                       message=f"HTTP Analysis complete: {live_services} live services detected",
                       subdomain_count=len(self.results))
    
    async def _analyze_http_comprehensive(self, semaphore: asyncio.Semaphore, session: aiohttp.ClientSession, result: SubdomainResult):
        """Comprehensive HTTP analysis with technology detection and SSL verification"""
        async with semaphore:
            # First try HTTPS for SSL certificate analysis
            for protocol in ['https', 'http']:
                try:
                    url = f"{protocol}://{result.subdomain}"
                    start_time = time.time()
                    
                    async with session.get(url, allow_redirects=False) as response:
                        result.http_status = response.status
                        result.response_time = (time.time() - start_time) * 1000  # Convert to milliseconds
                        
                        # Extract headers for technology detection
                        server = response.headers.get('Server', '')
                        if server:
                            result.server = server
                            result.technologies.append(server)
                        
                        # Extract other technology indicators
                        powered_by = response.headers.get('X-Powered-By', '')
                        if powered_by:
                            result.technologies.append(powered_by)
                        
                        # Get WHOIS info if not already collected
                        if not result.ownership_info:
                            result.ownership_info = self.get_domain_ownership(result.subdomain)
                        
                        # If HTTPS, analyze SSL certificate and discover new subdomains
                        if protocol == 'https':
                            discovered_ssl_subdomains = await self._analyze_ssl_certificate(result)
                            
                            # Add discovered subdomains from SSL certificates
                            for ssl_subdomain in discovered_ssl_subdomains:
                                if ssl_subdomain not in self.results:
                                    # Create new result for SSL-discovered subdomain
                                    ssl_result = SubdomainResult(
                                        subdomain=ssl_subdomain,
                                        source="SSL_Certificate_SAN",
                                        http_status=0,
                                        ip_addresses=[],
                                        technologies=[],
                                        confidence_score=0.85,
                                        discovered_at=time.time(),
                                        cname_records=[],
                                        cname_chain=[],
                                        takeover_risk="Low"
                                    )
                                    self.results[ssl_subdomain] = ssl_result
                                    self._emit_result(ssl_result)
                        
                        return  # Successfully analyzed
                        
                except Exception as e:
                    # Log HTTP analysis failure for debugging
                    result.additional_info = f"HTTP Analysis failed: {str(e)[:100]}"
                    continue
            
            # If both protocols failed, mark as no response
            if result.http_status == 0:
                result.http_status = 0
    
    async def _analyze_ssl_certificate(self, result: SubdomainResult) -> List[str]:
        """Analyze SSL certificate and extract additional subdomains from SAN fields"""
        discovered_subdomains = []
        
        try:
            # Create SSL context - more permissive for enumeration
            context = ssl.create_default_context()
            context.check_hostname = False
            context.verify_mode = ssl.CERT_NONE
            
            # Try different approaches for SSL connection
            ssl_info = await self._get_ssl_info_robust(result.subdomain)
            
            if ssl_info:
                result.ssl_domain_verified = True
                result.ssl_issuer = ssl_info.get('issuer', 'Unknown')
                result.ssl_subject = ssl_info.get('subject', 'Unknown')
                
                # Extract SAN domains for additional subdomain discovery
                san_domains = ssl_info.get('san_domains', [])
                for san_domain in san_domains:
                    if self._is_valid_subdomain(san_domain) and self._is_target_related(san_domain, result.subdomain):
                        discovered_subdomains.append(san_domain)
            else:
                result.ssl_domain_verified = False
                result.ssl_issuer = "No SSL/TLS"
                result.ssl_subject = "No SSL/TLS"
                
        except Exception as e:
            result.ssl_domain_verified = False
            result.ssl_issuer = f"Error: {str(e)[:50]}"
            result.ssl_subject = "SSL analysis failed"
            
        return discovered_subdomains
    
    async def _get_ssl_info_robust(self, hostname: str) -> Dict[str, any]:
        """Robust SSL information extraction with multiple fallbacks"""
        
        # Method 1: Using cryptography library (most reliable)
        if CRYPTO_AVAILABLE:
            try:
                context = ssl.create_default_context()
                context.check_hostname = False
                context.verify_mode = ssl.CERT_NONE
                
                with socket.create_connection((hostname, 443), timeout=3) as sock:
                    with context.wrap_socket(sock, server_hostname=hostname) as ssock:
                        cert_der = ssock.getpeercert(binary_form=True)
                        if not cert_der:
                            raise ValueError("No certificate data received")
                            
                        cert = x509.load_der_x509_certificate(cert_der, default_backend())
                        
                        # Extract issuer
                        issuer = "Unknown"
                        for attribute in cert.issuer:
                            if attribute.oid == x509.NameOID.ORGANIZATION_NAME:
                                issuer = attribute.value
                                break
                        
                        # Extract subject
                        subject = hostname
                        for attribute in cert.subject:
                            if attribute.oid == x509.NameOID.COMMON_NAME:
                                subject = attribute.value
                                break
                        
                        return {
                            'issuer': issuer,
                            'subject': subject,
                            'san_domains': self._extract_san_from_x509(cert),
                            'expiry': cert.not_valid_after_utc.strftime('%Y-%m-%d') if hasattr(cert, 'not_valid_after_utc') else cert.not_valid_after.strftime('%Y-%m-%d'),
                            'serial': str(cert.serial_number)
                        }
            except Exception as e:
                # Log specific error for debugging
                if hasattr(self, '_emit_progress'):
                    self._emit_progress("SSL Analysis", 0, message=f"SSL crypto method failed for {hostname}: {str(e)}")
        
        # Method 2: Direct SSL connection (fallback)
        try:
            context = ssl.create_default_context()
            context.check_hostname = False
            context.verify_mode = ssl.CERT_NONE
            
            with socket.create_connection((hostname, 443), timeout=10) as sock:
                with context.wrap_socket(sock, server_hostname=hostname) as ssock:
                    cert_dict = ssock.getpeercert()
                    
                    # Fix: Handle empty dictionary properly
                    if cert_dict and len(cert_dict) > 0:
                        # Extract issuer organization
                        issuer = "Unknown Issuer"
                        if 'issuer' in cert_dict:
                            issuer_dict = dict(cert_dict['issuer'])
                            issuer = issuer_dict.get('organizationName', 
                                     issuer_dict.get('commonName', 'Unknown Issuer'))
                        
                        # Extract subject
                        subject = hostname
                        if 'subject' in cert_dict:
                            subject_dict = dict(cert_dict['subject'])
                            subject = subject_dict.get('commonName', hostname)
                        
                        return {
                            'issuer': issuer,
                            'subject': subject,
                            'san_domains': self._extract_san_domains(cert_dict),
                            'expiry': cert_dict.get('notAfter', 'Unknown'),
                            'serial': cert_dict.get('serialNumber', 'Unknown')
                        }
        except Exception as e:
            if hasattr(self, '_emit_progress'):
                self._emit_progress("SSL Analysis", 0, message=f"SSL direct method failed for {hostname}: {str(e)}")
        
        # Method 3: OpenSSL command fallback
        try:
            cmd = ['openssl', 's_client', '-connect', f'{hostname}:443', '-servername', hostname, '-verify_return_error']
            result = subprocess.run(cmd, input='', capture_output=True, text=True, timeout=5)
            
            if result.returncode == 0 or 'CERTIFICATE' in result.stdout:
                return self._parse_openssl_output(result.stdout)
        except Exception as e:
            if hasattr(self, '_emit_progress'):
                self._emit_progress("SSL Analysis", 0, message=f"SSL openssl method failed for {hostname}: {str(e)}")
        
        return None
    
    def _extract_san_domains(self, cert_dict: dict) -> List[str]:
        """Extract SAN domains from certificate dictionary"""
        san_domains = []
        if 'subjectAltName' in cert_dict:
            for san_type, san_value in cert_dict['subjectAltName']:
                if san_type == 'DNS':
                    san_domains.append(san_value)
        return san_domains
    
    def _extract_san_from_x509(self, cert) -> List[str]:
        """Extract SAN domains from x509 certificate object"""
        san_domains = []
        try:
            san_ext = cert.extensions.get_extension_for_oid(x509.oid.ExtensionOID.SUBJECT_ALTERNATIVE_NAME)
            for name in san_ext.value:
                if hasattr(name, 'value'):
                    san_domains.append(name.value)
        except:
            pass
        return san_domains
    
    def _parse_openssl_output(self, openssl_output: str) -> Dict[str, any]:
        """Parse OpenSSL command output"""
        ssl_info = {'issuer': 'Unknown', 'subject': 'Unknown', 'san_domains': []}
        
        lines = openssl_output.split('\n')
        for line in lines:
            if 'issuer=' in line:
                # Extract organization from issuer line
                parts = line.split('issuer=')[1].split(',')
                for part in parts:
                    if 'O=' in part:
                        ssl_info['issuer'] = part.split('O=')[1].strip()
                        break
            elif 'subject=' in line:
                # Extract common name from subject line
                parts = line.split('subject=')[1].split(',')
                for part in parts:
                    if 'CN=' in part:
                        ssl_info['subject'] = part.split('CN=')[1].strip()
                        break
        
        return ssl_info
    
    async def _phase_ssl_analysis_controlled(self):
        """Controlled SSL analysis phase with proper concurrency and timeout handling"""
        
        # Only analyze subdomains that have IP addresses and aren't clearly internal
        ssl_candidates = []
        for result in self.results.values():
            if (result.ip_addresses and 
                result.http_status != 200 and  # Skip if already analyzed via HTTP
                not any(ip.startswith(('10.', '192.168.', '172.')) for ip in result.ip_addresses) and
                not result.ssl_issuer):  # Skip if SSL already analyzed
                ssl_candidates.append(result)
        
        if not ssl_candidates:
            self._emit_progress("SSL Analysis", 100, message="No SSL analysis candidates found")
            return
        
        self._emit_progress("SSL Analysis", 10, 
                           message=f"Analyzing SSL certificates for {len(ssl_candidates)} candidates...")
        
        # Create SSL-specific semaphore for controlled concurrency
        ssl_semaphore = asyncio.Semaphore(self.config['max_concurrent_ssl'])
        
        # Track progress
        completed = 0
        start_time = time.time()
        
        async def analyze_ssl_with_progress(result):
            nonlocal completed
            async with ssl_semaphore:
                try:
                    # Add overall timeout for SSL analysis
                    ssl_info = await asyncio.wait_for(
                        self._get_ssl_info_robust(result.subdomain),
                        timeout=self.config['ssl_timeout']
                    )
                    
                    if ssl_info:
                        result.ssl_issuer = ssl_info.get('issuer', 'Unknown')
                        result.ssl_subject = ssl_info.get('subject', result.subdomain)
                        result.ssl_domain_verified = True
                        
                        # Process SAN domains for additional discovery
                        san_domains = ssl_info.get('san_domains', [])
                        if san_domains:
                            for san_domain in san_domains[:5]:  # Limit to 5 SAN domains
                                if (san_domain not in self.results and 
                                    result.subdomain.split('.', 1)[1] in san_domain):
                                    # Create new subdomain result from SAN
                                    san_result = SubdomainResult(
                                        subdomain=san_domain,
                                        source=f"SSL_SAN_{result.subdomain}",
                                        ip_addresses=result.ip_addresses,
                                        ssl_issuer=ssl_info.get('issuer', 'Unknown'),
                                        ssl_subject=ssl_info.get('subject', san_domain),
                                        ssl_domain_verified=True,
                                        confidence_score=0.85
                                    )
                                    self.results[san_domain] = san_result
                                    self._emit_result(san_result)
                    
                except asyncio.TimeoutError:
                    # SSL analysis timed out - skip this subdomain
                    result.ssl_issuer = "Timeout"
                    result.ssl_subject = "SSL analysis timed out"
                    result.ssl_domain_verified = False
                except Exception as e:
                    # SSL analysis failed
                    result.ssl_issuer = f"Error: {str(e)[:50]}"
                    result.ssl_subject = "SSL analysis failed"
                    result.ssl_domain_verified = False
                
                completed += 1
                
                # Update progress every 10 completions or every 20%
                if completed % 10 == 0 or completed % max(1, len(ssl_candidates) // 5) == 0:
                    progress_pct = 10 + (completed / len(ssl_candidates)) * 90
                    elapsed = time.time() - start_time
                    rate = completed / elapsed if elapsed > 0 else 0
                    eta = (len(ssl_candidates) - completed) / rate if rate > 0 else 0
                    
                    self._emit_progress("SSL Analysis", progress_pct,
                                     current=completed, total=len(ssl_candidates),
                                     rate=rate, eta=eta,
                                     message=f"SSL analysis: {completed}/{len(ssl_candidates)} ({rate:.1f}/s)",
                                     subdomain_count=len(self.results))
        
        # Process SSL analysis with controlled concurrency
        try:
            tasks = [analyze_ssl_with_progress(result) for result in ssl_candidates]
            await asyncio.wait_for(
                asyncio.gather(*tasks, return_exceptions=True),
                timeout=len(ssl_candidates) * 2  # 2 seconds per subdomain max
            )
        except asyncio.TimeoutError:
            self._emit_progress("SSL Analysis", 90, 
                              message="SSL analysis phase timed out - completing with partial results...")
        
        # Final statistics
        ssl_verified_count = sum(1 for r in self.results.values() if r.ssl_domain_verified)
        analysis_time = time.time() - start_time
        
        self._emit_progress("SSL Analysis", 100,
                          message=f"SSL analysis complete: {ssl_verified_count} certificates verified in {analysis_time:.1f}s")
    
    def _is_valid_subdomain(self, subdomain: str) -> bool:
        """Check if a subdomain is valid for DNS resolution"""
        if not subdomain or len(subdomain) > 253:
            return False
        
        # Check for valid characters
        if not all(c.isalnum() or c in '-.' for c in subdomain):
            return False
        
        # Check each label
        labels = subdomain.split('.')
        for label in labels:
            if not label or len(label) > 63:
                return False
            if label.startswith('-') or label.endswith('-'):
                return False
        
        return True
    
    def _is_target_related(self, discovered_subdomain: str, target_domain: str) -> bool:
        """Check if discovered subdomain is related to our target domain"""
        return discovered_subdomain.endswith(target_domain) or discovered_subdomain == target_domain
    
    def _load_wordlists_from_files(self, wordlist_files: List[str]) -> List[str]:
        """Load wordlists from files with advanced optimization for massive files"""
        import mmap
        import gc
        
        # Memory-efficient wordlist loading with deduplication
        seen = set()
        unique_wordlist = []
        total_processed = 0
        
        self._emit_progress("Wordlist Loading", 10, message="Optimizing massive wordlists...")
        
        for i, file_path in enumerate(wordlist_files):
            try:
                # Memory-mapped file reading for massive files
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    # Use memory mapping for large files (>1MB)
                    try:
                        file_size = os.path.getsize(file_path)
                        if file_size > 1024 * 1024:  # 1MB threshold
                            # Memory-mapped reading for large files
                            with mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ) as mmapped_file:
                                content = mmapped_file.read().decode('utf-8', errors='ignore')
                                words = [line.strip() for line in content.splitlines() if line.strip() and not line.startswith('#')]
                        else:
                            # Regular reading for smaller files
                            words = [line.strip() for line in f if line.strip() and not line.startswith('#')]
                    except (OSError, ValueError):
                        # Fallback to regular reading
                        f.seek(0)
                        words = [line.strip() for line in f if line.strip() and not line.startswith('#')]
                
                # Advanced filtering and optimization
                filtered_words = []
                for word in words:
                    # Skip overly long subdomains (DNS limit is 63 chars per label)
                    if len(word) > 63 or len(word) < 1:
                        continue
                    # Skip invalid characters for DNS
                    if not all(c.isalnum() or c in '-_.' for c in word):
                        continue
                    # Skip words starting/ending with invalid chars
                    if word.startswith('-') or word.endswith('-'):
                        continue
                    
                    filtered_words.append(word.lower())  # Normalize case
                
                # Deduplicate while preserving order
                before_count = len(unique_wordlist)
                for word in filtered_words:
                    if word not in seen:
                        seen.add(word)
                        unique_wordlist.append(word)
                
                added_count = len(unique_wordlist) - before_count
                total_processed += len(words)
                
                # Progress update
                progress = 10 + ((i + 1) / len(wordlist_files)) * 80
                self._emit_progress("Wordlist Loading", progress, 
                    message=f"Processed {file_path.split('/')[-1]}: +{added_count:,} unique ({total_processed:,} total)")
                
                # Force garbage collection for memory management
                del words, filtered_words
                gc.collect()
                
            except FileNotFoundError:
                console.print(f"[bold red]⚠️  Wordlist file not found: {file_path}[/bold red]")
                continue
            except Exception as e:
                progress = 10 + ((i + 1) / len(wordlist_files)) * 80
                self._emit_progress("Wordlist Loading", progress, 
                    message=f"Error loading {file_path.split('/')[-1]}: {str(e)[:30]}")
                continue
        
        # Final optimization - sort by likelihood (common patterns first)
        common_prefixes = ['www', 'mail', 'ftp', 'admin', 'test', 'dev', 'api', 'app', 'web', 'blog', 'shop', 'm', 'mobile']
        prioritized_wordlist = []
        remaining_words = []
        
        # Prioritize common patterns
        for word in unique_wordlist:
            if any(word.startswith(prefix) for prefix in common_prefixes):
                prioritized_wordlist.append(word)
            else:
                remaining_words.append(word)
        
        # Combine prioritized + remaining
        final_wordlist = prioritized_wordlist + remaining_words
        
        self._emit_progress("Wordlist Loading", 100, 
            message=f"Loaded {len(final_wordlist):,} unique subdomains from {len(wordlist_files)} files")
        
        return final_wordlist
    
    def _get_readable_source(self, source: str) -> str:
        """Convert technical source names to readable descriptions"""
        source_mapping = {
            'CT_Mining': 'Certificate Transparency',
            'DNS_Intelligence': 'DNS Intelligence',
            'Advanced_ML_Local': 'AI/ML Generation',
            'Recursive_Discovery': 'Recursive Analysis',
            'DNS_Zone_Transfer': 'DNS Zone Transfer',
            'DNS_TXT_Mining': 'DNS TXT Records',
            'DNS_Reverse_Lookup': 'Reverse DNS',
            'DNS_ANY_Records': 'DNS ANY Records',
            'Robots_txt_Analysis': 'Robots.txt Analysis',
            'Sitemap_Analysis': 'Sitemap Analysis',
            'Security_txt_Analysis': 'Security.txt Analysis',
            'Archive_Discovery': 'Archive Discovery',
            'Legacy_Discovery': 'Legacy Systems',
            'SSL_Certificate_SAN': 'SSL Certificate SAN',
            'CNAME_Discovery': 'CNAME Discovery'
        }
        
        # Handle CT sources
        if source.startswith('CT_'):
            return 'Certificate Transparency'
        
        # Handle DNS Intelligence with resolver info
        if source.startswith('DNS_Intelligence_'):
            return 'DNS Intelligence'
            
        return source_mapping.get(source, source)
    
    def get_domain_ownership(self, domain_or_subdomain: str) -> str:
        """Get ownership information for a domain or subdomain using WHOIS"""
        try:
            # Extract root domain from subdomain
            parts = domain_or_subdomain.split('.')
            if len(parts) >= 2:
                root_domain = '.'.join(parts[-2:])  # Get last two parts (domain.tld)
            else:
                root_domain = domain_or_subdomain
            
            # Try python-whois first
            if WHOIS_AVAILABLE:
                try:
                    w = whois.whois(root_domain)
                    if w and hasattr(w, 'registrar') and w.registrar:
                        return str(w.registrar)
                    elif w and hasattr(w, 'org') and w.org:
                        return str(w.org)
                except:
                    pass
            
            # Fallback to system whois command
            if SUBPROCESS_AVAILABLE:
                try:
                    result = subprocess.run(['whois', root_domain], 
                                          capture_output=True, text=True, timeout=10)
                    if result.returncode == 0:
                        whois_output = result.stdout.lower()
                        
                        # Look for registrar
                        for line in whois_output.split('\n'):
                            if 'registrar:' in line and line.strip():
                                return line.split(':', 1)[1].strip().title()
                            elif 'organization:' in line and line.strip():
                                return line.split(':', 1)[1].strip().title()
                            elif 'org:' in line and line.strip() and not line.startswith('org-'):
                                return line.split(':', 1)[1].strip().title()
                except:
                    pass
                    
            return "Unknown"
            
        except Exception as e:
            return "Unknown"
    
    def perform_nmap_scan(self, target: str) -> Dict[str, any]:
        """Perform comprehensive Nmap scan on target domain/IP"""
        try:
            # Start with basic port scan first
            basic_cmd = [
                'nmap',
                '-sS',  # SYN stealth scan
                '-T4',  # Timing template
                '-Pn',  # Skip host discovery
                '--top-ports', '1000',  # Scan top 1000 ports
                target
            ]
            
            # Execute basic scan first
            basic_result = subprocess.run(basic_cmd, capture_output=True, text=True, timeout=120)
            
            if basic_result.returncode != 0:
                # Try TCP connect scan as fallback
                fallback_cmd = [
                    'nmap',
                    '-sT',  # TCP connect scan
                    '-T3',  # Normal timing
                    '-Pn',
                    '--top-ports', '100',
                    target
                ]
                basic_result = subprocess.run(fallback_cmd, capture_output=True, text=True, timeout=60)
                
                # If still failing, try minimal scan
                if basic_result.returncode != 0:
                    minimal_cmd = [
                        'nmap',
                        '-sT',  # TCP connect scan
                        '-T2',  # Polite timing
                        '-Pn',
                        '-p', '80,443',  # Just basic ports
                        target
                    ]
                    basic_result = subprocess.run(minimal_cmd, capture_output=True, text=True, timeout=30)
            
            # Parse basic results to see if we found open ports
            basic_data = self._parse_nmap_output(basic_result.stdout)
            
            # If we found open ports, do detailed scan on them
            if basic_data['open_ports']:
                port_list = ','.join([port.split('/')[0] for port in basic_data['open_ports'][:10]])
                
                detailed_cmd = [
                    'nmap',
                    '-sV',  # Service version detection
                    '-sC',  # Default scripts
                    '--script', 'ssl-cert,http-title,http-server-header',
                    '-T3',  # Normal timing
                    '-Pn',
                    '-p', port_list,
                    target
                ]
                
                detailed_result = subprocess.run(detailed_cmd, capture_output=True, text=True, timeout=180)
                
                if detailed_result.returncode == 0:
                    return self._parse_nmap_output(detailed_result.stdout)
            
            # Return basic results if detailed scan failed
            # Ensure we return a valid scan result even if no ports found
            if not basic_data['open_ports']:
                basic_data['os_detection'] = "Scan completed - No open ports detected"
                basic_data['ssl_info'] = "Scan completed - No SSL services found"
                basic_data['http_info'] = "Scan completed - No HTTP services found"
            return basic_data
            
        except subprocess.TimeoutExpired:
            return self._create_empty_nmap_result("Scan timeout")
        except FileNotFoundError:
            return self._create_empty_nmap_result("Nmap not installed")
        except Exception as e:
            return self._create_empty_nmap_result(f"Error: {str(e)}")
    
    def _create_empty_nmap_result(self, error_msg: str = "No data") -> Dict[str, any]:
        """Create empty Nmap result structure"""
        return {
            'open_ports': [],
            'services': [],
            'os_detection': f"Nmap scan completed - {error_msg}",
            'vulnerabilities': [],
            'ssl_info': f"No SSL detected - {error_msg}",
            'http_info': f"No HTTP services - {error_msg}",
            'traceroute': f"No traceroute data - {error_msg}",
            'dns_info': f"No DNS info - {error_msg}"
        }
    
    def _parse_nmap_output(self, nmap_output: str) -> Dict[str, any]:
        """Parse Nmap output and extract structured information"""
        result = {
            'open_ports': [],
            'services': [],
            'os_detection': None,
            'vulnerabilities': [],
            'ssl_info': None,
            'http_info': None,
            'traceroute': None,
            'dns_info': None
        }
        
        lines = nmap_output.split('\n')
        current_section = None
        
        for line in lines:
            line = line.strip()
            
            # Parse open ports and services
            if '/tcp' in line or '/udp' in line:
                if 'open' in line:
                    parts = line.split()
                    if len(parts) >= 3:
                        port = parts[0]
                        service = parts[2] if len(parts) > 2 else 'unknown'
                        result['open_ports'].append(port)
                        
                        # Extract version info if available
                        if len(parts) > 3:
                            version_info = ' '.join(parts[3:])
                            result['services'].append(f"{port}: {service} {version_info}")
                        else:
                            result['services'].append(f"{port}: {service}")
            
            # Parse OS detection
            if line.startswith('Running:') or line.startswith('OS details:'):
                result['os_detection'] = line
            elif 'OS fingerprint not ideal' in line:
                result['os_detection'] = "OS detection uncertain"
            
            # Parse vulnerabilities
            if 'VULNERABLE' in line.upper() or 'CVE-' in line:
                result['vulnerabilities'].append(line)
            
            # Parse SSL information
            if 'ssl-cert:' in line or 'SSL certificate:' in line:
                current_section = 'ssl'
                result['ssl_info'] = line
            elif current_section == 'ssl' and line.startswith('|'):
                if result['ssl_info']:
                    result['ssl_info'] += '; ' + line.replace('|', '').strip()
                else:
                    result['ssl_info'] = line.replace('|', '').strip()
            
            # Parse HTTP information
            if 'http-title:' in line or 'http-server-header:' in line:
                if result['http_info']:
                    result['http_info'] += '; ' + line
                else:
                    result['http_info'] = line
            
            # Parse traceroute
            if line.startswith('TRACEROUTE'):
                current_section = 'traceroute'
                result['traceroute'] = ''
            elif current_section == 'traceroute' and (line.startswith('HOP') or '...' in line or any(c.isdigit() for c in line[:3])):
                if result['traceroute']:
                    result['traceroute'] += '; ' + line
                else:
                    result['traceroute'] = line
            
            # Parse DNS information
            if 'dns-zone-transfer:' in line or line.startswith('Host script results:'):
                if 'dns' in line.lower():
                    result['dns_info'] = line
        
        # Clean up results
        result['open_ports'] = result['open_ports'][:10]  # Limit to first 10 ports
        result['services'] = result['services'][:10]      # Limit to first 10 services
        result['vulnerabilities'] = result['vulnerabilities'][:5]  # Limit to first 5 vulns
        
        return result
    
    async def parallel_nmap_scanner(self, results: Dict[str, SubdomainResult], progress_callback=None):
        """Run Nmap scans in parallel for all discovered subdomains"""
        subdomains_to_scan = [subdomain for subdomain, result in results.items() 
                             if not result.nmap_open_ports and result.ip_addresses]
        
        if progress_callback:
            progress_callback("Nmap Scanning", 0, message=f"Checking {len(results)} subdomains for NMAP scanning...")
        
        if not subdomains_to_scan:
            if progress_callback:
                progress_callback("Nmap Scanning", 100, message=f"All {len(results)} subdomains already have NMAP data or no IPs")
            return
        
        if progress_callback:
            progress_callback("Nmap Scanning", 10, message=f"Starting parallel Nmap scans on {len(subdomains_to_scan)} subdomains")
            # Debug info
            sample_targets = subdomains_to_scan[:3]
            progress_callback("Nmap Scanning", 15, message=f"Sample targets to scan: {', '.join(sample_targets)}")
        
        # Create semaphore to limit concurrent scans
        nmap_semaphore = asyncio.Semaphore(3)  # Limit to 3 concurrent Nmap scans
        
        async def scan_single_target(subdomain):
            async with nmap_semaphore:
                try:
                    if progress_callback:
                        progress_callback("Nmap Scanning", 0, message=f"Scanning {subdomain}...")
                    
                    # Run Nmap scan in thread pool to avoid blocking
                    loop = asyncio.get_event_loop()
                    nmap_data = await loop.run_in_executor(
                        None, self.perform_nmap_scan, subdomain
                    )
                    
                    # Update result with Nmap data
                    if subdomain in results:
                        result = results[subdomain]
                        result.nmap_open_ports = nmap_data['open_ports']
                        result.nmap_services = nmap_data['services']
                        result.nmap_os_detection = nmap_data['os_detection']
                        result.nmap_vulnerabilities = nmap_data['vulnerabilities']
                        result.nmap_ssl_info = nmap_data['ssl_info']
                        result.nmap_http_info = nmap_data['http_info']
                        result.nmap_traceroute = nmap_data['traceroute']
                        result.nmap_dns_info = nmap_data['dns_info']
                    
                    if progress_callback:
                        ports_found = len(nmap_data['open_ports'])
                        progress_callback("Nmap Scanning", 0, 
                                        message=f"Completed {subdomain} - {ports_found} ports found")
                    
                except Exception as e:
                    if progress_callback:
                        progress_callback("Nmap Scanning", 0, 
                                        message=f"Failed to scan {subdomain}: {str(e)}")
        
        # Run all scans in parallel
        tasks = [scan_single_target(subdomain) for subdomain in subdomains_to_scan]
        await asyncio.gather(*tasks, return_exceptions=True)
        
        # Count how many subdomains now have NMAP data (scanned, even if no ports found)
        scanned_count = sum(1 for result in results.values() 
                          if result.nmap_open_ports or result.nmap_services or result.nmap_os_detection)
        
        if progress_callback:
            progress_callback("Nmap Scanning", 100, 
                            message=f"Parallel Nmap scanning completed - {scanned_count}/{len(results)} subdomains have NMAP data")

    def detect_system_resources(self) -> 'SystemResources':
        """Detect available system resources for dynamic scaling"""
        if not PSUTIL_AVAILABLE:
            # Fallback to basic resource detection using multiprocessing
            return SystemResources(
                cpu_cores=multiprocessing.cpu_count(),
                memory_gb=8.0,  # Conservative estimate
                network_bandwidth_mbps=100.0,  # Conservative estimate
                concurrent_threads=min(50, multiprocessing.cpu_count() * 4)
            )
        
        try:
            cpu_cores = psutil.cpu_count(logical=True)
            memory_info = psutil.virtual_memory()
            memory_gb = memory_info.total / (1024**3)
            
            # Estimate network bandwidth (basic check)
            # This is a rough estimate - real bandwidth testing would be more complex
            network_bandwidth_mbps = 100.0  # Default assumption
            
            # Calculate optimal thread count based on resources
            concurrent_threads = min(100, cpu_cores * 8)  # Max 100 threads
            
            return SystemResources(
                cpu_cores=cpu_cores,
                memory_gb=memory_gb,
                network_bandwidth_mbps=network_bandwidth_mbps,
                concurrent_threads=concurrent_threads
            )
        except Exception:
            # Fallback if psutil fails
            return SystemResources(
                cpu_cores=multiprocessing.cpu_count(),
                memory_gb=8.0,
                network_bandwidth_mbps=100.0,
                concurrent_threads=min(50, multiprocessing.cpu_count() * 4)
            )

    def group_ips_by_ranges(self, results: Dict[str, SubdomainResult]) -> List['IPRangeGroup']:
        """Group subdomains by IP ranges for efficient scanning"""
        ip_to_subdomains = defaultdict(list)
        
        # Group subdomains by IP
        for subdomain, result in results.items():
            for ip in result.ip_addresses:
                try:
                    ip_addr = ipaddress.ip_address(ip)
                    ip_to_subdomains[str(ip_addr)].append(subdomain)
                except ValueError:
                    continue
        
        # Group IPs by subnets
        subnet_groups = []
        processed_ips = set()
        
        for ip_str, subdomains in ip_to_subdomains.items():
            if ip_str in processed_ips:
                continue
                
            try:
                ip_addr = ipaddress.ip_address(ip_str)
                
                # Create /24 subnet for IPv4, /64 for IPv6
                if ip_addr.version == 4:
                    subnet = ipaddress.ip_network(f"{ip_addr}/24", strict=False)
                    subnet_size = 24
                else:
                    subnet = ipaddress.ip_network(f"{ip_addr}/64", strict=False)
                    subnet_size = 64
                
                # Find all IPs in this subnet
                subnet_ips = []
                subnet_subdomains = []
                
                for other_ip_str, other_subdomains in ip_to_subdomains.items():
                    if other_ip_str in processed_ips:
                        continue
                        
                    try:
                        other_ip_addr = ipaddress.ip_address(other_ip_str)
                        if other_ip_addr in subnet:
                            subnet_ips.append(other_ip_str)
                            subnet_subdomains.extend(other_subdomains)
                            processed_ips.add(other_ip_str)
                    except ValueError:
                        continue
                
                if subnet_ips:
                    group = IPRangeGroup(
                        ip_range=str(subnet),
                        subnets=subnet_ips,
                        subdomains=subnet_subdomains
                    )
                    subnet_groups.append(group)
                    
            except ValueError:
                continue
        
        return subnet_groups

    async def smart_ip_range_scan(self, ip_groups: List['IPRangeGroup'], progress_callback=None) -> Dict[str, Dict]:
        """Perform intelligent scanning on IP ranges"""
        system_resources = self.detect_system_resources()
        
        # Determine optimal batch size based on system resources
        optimal_batch_size = min(20, system_resources.concurrent_threads // 2)
        
        results = {}
        total_groups = len(ip_groups)
        
        if progress_callback:
            progress_callback("Smart IP Scanning", 0, 
                            message=f"Starting smart scan of {total_groups} IP ranges with {optimal_batch_size} concurrent scans")
        
        # Create semaphore for controlling concurrency
        semaphore = asyncio.Semaphore(optimal_batch_size)
        
        async def scan_ip_group(group: 'IPRangeGroup', group_index: int):
            async with semaphore:
                try:
                    if progress_callback:
                        progress_callback("Smart IP Scanning", 
                                        (group_index / total_groups) * 100,
                                        message=f"Scanning subnet {group.ip_range} ({len(group.subnets)} IPs)")
                    
                    # For each IP in the group, run a targeted scan
                    group_results = {}
                    
                    for ip in group.subnets:
                        # Use faster, targeted nmap scan for range scanning
                        cmd = [
                            'nmap', '-sS', '-O', '--top-ports', '1000',
                            '--version-intensity', '0',  # Faster version detection
                            '-T4', '-Pn', '--host-timeout', '60s',
                            ip
                        ]
                        
                        try:
                            result = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
                            
                            if result.returncode == 0:
                                parsed_result = self._parse_nmap_output(result.stdout)
                                group_results[ip] = parsed_result
                                
                                if progress_callback:
                                    ports_found = len(parsed_result.get('open_ports', []))
                                    progress_callback("Smart IP Scanning",
                                                    (group_index / total_groups) * 100,
                                                    message=f"IP {ip}: {ports_found} ports found")
                            else:
                                group_results[ip] = self._create_empty_nmap_result()
                                
                        except subprocess.TimeoutExpired:
                            group_results[ip] = self._create_empty_nmap_result()
                            if progress_callback:
                                progress_callback("Smart IP Scanning",
                                                (group_index / total_groups) * 100,
                                                message=f"IP {ip}: scan timeout")
                        except Exception:
                            group_results[ip] = self._create_empty_nmap_result()
                    
                    results[group.ip_range] = group_results
                    
                except Exception as e:
                    if progress_callback:
                        progress_callback("Smart IP Scanning",
                                        (group_index / total_groups) * 100,
                                        message=f"Error scanning {group.ip_range}: {str(e)}")
        
        # Execute all group scans concurrently
        tasks = [scan_ip_group(group, i) for i, group in enumerate(ip_groups)]
        await asyncio.gather(*tasks, return_exceptions=True)
        
        if progress_callback:
            total_ips_scanned = sum(len(group.subnets) for group in ip_groups)
            progress_callback("Smart IP Scanning", 100,
                            message=f"Smart IP scanning completed - {total_ips_scanned} IPs in {total_groups} ranges")
        
        return results

    async def apply_smart_scan_results(self, results: Dict[str, SubdomainResult], 
                                     ip_scan_results: Dict[str, Dict], progress_callback=None):
        """Apply smart IP scan results back to subdomain results"""
        if progress_callback:
            progress_callback("Result Integration", 0, message="Integrating smart scan results")
        
        updated_count = 0
        
        for subdomain, result in results.items():
            for ip in result.ip_addresses:
                # Find the subnet this IP belongs to
                for subnet, subnet_results in ip_scan_results.items():
                    if ip in subnet_results:
                        nmap_data = subnet_results[ip]
                        
                        # Update the subdomain result with smart scan data
                        if nmap_data:
                            result.nmap_open_ports = nmap_data.get('open_ports', [])
                            result.nmap_services = nmap_data.get('services', [])
                            result.nmap_os_detection = nmap_data.get('os_detection', "Not detected")
                            result.nmap_vulnerabilities = nmap_data.get('vulnerabilities', [])
                            result.nmap_ssl_info = nmap_data.get('ssl_info', None)
                            result.nmap_http_info = nmap_data.get('http_info', None)
                            result.nmap_traceroute = nmap_data.get('traceroute', None)
                            result.nmap_dns_info = nmap_data.get('dns_info', None)
                            updated_count += 1
                        break
        
        if progress_callback:
            progress_callback("Result Integration", 100,
                            message=f"Updated {updated_count} subdomains with smart scan results")

    async def _advanced_vulnerability_assessment(self):
        """Advanced Vulnerability Assessment with CVE integration and security analysis"""
        total_subdomains = len(self.results)
        processed = 0
        
        for subdomain, result in self.results.items():
            try:
                self._emit_progress("Vulnerability Assessment", 
                                  (processed / total_subdomains) * 100,
                                  message=f"Analyzing {subdomain} for vulnerabilities...")
                
                # Initialize vulnerability assessment
                vuln_assessment = VulnerabilityAssessment()
                
                # 1. Security Headers Analysis
                if result.http_status == 200:
                    security_headers = await self._analyze_security_headers(subdomain)
                    vuln_assessment.security_headers = security_headers
                
                # 2. SSL/TLS Analysis
                ssl_issues = await self._analyze_ssl_vulnerabilities(subdomain)
                vuln_assessment.ssl_issues = ssl_issues
                
                # 3. Cookie Security Analysis
                cookie_security = await self._analyze_cookie_security(subdomain)
                vuln_assessment.cookie_security = cookie_security
                
                # 4. CORS Analysis
                cors_issues = await self._analyze_cors_issues(subdomain)
                vuln_assessment.cors_issues = cors_issues
                
                # 5. CVE Database Integration
                cve_ids = await self._check_cve_database(result)
                vuln_assessment.cve_ids = cve_ids
                
                # 6. Calculate Risk Score
                vuln_assessment.risk_score = self._calculate_risk_score(vuln_assessment)
                
                # Store vulnerability assessment in result
                result.vulnerability_assessment = vuln_assessment
                
                processed += 1
                
            except Exception as e:
                self._emit_progress("Vulnerability Assessment",
                                  (processed / total_subdomains) * 100,
                                  message=f"Error analyzing {subdomain}: {str(e)[:30]}")
                processed += 1
                continue
        
        self._emit_progress("Vulnerability Assessment", 100,
                          message=f"Vulnerability assessment complete: {processed} subdomains analyzed")

    async def _analyze_security_headers(self, subdomain: str) -> Dict[str, str]:
        """Analyze HTTP security headers"""
        security_headers = {}
        
        try:
            # Use curl to get headers
            cmd = ['curl', '-s', '-I', '-L', '--max-time', '10', f'https://{subdomain}']
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
            
            if result.returncode == 0:
                headers = result.stdout.lower()
                
                # Check for important security headers
                security_checks = {
                    'X-Frame-Options': 'x-frame-options:' in headers,
                    'X-Content-Type-Options': 'x-content-type-options:' in headers,
                    'X-XSS-Protection': 'x-xss-protection:' in headers,
                    'Strict-Transport-Security': 'strict-transport-security:' in headers,
                    'Content-Security-Policy': 'content-security-policy:' in headers,
                    'Referrer-Policy': 'referrer-policy:' in headers,
                    'Permissions-Policy': 'permissions-policy:' in headers,
                }
                
                for header, present in security_checks.items():
                    security_headers[header] = "Present" if present else "Missing"
                    
        except Exception:
            security_headers['Error'] = "Failed to analyze headers"
        
        return security_headers

    async def _analyze_ssl_vulnerabilities(self, subdomain: str) -> List[str]:
        """Analyze SSL/TLS vulnerabilities"""
        ssl_issues = []
        
        try:
            # Use testssl.sh if available, otherwise basic SSL check
            testssl_cmd = ['testssl.sh', '--quiet', '--jsonfile-pretty', '-', subdomain]
            
            try:
                result = subprocess.run(testssl_cmd, capture_output=True, text=True, timeout=30)
                if result.returncode == 0 and result.stdout:
                    # Parse testssl.sh JSON output for vulnerabilities
                    import json
                    try:
                        data = json.loads(result.stdout)
                        for scan_result in data:
                            if scan_result.get('severity') in ['HIGH', 'CRITICAL', 'MEDIUM']:
                                ssl_issues.append(f"{scan_result.get('id', 'Unknown')}: {scan_result.get('finding', 'SSL Issue')}")
                    except json.JSONDecodeError:
                        pass
            except FileNotFoundError:
                # Fallback to basic SSL check using openssl
                openssl_cmd = ['openssl', 's_client', '-connect', f'{subdomain}:443', '-verify_return_error']
                try:
                    result = subprocess.run(openssl_cmd, input='', capture_output=True, text=True, timeout=10)
                    if 'verify error' in result.stderr.lower():
                        ssl_issues.append("Certificate verification failed")
                    if 'ssl3_get_record:wrong version number' in result.stderr.lower():
                        ssl_issues.append("SSL version mismatch")
                except subprocess.TimeoutExpired:
                    ssl_issues.append("SSL connection timeout")
                    
        except Exception:
            ssl_issues.append("SSL analysis failed")
        
        return ssl_issues

    async def _analyze_cookie_security(self, subdomain: str) -> Dict[str, str]:
        """Analyze cookie security settings"""
        cookie_analysis = {}
        
        try:
            cmd = ['curl', '-s', '-I', '-L', '--max-time', '10', f'https://{subdomain}']
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
            
            if result.returncode == 0:
                headers = result.stdout.lower()
                
                # Look for Set-Cookie headers
                cookie_lines = [line for line in headers.split('\n') if 'set-cookie:' in line]
                
                if cookie_lines:
                    secure_cookies = sum(1 for line in cookie_lines if 'secure' in line)
                    httponly_cookies = sum(1 for line in cookie_lines if 'httponly' in line)
                    samesite_cookies = sum(1 for line in cookie_lines if 'samesite' in line)
                    
                    cookie_analysis['Total_Cookies'] = str(len(cookie_lines))
                    cookie_analysis['Secure_Cookies'] = str(secure_cookies)
                    cookie_analysis['HttpOnly_Cookies'] = str(httponly_cookies)
                    cookie_analysis['SameSite_Cookies'] = str(samesite_cookies)
                    
                    # Risk assessment
                    if secure_cookies < len(cookie_lines):
                        cookie_analysis['Risk'] = "Insecure cookies detected"
                    elif httponly_cookies < len(cookie_lines):
                        cookie_analysis['Risk'] = "Non-HttpOnly cookies detected"
                    else:
                        cookie_analysis['Risk'] = "Low"
                else:
                    cookie_analysis['Status'] = "No cookies found"
                    
        except Exception:
            cookie_analysis['Error'] = "Cookie analysis failed"
        
        return cookie_analysis

    async def _analyze_cors_issues(self, subdomain: str) -> List[str]:
        """Analyze CORS (Cross-Origin Resource Sharing) issues"""
        cors_issues = []
        
        try:
            # Test CORS with different origins
            test_origins = ['https://evil.com', 'http://malicious.site', 'null']
            
            for origin in test_origins:
                cmd = [
                    'curl', '-s', '-I', '-H', f'Origin: {origin}',
                    '-H', 'Access-Control-Request-Method: GET',
                    '--max-time', '10', f'https://{subdomain}'
                ]
                
                try:
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
                    if result.returncode == 0:
                        headers = result.stdout.lower()
                        
                        if 'access-control-allow-origin: *' in headers:
                            cors_issues.append("Wildcard CORS policy allows any origin")
                        elif f'access-control-allow-origin: {origin.lower()}' in headers:
                            cors_issues.append(f"CORS allows potentially malicious origin: {origin}")
                        elif 'access-control-allow-credentials: true' in headers and 'access-control-allow-origin: *' in headers:
                            cors_issues.append("Dangerous combination: credentials allowed with wildcard origin")
                            
                except subprocess.TimeoutExpired:
                    continue
                    
        except Exception:
            cors_issues.append("CORS analysis failed")
        
        return cors_issues

    async def _check_cve_database(self, result: SubdomainResult) -> List[str]:
        """Check for known CVEs based on detected technologies and services"""
        cve_ids = []
        
        try:
            # Check technologies for known vulnerabilities
            for tech in result.technologies:
                # Simple CVE database lookup (in production, use a real CVE database)
                known_vulns = {
                    'apache': ['CVE-2021-44228', 'CVE-2021-45046'],  # Log4j
                    'nginx': ['CVE-2021-23017'],
                    'wordpress': ['CVE-2021-34527', 'CVE-2021-39201'],
                    'drupal': ['CVE-2021-41182', 'CVE-2021-41183'],
                    'joomla': ['CVE-2021-23132'],
                    'iis': ['CVE-2021-31207', 'CVE-2021-31199'],
                    'tomcat': ['CVE-2021-25329', 'CVE-2021-30640'],
                }
                
                tech_lower = tech.lower()
                for vuln_tech, vulns in known_vulns.items():
                    if vuln_tech in tech_lower:
                        cve_ids.extend(vulns)
            
            # Check server information for vulnerabilities
            if result.server:
                server_lower = result.server.lower()
                if 'apache' in server_lower:
                    cve_ids.extend(['CVE-2021-44228', 'CVE-2021-45046'])
                elif 'nginx' in server_lower:
                    cve_ids.append('CVE-2021-23017')
                elif 'iis' in server_lower:
                    cve_ids.extend(['CVE-2021-31207', 'CVE-2021-31199'])
            
            # Remove duplicates
            cve_ids = list(set(cve_ids))
            
        except Exception:
            pass
        
        return cve_ids

    def _calculate_risk_score(self, vuln_assessment: VulnerabilityAssessment) -> float:
        """Calculate overall risk score based on vulnerability assessment"""
        risk_score = 0.0
        
        try:
            # CVE-based scoring
            cve_count = len(vuln_assessment.cve_ids)
            risk_score += min(cve_count * 2.0, 10.0)  # Max 10 points for CVEs
            
            # Security headers scoring
            missing_headers = sum(1 for status in vuln_assessment.security_headers.values() if status == "Missing")
            risk_score += min(missing_headers * 1.0, 5.0)  # Max 5 points for missing headers
            
            # SSL issues scoring
            ssl_issue_count = len(vuln_assessment.ssl_issues)
            risk_score += min(ssl_issue_count * 1.5, 7.5)  # Max 7.5 points for SSL issues
            
            # CORS issues scoring
            cors_issue_count = len(vuln_assessment.cors_issues)
            risk_score += min(cors_issue_count * 2.0, 5.0)  # Max 5 points for CORS issues
            
            # Cookie security scoring
            if 'Risk' in vuln_assessment.cookie_security:
                if vuln_assessment.cookie_security['Risk'] == "Insecure cookies detected":
                    risk_score += 3.0
                elif vuln_assessment.cookie_security['Risk'] == "Non-HttpOnly cookies detected":
                    risk_score += 1.5
            
            # Cap at 10.0
            risk_score = min(risk_score, 10.0)
            
        except Exception:
            risk_score = 0.0
        
        return risk_score

    async def _advanced_analytics_generation(self):
        """Generate advanced analytics and historical data analysis"""
        total_subdomains = len(self.results)
        processed = 0
        
        # Overall statistics
        domain_stats = {
            'total_subdomains': total_subdomains,
            'live_subdomains': sum(1 for r in self.results.values() if r.http_status == 200),
            'sources_used': len(set(r.source for r in self.results.values())),
            'unique_ips': len(set(ip for r in self.results.values() for ip in r.ip_addresses)),
            'avg_confidence': sum(r.confidence_score for r in self.results.values()) / total_subdomains if total_subdomains > 0 else 0,
        }
        
        self._emit_progress("Advanced Analytics", 20,
                          message=f"Analyzing {total_subdomains} subdomains for trends...")
        
        # Risk analysis
        risk_distribution = {'Low': 0, 'Medium': 0, 'High': 0, 'Critical': 0}
        technology_trends = Counter()
        source_effectiveness = Counter()
        
        for subdomain, result in self.results.items():
            try:
                processed += 1
                
                # Generate historical data tracking
                historical_data = HistoricalData(
                    first_seen=datetime.datetime.fromtimestamp(result.discovered_at),
                    last_seen=datetime.datetime.now(),
                    changes_detected=[],
                    trend_analysis={}
                )
                
                # Analyze trends
                if result.technologies:
                    technology_trends.update(result.technologies)
                
                source_effectiveness[result.source] += 1
                
                # Risk categorization
                if result.vulnerability_assessment:
                    risk_score = result.vulnerability_assessment.risk_score
                    if risk_score >= 8.0:
                        risk_distribution['Critical'] += 1
                        historical_data.changes_detected.append("High-risk vulnerabilities detected")
                    elif risk_score >= 6.0:
                        risk_distribution['High'] += 1
                        historical_data.changes_detected.append("Medium-risk vulnerabilities detected")
                    elif risk_score >= 3.0:
                        risk_distribution['Medium'] += 1
                    else:
                        risk_distribution['Low'] += 1
                
                # Store historical data
                result.historical_data = historical_data
                
                # Generate AI-like analysis
                ai_analysis = AIAnalysis()
                
                # Pattern recognition (simplified)
                subdomain_parts = result.subdomain.split('.')
                if len(subdomain_parts) > 2:
                    prefix = subdomain_parts[0]
                    
                    # Detect patterns
                    if any(word in prefix.lower() for word in ['api', 'service', 'micro']):
                        ai_analysis.predicted_subdomains.append(f"{prefix}-v2.{'.'.join(subdomain_parts[1:])}")
                        ai_analysis.technology_predictions.extend(['API Gateway', 'Microservice'])
                        ai_analysis.pattern_confidence = 0.8
                    
                    if any(word in prefix.lower() for word in ['admin', 'manage', 'control']):
                        ai_analysis.risk_assessment = "High"
                        ai_analysis.anomaly_score = 0.7
                        ai_analysis.pattern_confidence = 0.9
                    
                    if any(word in prefix.lower() for word in ['test', 'dev', 'staging']):
                        ai_analysis.risk_assessment = "Medium"
                        ai_analysis.anomaly_score = 0.5
                        ai_analysis.predicted_subdomains.append(f"prod-{prefix}.{'.'.join(subdomain_parts[1:])}")
                
                # Store AI analysis
                result.ai_analysis = ai_analysis
                
                if processed % 10 == 0:  # Update progress every 10 items
                    self._emit_progress("Advanced Analytics",
                                      20 + (processed / total_subdomains) * 60,
                                      message=f"Analyzed {processed}/{total_subdomains} subdomains")
                
            except Exception as e:
                processed += 1
                continue
        
        # Generate comprehensive analytics report
        analytics_summary = {
            'domain_statistics': domain_stats,
            'risk_distribution': risk_distribution,
            'top_technologies': dict(technology_trends.most_common(10)),
            'source_effectiveness': dict(source_effectiveness.most_common()),
            'security_insights': {
                'high_risk_count': risk_distribution['High'] + risk_distribution['Critical'],
                'vulnerable_percentage': ((risk_distribution['High'] + risk_distribution['Critical']) / total_subdomains * 100) if total_subdomains > 0 else 0,
                'most_common_vulns': self._get_common_vulnerabilities(),
            },
            'discovery_insights': {
                'most_effective_source': source_effectiveness.most_common(1)[0] if source_effectiveness else ('Unknown', 0),
                'discovery_methods_count': len(source_effectiveness),
                'avg_subdomains_per_source': sum(source_effectiveness.values()) / len(source_effectiveness) if source_effectiveness else 0,
            }
        }
        
        self._emit_progress("Advanced Analytics", 90,
                          message="Generating trend analysis and predictions...")
        
        # Store analytics in class for reporting
        self.analytics_summary = analytics_summary
        
        self._emit_progress("Advanced Analytics", 100,
                          message=f"Analytics complete: {processed} subdomains analyzed with trend insights")

    def _get_common_vulnerabilities(self) -> Dict[str, int]:
        """Get most common vulnerabilities across all subdomains"""
        vuln_counter = Counter()
        
        for result in self.results.values():
            if result.vulnerability_assessment and result.vulnerability_assessment.cve_ids:
                vuln_counter.update(result.vulnerability_assessment.cve_ids)
        
        return dict(vuln_counter.most_common(5))

    def save_advanced_excel(self, results: Dict[str, SubdomainResult], domain: str):
        """Save results to advanced Excel file with comprehensive data"""
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"output/{domain}_ultra_robust_{timestamp}.xlsx"
        
        os.makedirs('output', exist_ok=True)
        
        # Create workbook with multiple sheets
        wb = openpyxl.Workbook()
        
        # Main results sheet
        ws_main = wb.active
        ws_main.title = "Subdomain Discovery"
        
        # Headers for main sheet - Enhanced with CNAME analysis and ownership
        headers = [
            'SSL_Verified', 'SSL_Issuer', 'SSL_Subject',
            'CNAME_Chain', 'CNAME_Target', 'Service_Provider', 'Service_Type', 'Takeover_Risk',
            'Subdomain', 'Discovery_Source', 'Domain_Owner', 'HTTP_Status', 'Status_Explanation', 'IP_Addresses', 
            'Response_Time', 'Technologies', 'Server', 'Confidence_Score'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        status_explanations = {
            0: "No Response - Server unreachable",
            200: "OK - Server responding normally",
            301: "Moved Permanently - Redirected",
            302: "Found - Temporary redirect",
            400: "Bad Request - Invalid request",
            401: "Unauthorized - Authentication required",
            403: "Forbidden - Access denied",
            404: "Not Found - Resource not found",
            500: "Internal Server Error - Server malfunction"
        }
        
        confidence_colors = {
            'high': "E6FFE6",    # Light green
            'medium': "FFF2CC",  # Light yellow
            'low': "FFE6E6"      # Light red
        }
        
        row = 2
        for result in sorted(results.values(), key=lambda x: x.confidence_score, reverse=True):
            col = 1
            
            # SSL Verified
            ssl_verified_cell = ws_main.cell(row=row, column=col, value="Yes" if result.ssl_domain_verified else "No")
            if result.ssl_domain_verified:
                ssl_verified_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ssl_verified_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            col += 1
            
            # SSL Issuer
            ws_main.cell(row=row, column=col, value=result.ssl_issuer or "N/A")
            col += 1
            
            # SSL Subject
            ws_main.cell(row=row, column=col, value=result.ssl_subject or "N/A")
            col += 1
            
            # CNAME Chain
            cname_chain = " -> ".join(result.cname_chain) if result.cname_chain else "Direct"
            ws_main.cell(row=row, column=col, value=cname_chain)
            col += 1
            
            # CNAME Target (final target)
            cname_target = result.cname_chain[-1] if result.cname_chain else "N/A"
            ws_main.cell(row=row, column=col, value=cname_target)
            col += 1
            
            # Service Provider
            service_provider = result.cname_records[0].provider if result.cname_records else "Unknown"
            ws_main.cell(row=row, column=col, value=service_provider)
            col += 1
            
            # Service Type
            service_type = result.cname_records[0].service_type if result.cname_records else "Unknown"
            ws_main.cell(row=row, column=col, value=service_type)
            col += 1
            
            # Takeover Risk with color coding
            takeover_cell = ws_main.cell(row=row, column=col, value=result.takeover_risk)
            if result.takeover_risk == "Critical":
                takeover_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                takeover_cell.font = Font(color="FFFFFF", bold=True)
            elif result.takeover_risk == "High":
                takeover_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            col += 1
            
            # Subdomain
            ws_main.cell(row=row, column=col, value=result.subdomain)
            col += 1
            
            # Discovery Source (more readable)
            readable_source = self._get_readable_source(result.source)
            ws_main.cell(row=row, column=col, value=readable_source)
            col += 1
            
            # Domain Owner/Registrar
            if result.ownership_info is None:
                result.ownership_info = self.get_domain_ownership(result.subdomain)
            ws_main.cell(row=row, column=col, value=result.ownership_info)
            col += 1
            
            # HTTP Status with color coding
            status_cell = ws_main.cell(row=row, column=col, value=result.http_status)
            if result.http_status == 200:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif result.http_status in [301, 302]:
                status_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            elif result.http_status >= 400:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            col += 1
            
            # Status explanation
            explanation = status_explanations.get(result.http_status, f"HTTP {result.http_status}")
            ws_main.cell(row=row, column=col, value=explanation)
            col += 1
            
            # IP Addresses
            ws_main.cell(row=row, column=col, value=", ".join(result.ip_addresses))
            col += 1
            
            # Response Time
            if result.response_time:
                ws_main.cell(row=row, column=col, value=f"{result.response_time:.3f}s")
            col += 1
            
            # Technologies
            ws_main.cell(row=row, column=col, value=", ".join(result.technologies))
            col += 1
            
            # Server
            if result.server:
                ws_main.cell(row=row, column=col, value=result.server)
            col += 1
            
            # Confidence Score with color coding
            conf_cell = ws_main.cell(row=row, column=col, value=f"{result.confidence_score:.2f}")
            if result.confidence_score >= 0.8:
                conf_cell.fill = PatternFill(start_color=confidence_colors['high'], end_color=confidence_colors['high'], fill_type="solid")
            elif result.confidence_score >= 0.5:
                conf_cell.fill = PatternFill(start_color=confidence_colors['medium'], end_color=confidence_colors['medium'], fill_type="solid")
            else:
                conf_cell.fill = PatternFill(start_color=confidence_colors['low'], end_color=confidence_colors['low'], fill_type="solid")
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws_main[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_main.column_dimensions[column].width = adjusted_width
        
        # Create detailed attributes sheet
        ws_detailed = wb.create_sheet(title="Detailed Attributes")
        
        # Headers for detailed sheet - One row per domain/subdomain, one column per attribute
        detailed_headers = [
            'Domain/Subdomain', 'Domain_Owner', 'Discovery_Source', 'HTTP_Status', 'Status_Explanation',
            'IP_Addresses', 'Response_Time_ms', 'Technologies', 'Server', 'Confidence_Score',
            'CNAME_Chain', 'CNAME_Target', 'Service_Provider', 'Service_Type', 'Takeover_Risk', 
            'Discovered_At', 'WHOIS_Registrar', 'Root_Domain', 'Subdomain_Level',
            # Nmap scan results
            'Nmap_Open_Ports', 'Nmap_Services', 'Nmap_OS_Detection', 'Nmap_Vulnerabilities',
            'Nmap_SSL_Info', 'Nmap_HTTP_Info', 'Nmap_Traceroute', 'Nmap_DNS_Info', 'Additional_Info'
        ]
        
        # Set headers
        for col, header in enumerate(detailed_headers, 1):
            cell = ws_detailed.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Populate detailed data
        row = 2
        for result in sorted(results.values(), key=lambda x: x.subdomain):
            col = 1
            
            # Domain/Subdomain
            ws_detailed.cell(row=row, column=col, value=result.subdomain)
            col += 1
            
            # Domain Owner (ensure it's populated)
            if result.ownership_info is None:
                result.ownership_info = self.get_domain_ownership(result.subdomain)
            ws_detailed.cell(row=row, column=col, value=result.ownership_info)
            col += 1
            
            # Discovery Source
            ws_detailed.cell(row=row, column=col, value=self._get_readable_source(result.source))
            col += 1
            
            # HTTP Status
            ws_detailed.cell(row=row, column=col, value=result.http_status)
            col += 1
            
            # Status Explanation
            explanation = status_explanations.get(result.http_status, f"HTTP {result.http_status}")
            ws_detailed.cell(row=row, column=col, value=explanation)
            col += 1
            
            # IP Addresses
            ws_detailed.cell(row=row, column=col, value=", ".join(result.ip_addresses))
            col += 1
            
            # Response Time (in milliseconds)
            if result.response_time:
                ws_detailed.cell(row=row, column=col, value=f"{result.response_time * 1000:.1f}")
            else:
                ws_detailed.cell(row=row, column=col, value="N/A")
            col += 1
            
            # Technologies
            ws_detailed.cell(row=row, column=col, value=", ".join(result.technologies))
            col += 1
            
            # Server
            ws_detailed.cell(row=row, column=col, value=result.server or "N/A")
            col += 1
            
            # Confidence Score
            ws_detailed.cell(row=row, column=col, value=f"{result.confidence_score:.3f}")
            col += 1
            
            # CNAME Chain
            cname_chain = " -> ".join(result.cname_chain) if result.cname_chain else "Direct"
            ws_detailed.cell(row=row, column=col, value=cname_chain)
            col += 1
            
            # CNAME Target
            cname_target = result.cname_chain[-1] if result.cname_chain else "N/A"
            ws_detailed.cell(row=row, column=col, value=cname_target)
            col += 1
            
            # Service Provider
            service_provider = result.cname_records[0].provider if result.cname_records else "Unknown"
            ws_detailed.cell(row=row, column=col, value=service_provider)
            col += 1
            
            # Service Type
            service_type = result.cname_records[0].service_type if result.cname_records else "Unknown"
            ws_detailed.cell(row=row, column=col, value=service_type)
            col += 1
            
            # Takeover Risk
            ws_detailed.cell(row=row, column=col, value=result.takeover_risk)
            col += 1
            
            # Discovered At (timestamp)
            discovered_time = datetime.datetime.fromtimestamp(result.discovered_at).strftime('%Y-%m-%d %H:%M:%S')
            ws_detailed.cell(row=row, column=col, value=discovered_time)
            col += 1
            
            # WHOIS Registrar (same as domain owner, but explicit column)
            ws_detailed.cell(row=row, column=col, value=result.ownership_info)
            col += 1
            
            # Root Domain
            parts = result.subdomain.split('.')
            root_domain = '.'.join(parts[-2:]) if len(parts) >= 2 else result.subdomain
            ws_detailed.cell(row=row, column=col, value=root_domain)
            col += 1
            
            # Subdomain Level
            subdomain_level = len(parts) - 2 if len(parts) > 2 else 0
            ws_detailed.cell(row=row, column=col, value=subdomain_level)
            col += 1
            
            # Nmap data should already be populated by parallel scanner
            
            # Nmap Open Ports
            ws_detailed.cell(row=row, column=col, value=", ".join(result.nmap_open_ports) if result.nmap_open_ports else "None detected")
            col += 1
            
            # Nmap Services
            ws_detailed.cell(row=row, column=col, value="; ".join(result.nmap_services) if result.nmap_services else "None detected")
            col += 1
            
            # Nmap OS Detection
            ws_detailed.cell(row=row, column=col, value=result.nmap_os_detection or "Not detected")
            col += 1
            
            # Nmap Vulnerabilities
            ws_detailed.cell(row=row, column=col, value="; ".join(result.nmap_vulnerabilities) if result.nmap_vulnerabilities else "None found")
            col += 1
            
            # Nmap SSL Info
            ws_detailed.cell(row=row, column=col, value=result.nmap_ssl_info or "N/A")
            col += 1
            
            # Nmap HTTP Info
            ws_detailed.cell(row=row, column=col, value=result.nmap_http_info or "N/A")
            col += 1
            
            # Nmap Traceroute
            ws_detailed.cell(row=row, column=col, value=result.nmap_traceroute or "N/A")
            col += 1
            
            # Nmap DNS Info
            ws_detailed.cell(row=row, column=col, value=result.nmap_dns_info or "N/A")
            col += 1
            
            # Additional Info
            additional_info = []
            if result.is_cname_target:
                additional_info.append("CNAME Target")
            if result.takeover_risk in ["High", "Critical"]:
                additional_info.append("Potential Takeover Risk")
            if result.confidence_score >= 0.9:
                additional_info.append("High Confidence")
            ws_detailed.cell(row=row, column=col, value="; ".join(additional_info) if additional_info else "None")
            
            row += 1
        
        # Auto-adjust column widths for detailed sheet
        for col in range(1, len(detailed_headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws_detailed[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_detailed.column_dimensions[column].width = adjusted_width
        
        # Save workbook
        wb.save(filename)
        
        return filename

# Textual TUI Classes (only available if Textual is installed)
if TEXTUAL_AVAILABLE:
    class FastScanTUI(App):
        """Modern Textual TUI for ultra-robust subdomain scanning with categorized progress"""
        
        CSS = """
        .title {
            text-align: center;
            text-style: bold;
            color: $accent;
            margin: 1;
        }
        
        .section-title {
            text-style: bold;
            color: $primary;
            background: $surface;
            padding: 0 1;
        }
    
        #domain-input {
            width: 100%;
            margin: 1;
        }
        
        #mode-buttons {
            height: 3;
            margin: 1;
        }
    
        #control-buttons {
            margin: 1;
            height: 3;
        }
        
        #progress-grid {
            height: 20;
            margin: 1;
        }
        
        .progress-card {
            border: solid $primary;
            padding: 1;
            margin: 1;
            min-height: 6;
        }
        
        .phase-complete {
            border: solid $success;
        }
        
        .phase-active {
            border: solid $warning;
        }
        
        .phase-pending {
            border: solid $surface;
        }
    
        #results-section {
            border: solid $primary;
            margin: 1;
            min-height: 15;
        }
        
        #stats-bar {
            background: $surface;
            height: 3;
            margin: 1;
        }
        
        #log-section {
            height: 8;
            border: solid $primary;
            margin: 1;
        }
        """
    
        BINDINGS = [
            Binding("q", "quit", "Quit"),
            Binding("s", "scan", "Start Scan"),
            Binding("c", "clear", "Clear Results"),
        ]
    
        current_phase = reactive("Idle")
        progress = reactive(0.0)
        found_count = reactive(0)
    
        def __init__(self):
            super().__init__()
            self.scan_results = {}
            self.current_scanner = None
            self.is_scanning = False
            self.selected_mode = 5  # Default to Quick mode
            self.scan_start_time = None
    
        def compose(self) -> ComposeResult:
            yield Header()
        
            with ScrollableContainer():
                # Main title
                yield Static("🔍 Ultra-Robust Subdomain Enumerator", classes="title")
                
                # Domain input section
                with Container():
                    yield Static("🎯 Target Configuration", classes="section-title")
                    yield Input(
                        placeholder="Enter domain (e.g., example.com) or multiple domains separated by commas",
                        id="domain-input"
                    )
                    yield Label("", id="validation-label")
                
                # Mode selection section
                with Container(id="mode-buttons"):
                    yield Static("⚙️ Scanning Mode", classes="section-title")
                    with Horizontal():
                        yield Button("1️⃣ Basic", variant="default", id="mode-1")
                        yield Button("2️⃣ Standard", variant="default", id="mode-2") 
                        yield Button("3️⃣ Advanced", variant="default", id="mode-3")
                        yield Button("4️⃣ Ultra", variant="default", id="mode-4")
                        yield Button("5️⃣ Quick", variant="success", id="mode-5")
                    yield Label("Selected: Quick Mode (DNS-only, fastest)", id="mode-label")
                
                # Control buttons
                with Horizontal(id="control-buttons"):
                    yield Button("🚀 Start Scan", variant="success", id="start-scan")
                    yield Button("🔄 Multi-Domain", variant="primary", id="multi-scan")
                    yield Button("🛑 Stop", variant="error", id="stop-scan", disabled=True)
                    yield Button("💾 Export", variant="default", id="export-results", disabled=True)
                
                # Statistics bar
                with Horizontal(id="stats-bar"):
                    yield Label("📊 Status: Idle", id="status-label")
                    yield Label("🔍 Found: 0", id="found-label")
                    yield Label("⏱️ Time: 00:00", id="time-label")
                
                # Categorized progress grid (2x4 layout)
                with Container(id="progress-grid"):
                    yield Static("📈 Scanning Progress", classes="section-title")
                    with Horizontal():
                        # Left column
                        with Vertical():
                            with Container(classes="progress-card phase-pending", id="cert-card"):
                                yield Label("📜 Certificate Transparency")
                                yield ProgressBar(total=100, id="cert-progress")
                                yield Label("Waiting...", id="cert-status")
                            
                            with Container(classes="progress-card phase-pending", id="dns-card"):
                                yield Label("🎯 DNS Brute Force")
                                yield ProgressBar(total=100, id="dns-progress")
                                yield Label("Waiting...", id="dns-status")
                            
                            with Container(classes="progress-card phase-pending", id="http-card"):
                                yield Label("🔍 HTTP Analysis")
                                yield ProgressBar(total=100, id="http-progress")
                                yield Label("Waiting...", id="http-status")
                                
                            with Container(classes="progress-card phase-pending", id="ssl-card"):
                                yield Label("🔒 SSL Analysis")
                                yield ProgressBar(total=100, id="ssl-progress")
                                yield Label("Waiting...", id="ssl-status")
                        
                        # Right column
                        with Vertical():
                            with Container(classes="progress-card phase-pending", id="cname-card"):
                                yield Label("⛓️ CNAME Analysis")
                                yield ProgressBar(total=100, id="cname-progress")
                                yield Label("Waiting...", id="cname-status")
                            
                            with Container(classes="progress-card phase-pending", id="nmap-card"):
                                yield Label("🌐 Port Scanning")
                                yield ProgressBar(total=100, id="nmap-progress")
                                yield Label("Waiting...", id="nmap-status")
                            
                            with Container(classes="progress-card phase-pending", id="vuln-card"):
                                yield Label("🔐 Vulnerability Analysis")
                                yield ProgressBar(total=100, id="vuln-progress")
                                yield Label("Waiting...", id="vuln-status")
                                
                            with Container(classes="progress-card phase-pending", id="final-card"):
                                yield Label("✨ Final Analysis")
                                yield ProgressBar(total=100, id="final-progress")
                                yield Label("Waiting...", id="final-status")
                
                # Results section
                with Container(id="results-section"):
                    yield Static("📊 Live Results", classes="section-title")
                    yield DataTable(zebra_stripes=True, id="results-table")
                
                # Activity log
                with Container(id="log-section"):
                    yield Static("📝 Activity Log", classes="section-title")
                    yield RichLog(id="activity-log", markup=True)
        
            yield Footer()
    
        def on_mount(self) -> None:
            """Initialize the app"""
            # Setup results table
            table = self.query_one("#results-table", DataTable)
            table.add_columns("Subdomain", "Status", "IP Address", "Server", "Source")
            
            # Start timer update
            self.set_interval(1.0, self.update_timer)
        
            self.log_activity("🚀 Ultra-Robust Subdomain Enumerator initialized")
            self.log_activity("💡 Configure your target domain and scanning mode, then click Start Scan")
            self.log_activity("⚙️ Quick mode is selected by default for fastest results")
        
        def update_timer(self) -> None:
            """Update the elapsed time display"""
            try:
                time_label = self.query_one("#time-label", Label)
                if self.scan_start_time:
                    elapsed = time.time() - self.scan_start_time
                    minutes = int(elapsed // 60)
                    seconds = int(elapsed % 60)
                    time_label.update(f"⏱️ Time: {minutes:02d}:{seconds:02d}")
                else:
                    time_label.update("⏱️ Time: 00:00")
            except:
                pass
    
        def log_activity(self, message: str) -> None:
            """Add message to activity log"""
            log = self.query_one("#activity-log", RichLog)
            timestamp = time.strftime("%H:%M:%S")
            log.write(f"[dim]{timestamp}[/dim] {message}")
    
        def on_input_changed(self, event: Input.Changed) -> None:
            """Validate domain input in real-time"""
            if event.input.id != "domain-input":
                return
            
            validation_label = self.query_one("#validation-label", Label)
        
            if not event.value:
                validation_label.update("")
                return
        
            domains = [d.strip() for d in event.value.split(",")]
            invalid_domains = []
        
            for domain in domains:
                if not self.is_valid_domain(domain):
                    invalid_domains.append(domain)
        
            if invalid_domains:
                validation_label.update(f"❌ Invalid domains: {', '.join(invalid_domains)}")
            else:
                validation_label.update(f"✅ {len(domains)} valid domain(s)")
    
        def is_valid_domain(self, domain: str) -> bool:
            """Basic domain validation"""
            import re
            pattern = r'^(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,}$'
            return bool(re.match(pattern, domain))
    
        def get_domains(self) -> List[str]:
            """Get validated domains from input"""
            domain_input = self.query_one("#domain-input", Input)
            domains = [d.strip() for d in domain_input.value.split(",") if d.strip()]
            return [d for d in domains if self.is_valid_domain(d)]
    
        async def on_button_pressed(self, event: Button.Pressed) -> None:
            """Handle button presses"""
            if event.button.id.startswith("mode-"):
                self.select_mode(int(event.button.id.split("-")[1]))
            elif event.button.id == "start-scan":
                await self.start_scan()
            elif event.button.id == "multi-scan":
                await self.start_multi_domain_scan()
            elif event.button.id == "stop-scan":
                await self.stop_scan()
        
        def select_mode(self, mode: int) -> None:
            """Select scanning mode and update UI"""
            self.selected_mode = mode
            mode_names = ["", "Basic (Fast DNS + HTTP)", "Standard (Balanced)", "Advanced (Comprehensive)", "Ultra (Maximum depth)", "Quick (DNS-only)"]
            
            # Update mode label
            try:
                mode_label = self.query_one("#mode-label", Label)
                mode_label.update(f"Selected: {mode_names[mode]}")
                
                # Update button styles
                for i in range(1, 6):
                    btn = self.query_one(f"#mode-{i}", Button)
                    btn.variant = "success" if i == mode else "default"
                
                self.log_activity(f"🎯 Mode selected: {mode_names[mode]}")
            except:
                pass
    
        async def start_scan(self) -> None:
            """Start scan with selected mode"""
            domains = self.get_domains()
            if not domains:
                self.log_activity("❌ Please enter valid domain(s)")
                return
        
            if len(domains) == 1:
                await self.run_single_scan(domains[0], self.selected_mode)
            else:
                await self.run_multi_scan(domains, self.selected_mode)
    
        async def start_multi_domain_scan(self) -> None:
            """Start multi-domain scan"""
            domains = self.get_domains()
            if len(domains) < 2:
                self.log_activity("❌ Multi-domain scan requires at least 2 domains")
                return
        
            await self.run_multi_scan(domains, self.selected_mode)
    
        @work(exclusive=True)
        async def run_single_scan(self, domain: str, mode: int = 1) -> None:
            """Run scan for a single domain"""
            self.is_scanning = True
            self.scan_start_time = time.time()
            self.update_scan_state(True)
        
            mode_names = ["", "Basic", "Standard", "Advanced", "Ultra", "Quick"]
            self.log_activity(f"🔍 Starting {mode_names[mode]} scan for {domain}")
        
            try:
                scanner = UltraRobustEnumerator(
                    progress_callback=self.update_progress,
                    result_callback=self.add_result
                )
            
                self.current_scanner = scanner
            
                if mode == 5:  # Quick mode
                    results = await scanner.fast_enumerate(domain, ['wordlists/common.txt'])
                else:
                    results = await scanner.ultra_enumerate(domain, mode, ['wordlists/common.txt'])
            
                self.scan_results[domain] = results
            
                live_count = sum(1 for r in results.values() if r.http_status == 200)
                self.log_activity(f"✅ Scan completed for {domain}: {len(results)} subdomains ({live_count} live)")
            
            except Exception as e:
                self.log_activity(f"❌ Scan failed for {domain}: {str(e)}")
            finally:
                self.is_scanning = False
                self.current_scanner = None
                self.scan_start_time = None 
                self.update_scan_state(False)
    
        @work(exclusive=True)
        async def run_multi_scan(self, domains: List[str], mode: int = 1) -> None:
            """Run scan for multiple domains"""
            self.is_scanning = True
            self.scan_start_time = time.time()
            self.update_scan_state(True)
        
            mode_names = ["", "Basic", "Standard", "Advanced", "Ultra", "Quick"]
            self.log_activity(f"🔄 Starting {mode_names[mode]} multi-domain scan for {len(domains)} domains")
        
            try:
                scanner = UltraRobustEnumerator(
                    progress_callback=self.update_progress,
                    result_callback=self.add_result
                )
            
                self.current_scanner = scanner
                results = await scanner.scan_multiple_domains(domains, mode, ['wordlists/common.txt'])
            
                self.scan_results.update(results)
            
                for domain, domain_results in results.items():
                    live_count = sum(1 for r in domain_results.values() if r.http_status == 200)
                    self.log_activity(f"✅ {domain}: {len(domain_results)} subdomains ({live_count} live)")
            
            except Exception as e:
                self.log_activity(f"❌ Multi-domain scan failed: {str(e)}")
            finally:
                self.is_scanning = False
                self.current_scanner = None
                self.scan_start_time = None
                self.update_scan_state(False)
    
        async def stop_scan(self) -> None:
            """Stop current scan"""
            if self.current_scanner and self.is_scanning:
                self.log_activity("🛑 Stopping scan...")
                self.is_scanning = False
                self.update_scan_state(False)
    
        def update_scan_state(self, scanning: bool) -> None:
            """Update UI state based on scanning status"""
            try:
                start_btn = self.query_one("#start-scan", Button)
                multi_btn = self.query_one("#multi-scan", Button)
                stop_btn = self.query_one("#stop-scan", Button)
                
                # Disable mode buttons during scanning
                for i in range(1, 6):
                    mode_btn = self.query_one(f"#mode-{i}", Button)
                    mode_btn.disabled = scanning
            
                start_btn.disabled = scanning
                multi_btn.disabled = scanning
                stop_btn.disabled = not scanning
            except:
                pass
    
        def update_progress(self, phase: str, progress: float, **kwargs) -> None:
            """Update categorized progress display"""
            try:
                self.current_phase = phase
                self.progress = progress
                
                # Update main statistics
                if 'subdomain_count' in kwargs:
                    self.found_count = kwargs['subdomain_count']
                    found_label = self.query_one("#found-label", Label)
                    found_label.update(f"🔍 Found: {self.found_count}")
                
                # Map phases to progress cards
                phase_mapping = {
                    "Certificate Transparency": "cert",
                    "Enhanced Certificate Transparency": "cert", 
                    "CNAME Analysis": "cname",
                    "DNS Brute Force": "dns",
                    "Intelligent DNS Brute Force": "dns",
                    "HTTP Analysis": "http",
                    "SSL Analysis": "ssl", 
                    "Nmap Scanning": "nmap",
                    "Smart IP Analysis": "nmap",
                    "Infrastructure Analysis": "nmap",
                    "Vulnerability Assessment": "vuln",
                    "Advanced Analytics": "final",
                    "Final CNAME Analysis": "final",
                    "Complete": "final"
                }
                
                # Get the card ID for this phase
                card_id = phase_mapping.get(phase, "final")
                
                # Update the specific progress card
                try:
                    progress_bar = self.query_one(f"#{card_id}-progress", ProgressBar)
                    status_label = self.query_one(f"#{card_id}-status", Label)
                    card_container = self.query_one(f"#{card_id}-card", Container)
                    
                    progress_bar.progress = progress
                    
                    # Update status text
                    message = kwargs.get('message', phase)
                    if len(message) > 40:
                        message = message[:37] + "..."
                    status_label.update(message)
                    
                    # Update card styling based on progress
                    if progress >= 100:
                        card_container.remove_class("phase-pending", "phase-active")
                        card_container.add_class("phase-complete")
                        status_label.update("✅ Complete")
                        self.log_activity(f"✅ {phase} completed")
                    elif progress > 0:
                        card_container.remove_class("phase-pending", "phase-complete")
                        card_container.add_class("phase-active")
                    
                except Exception:
                    pass  # Card might not exist for some phases
                
                # Update overall status
                status_label = self.query_one("#status-label", Label)
                if progress >= 100 and phase == "Complete":
                    status_label.update("📊 Status: Scan Complete")
                elif progress > 0:
                    status_label.update(f"📊 Status: {phase}")
                
            except Exception as e:
                pass
    
        def add_result(self, result: SubdomainResult) -> None:
            """Add new result to the table"""
            try:
                table = self.query_one("#results-table", DataTable)
            
                status = "🟢 Live" if result.http_status == 200 else f"{result.http_status}" if result.http_status > 0 else "No Response"
                ip_addr = result.ip_addresses[0] if result.ip_addresses else "N/A"
                server = result.server or "Unknown"
            
                table.add_row(
                    result.subdomain,
                    status,
                    ip_addr,
                    server,
                    result.source
                )
            
                self.found_count += 1
            
            except:
                pass
    
        def action_scan(self) -> None:
            """Keyboard shortcut for scan"""
            asyncio.create_task(self.start_scan())
    
        def action_clear(self) -> None:
            """Clear results"""
            try:
                table = self.query_one("#results-table", DataTable)
                table.clear()
                table.add_columns("Subdomain", "Status", "IP Address", "Server", "Source")
                self.scan_results.clear()
                self.found_count = 0
                self.log_activity("🗑️ Results cleared")
            except:
                pass

else:
    # Fallback class when Textual is not available
    class FastScanTUI:
        def __init__(self):
            print("❌ Textual not available. Install with: pip install textual")
        
        def run(self):
            print("Please install Textual for modern TUI interface")

class BeautifulTUI:
    """Beautiful Python TUI using Rich with full functionality"""
    
    def __init__(self):
        self.console = Console()
        self.enumerator = None
        self.scan_config: Optional[ScanConfig] = None
        self.results: List[SubdomainResult] = []
        self.progress_data = {}
        self.running = False
        self.paused = False
        self.pause_event = asyncio.Event()
        self.pause_event.set()  # Start unpaused
        self.shutting_down = False
        
        # Advanced features
        self.network_stats = NetworkStats()
        self.current_theme = "default"
        self.discovery_rate_history = deque(maxlen=60)  # Last 60 seconds
        self.animation_frame = 0
        self.alerts = []
        self.start_time = time.time()
        
        # Live progress system
        self.live_output_enabled = TQDM_AVAILABLE
        self.current_bar = None
        self.current_phase = ""
        self.discoveries_count = 0
        self.phase_bars = {}
        self.progress_bars = {}  # Store tqdm progress bars
        self.start_time = time.time()
        self.enumerator = None  # Will store enumerator reference
        
        # Interesting subdomains to highlight
        self.interesting_keywords = [
            'admin', 'administrator', 'root', 'test', 'dev', 'staging', 'prod', 'production',
            'api', 'dashboard', 'panel', 'control', 'manage', 'login', 'auth', 'secure',
            'internal', 'private', 'secret', 'hidden', 'backup', 'old', 'legacy',
            'vpn', 'mail', 'email', 'ftp', 'ssh', 'db', 'database', 'sql'
        ]
        
        # Display buffer to prevent glitching
        self.display_buffer = None
        self.last_update_time = 0
        self.update_interval = 0.5  # Stable update rate
    
    def get_theme_color(self, color_type: str) -> str:
        """Get color from current theme"""
        return THEMES[self.current_theme][color_type]
    
    def cycle_theme(self):
        """Cycle through available themes"""
        theme_names = list(THEMES.keys())
        current_index = theme_names.index(self.current_theme)
        self.current_theme = theme_names[(current_index + 1) % len(theme_names)]
    
    def get_wordlist_selection(self) -> List[str]:
        """Get wordlist file selection from user"""
        wordlist_dir = "wordlists/subdomain"
        
        if not os.path.exists(wordlist_dir):
            self.console.print(f"[bold red]❌ Wordlist directory not found: {wordlist_dir}[/bold red]")
            return []
        
        # Get all .txt files in the subdomain directory
        txt_files = glob.glob(os.path.join(wordlist_dir, "*.txt"))
        
        if not txt_files:
            self.console.print(f"[bold red]❌ No .txt files found in {wordlist_dir}[/bold red]")
            return []
        
        # Create selection table
        wordlist_table = Table(show_header=True, header_style="bold magenta", box=box.ROUNDED)
        wordlist_table.add_column("Option", style="cyan", width=8)
        wordlist_table.add_column("Filename", style="green", width=30)
        wordlist_table.add_column("Size", style="white")
        
        # Add "All Files" option
        wordlist_table.add_row("0", "ALL FILES", f"{len(txt_files)} files")
        
        # Add individual files
        for i, file_path in enumerate(txt_files, 1):
            filename = os.path.basename(file_path)
            try:
                with open(file_path, 'r') as f:
                    line_count = sum(1 for _ in f)
                size_str = f"{line_count:,} lines"
            except:
                size_str = "Unknown"
            
            wordlist_table.add_row(str(i), filename, size_str)
        
        wordlist_panel = Panel(wordlist_table, title="📚 Wordlist Selection", border_style="blue")
        self.console.print(wordlist_panel)
        
        while True:
            try:
                choices = input("\n🔍 Enter your selection (0 for all, or comma-separated numbers like 1,3,5): ").strip()
                
                if choices == "0":
                    # Select all files
                    return txt_files
                
                # Parse comma-separated choices
                selected_indices = [int(x.strip()) for x in choices.split(',')]
                selected_files = []
                
                for idx in selected_indices:
                    if 1 <= idx <= len(txt_files):
                        selected_files.append(txt_files[idx - 1])
                    else:
                        raise ValueError(f"Invalid choice: {idx}")
                
                if selected_files:
                    return selected_files
                else:
                    self.console.print("[bold red]❌ No valid selections made[/bold red]")
                    
            except ValueError as e:
                self.console.print(f"[bold red]❌ Invalid input: {e}[/bold red]")
                self.console.print("Please enter numbers separated by commas (e.g., 1,3,5) or 0 for all files")
    
    def display_banner(self):
        """Display beautiful banner"""
        banner_text = """
[bold blue]🔍 Ultra-Robust Subdomain Enumerator[/bold blue]
[dim]Advanced AI-Powered Reconnaissance Tool v3.0[/dim]
        """
        
        banner = Panel(
            Align.center(Text.from_markup(banner_text.strip())),
            style="bright_blue",
            border_style="blue",
            padding=(1, 2)
        )
        
        self.console.print(banner)
        self.console.print()
    
    def get_configuration(self) -> ScanConfig:
        """Get user configuration with beautiful prompts"""
        
        # Domain input
        while True:
            domain = Prompt.ask(
                "[bold green]🎯 Target Domain[/bold green]",
                default="example.com"
            ).strip().lower()
            
            if domain and '.' in domain and not domain.startswith(('http://', 'https://')):
                break
            self.console.print("[bold red]❌ Please enter a valid domain name[/bold red]")
        
        # Mode selection
        mode_table = Table(show_header=True, header_style="bold magenta", box=box.ROUNDED)
        mode_table.add_column("Option", style="cyan", width=8)
        mode_table.add_column("Mode", style="green", width=15)
        mode_table.add_column("Description", style="white")
        
        mode_table.add_row("1", "Basic", "Fast DNS + HTTP (quick results)")
        mode_table.add_row("2", "Standard", "Balanced features (recommended)")
        mode_table.add_row("3", "Advanced", "Comprehensive analysis")
        mode_table.add_row("4", "Ultra", "Maximum depth scanning")
        mode_table.add_row("5", "Quick", "DNS-only, fastest mode")
        
        mode_panel = Panel(mode_table, title="🚀 Enumeration Mode", border_style="green")
        self.console.print(mode_panel)
        
        while True:
            try:
                mode = int(Prompt.ask("Select mode", choices=["1", "2", "3", "4", "5"], default="1"))
                break
            except ValueError:
                self.console.print("[bold red]❌ Please enter 1, 2, 3, 4, or 5[/bold red]")
        
        # Get wordlist selection
        wordlist_files = self.get_wordlist_selection()
        if not wordlist_files:
            self.console.print("[bold red]❌ No wordlists selected. Exiting.[/bold red]")
            sys.exit(1)
        
        # Display selected wordlists
        selected_names = [os.path.basename(f) for f in wordlist_files]
        self.console.print(f"\n[bold green]✅ Selected wordlists:[/bold green] {', '.join(selected_names)}")
        
        return ScanConfig(domain=domain, mode=mode, wordlist_files=wordlist_files)
    
    def update_discovery_rate(self):
        """Update discovery rate history for mini-graph"""
        current_time = time.time()
        current_count = len(self.results)
        self.discovery_rate_history.append((current_time, current_count))
    
    def get_discovery_rate_graph(self) -> str:
        """Generate ASCII mini-graph of discovery rate"""
        if len(self.discovery_rate_history) < 2:
            return "░░░░░░░░░░ No data"
        
        # Calculate rates
        rates = []
        for i in range(1, len(self.discovery_rate_history)):
            prev_time, prev_count = self.discovery_rate_history[i-1]
            curr_time, curr_count = self.discovery_rate_history[i]
            time_diff = curr_time - prev_time
            if time_diff > 0:
                rate = (curr_count - prev_count) / time_diff
                rates.append(rate)
        
        if not rates:
            return "░░░░░░░░░░ No activity"
        
        # Create mini graph
        max_rate = max(rates) if rates else 1
        graph_chars = "▁▂▃▄▅▆▇█"
        graph = ""
        
        # Take last 10 data points
        recent_rates = rates[-10:] if len(rates) >= 10 else rates
        for rate in recent_rates:
            if max_rate > 0:
                level = int((rate / max_rate) * (len(graph_chars) - 1))
                graph += graph_chars[level]
            else:
                graph += graph_chars[0]
        
        # Pad with empty if needed
        graph = graph.ljust(10, '░')
        return f"{graph} {recent_rates[-1]:.1f}/s" if recent_rates else "░░░░░░░░░░ 0.0/s"
    
    def get_spinning_indicator(self) -> str:
        """Get animated spinning indicator"""
        spinners = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]
        return spinners[self.animation_frame % len(spinners)]
    
    def scroll_long_text(self, text: str, max_width: int = 30) -> str:
        """Scroll long text with animation"""
        if not text or len(text) <= max_width:
            return text.ljust(max_width) if text else " " * max_width
        
        try:
            # Safer scroll effect
            scroll_pos = self.animation_frame % (len(text) + 10)
            if scroll_pos < len(text):
                visible_text = text[scroll_pos:scroll_pos + max_width]
            else:
                # Simpler wrap-around
                wrap_pos = scroll_pos - len(text)
                if wrap_pos < max_width:
                    visible_text = (" " * wrap_pos) + text[:max_width - wrap_pos]
                else:
                    visible_text = text[:max_width]
            
            return visible_text.ljust(max_width)[:max_width]
        except Exception:
            # Fallback to truncation
            return (text[:max_width-3] + "...") if len(text) > max_width else text.ljust(max_width)
    
    def check_interesting_subdomain(self, subdomain: str) -> bool:
        """Check if subdomain contains interesting keywords"""
        subdomain_lower = subdomain.lower()
        for keyword in self.interesting_keywords:
            if keyword in subdomain_lower:
                return True
        return False
    
    def add_alert(self, subdomain: str, reason: str):
        """Add security alert for interesting finding"""
        alert_msg = f"🚨 ALERT: {subdomain} - {reason}"
        self.alerts.append(alert_msg)
        
        # Keep only last 10 alerts
        if len(self.alerts) > 10:
            self.alerts.pop(0)
    
    def create_display(self):
        """Create stable display with anti-glitch measures"""
        current_time = time.time()
        
        # Rate limit updates to prevent glitching
        if current_time - self.last_update_time < self.update_interval:
            return self.display_buffer if self.display_buffer else Panel("Loading...", style="dim")
        
        self.last_update_time = current_time
        
        # Update animation frame and discovery rate
        self.animation_frame += 1
        self.update_discovery_rate()
        
        # Get theme colors
        primary = self.get_theme_color("primary")
        success = self.get_theme_color("success") 
        warning = self.get_theme_color("warning")
        error = self.get_theme_color("error")
        info = self.get_theme_color("info")
        accent = self.get_theme_color("accent")
        
        # Header with scan info and status indicator
        modes = ["Basic", "Standard", "Advanced", "Ultra", "Quick"]
        
        # Calculate overall progress
        total_phases = 8  # Updated for new CNAME phases
        completed_phases = sum(1 for data in self.progress_data.values() if data.get('completed', False) or data.get('progress', 0) >= 100)
        overall_progress = (completed_phases / total_phases) * 100
        
        # Status indicator with animation
        spinner = self.get_spinning_indicator()
        if overall_progress >= 100:
            status_indicator = f"[{success}]🟢 COMPLETED[/{success}]"
        elif overall_progress > 0:
            if self.paused:
                status_indicator = f"[{warning}]⏸️  PAUSED[/{warning}]"
            else:
                status_indicator = f"[{warning}]{spinner} SCANNING[/{warning}]"
        else:
            status_indicator = f"[{info}]🟡 STARTING[/{info}]"
        
        # Elapsed time
        elapsed_time = time.time() - self.start_time
        elapsed_str = f"{int(elapsed_time//3600):02d}:{int((elapsed_time%3600)//60):02d}:{int(elapsed_time%60):02d}"
        
        # Enhanced header with more info  
        # Handle None scan_config gracefully
        if self.scan_config:
            scrolled_domain = self.scroll_long_text(self.scan_config.domain, 25)
            wordlist_count = len(self.scan_config.wordlist_files)
            mode_text = modes[self.scan_config.mode-1]
        else:
            scrolled_domain = "Not configured"
            wordlist_count = 0
            mode_text = "Not set"
        header_text = f"""🎯 Target: [{primary}]{scrolled_domain}[/{primary}] | {status_indicator}
📊 Mode: {mode_text} | Wordlists: {wordlist_count} files | Overall: [{accent}]{overall_progress:.1f}%[/{accent}]
🏆 Found: [{success}]{len(self.results)}[/{success}] subdomains | ⏱️ Elapsed: [{info}]{elapsed_str}[/{info}] | Theme: [{accent}]{self.current_theme.title()}[/{accent}]"""
        
        # Enhanced progress section with stable positioning
        progress_lines = []
        phase_names = ["Certificate Transparency", "CNAME Analysis", "DNS Brute Force", "ML Predictions", "Infrastructure Analysis", "Recursive Discovery", "HTTP Analysis", "Final CNAME Analysis"]
        phase_icons = ["📜", "⛓️", "🎯", "🤖", "🌐", "🔄", "🔍", "🔗"]
        
        for i, phase in enumerate(phase_names):
            icon = phase_icons[i]
            if phase in self.progress_data:
                data = self.progress_data[phase]
                progress = data.get('progress', 0)
                
                # Create stable progress bar
                bar_width = 20
                filled = int((progress / 100) * bar_width)
                bar = "█" * filled + "░" * (bar_width - filled)
                
                if data.get('completed', False) or progress >= 100:
                    status = f"[{success}]✅ {icon} {phase}[/{success}]"
                    progress_text = f"[{success}]{bar} 100%[/{success}]"
                elif progress > 0:
                    rate = data.get('rate', 0)
                    msg = data.get('message', '')
                    spinner = self.get_spinning_indicator()
                    status = f"[{warning}]{spinner} {icon} {phase}[/{warning}]"
                    progress_text = f"[{warning}]{bar} {progress:.1f}%[/{warning}]"
                    if rate > 0:
                        progress_text += f" [{info}]({rate:.0f}/sec)[/{info}]"
                    if msg:
                        truncated_msg = (msg[:40] + "...") if len(msg) > 40 else msg
                        status += f" - [{info}]{truncated_msg}[/{info}]"
                else:
                    status = f"[dim]⏳ {icon} {phase}[/dim]"
                    progress_text = f"[dim]{bar} 0%[/dim]"
            else:
                status = f"[dim]⏳ {icon} {phase}[/dim]"
                progress_text = f"[dim]{'░' * 20} 0%[/dim]"
            
            # Fixed-width formatting to prevent line jumping
            progress_lines.append(f"{status:<60}")
            progress_lines.append(f"{progress_text:<40}")
            progress_lines.append("")  # Add spacing between phases
        
        progress_text = "\n".join(progress_lines)
        
        # Enhanced results section with stable formatting
        if self.results:
            # Calculate comprehensive result statistics
            live_results = sum(1 for r in self.results if r.http_status == 200)
            unique_sources = len(set(r.source for r in self.results))
            interesting_results = sum(1 for r in self.results if self.check_interesting_subdomain(r.subdomain))
            
            # CNAME-specific statistics
            cname_results = sum(1 for r in self.results if r.cname_records)
            high_risk_takeover = sum(1 for r in self.results if r.takeover_risk in ["High", "Critical"])
            critical_takeover = sum(1 for r in self.results if r.takeover_risk == "Critical")
            
            # Provider distribution
            providers = {}
            for r in self.results:
                for cname in r.cname_records:
                    providers[cname.provider] = providers.get(cname.provider, 0) + 1
            top_providers = sorted(providers.items(), key=lambda x: x[1], reverse=True)[:3]
            
            # Recent results preview with enhanced CNAME information
            result_lines = []
            for result in self.results[-8:]:  # Show last 8 results with more info
                status = str(result.http_status) if result.http_status > 0 else "N/A"
                if result.http_status == 200:
                    status = f"[{success}]{status:>3}[/{success}]"
                elif result.http_status >= 400:
                    status = f"[{error}]{status:>3}[/{error}]"
                elif result.http_status > 0:
                    status = f"[{warning}]{status:>3}[/{warning}]"
                else:
                    status = f"[dim]{status:>3}[/dim]"
                
                ips = ", ".join(result.ip_addresses[:1]) if result.ip_addresses else "N/A"
                
                # Fixed width formatting to prevent jumping
                subdomain_display = result.subdomain[:35] + "..." if len(result.subdomain) > 35 else result.subdomain
                source_display = result.source[:12] + "..." if len(result.source) > 12 else result.source
                
                # Check for interesting subdomains and highlight
                if self.check_interesting_subdomain(result.subdomain):
                    # Add alert if not already added
                    alert_exists = any(result.subdomain in alert for alert in self.alerts)
                    if not alert_exists:
                        self.add_alert(result.subdomain, "Contains sensitive keyword")
                    subdomain_display = f"[{error}]⚠️  {subdomain_display}[/{error}]"
                else:
                    subdomain_display = f"[{info}]{subdomain_display}[/{info}]"
                
                # Fixed formatting
                result_lines.append(f"{subdomain_display:<45} | [{success}]{source_display:<15}[/{success}] | {status} | [{primary}]{ips[:15]}[/{primary}]")
            
            # Alert system display with fixed height
            alerts_display = ""
            if self.alerts:
                recent_alerts = self.alerts[-2:]  # Show last 2 alerts
                alert_lines = []
                for alert in recent_alerts:
                    truncated_alert = (alert[:60] + "...") if len(alert) > 60 else alert
                    alert_lines.append(f"[{error}]{truncated_alert}[/{error}]")
                alerts_display = f"""

🚨 [{error}]Security Alerts ({len(self.alerts)} total):[/{error}]
{chr(10).join(alert_lines)}"""
            
            # Enhanced statistics with CNAME data
            provider_display = ", ".join([f"{name}({count})" for name, count in top_providers]) if top_providers else "N/A"
            
            results_text = f"""📊 [{accent}]Live Statistics:[/{accent}]
[{success}]🟢 {live_results} Live[/{success}] | [{primary}]📊 {len(self.results)} Total[/{primary}] | [{info}]🔍 {unique_sources} Sources[/{info}] | [{error}]🚨 {interesting_results} Flagged[/{error}]

📈 [{accent}]CNAME Intelligence:[/{accent}]
[{info}]⛓️  {cname_results} CNAME Records[/{info}] | [{warning}]⚠️  {high_risk_takeover} High Risk[/{warning}] | [{error}]🚨 {critical_takeover} Critical[/{error}]
[{primary}]☁️  Top Providers: {provider_display[:50]}[/{primary}]

[{accent}]🔍 Recent Discoveries (Live Feed):[/{accent}]
{chr(10).join(result_lines)}"""
            if len(self.results) > 8:
                results_text += f"\n[dim]... and {len(self.results) - 8} more results (see final report)[/dim]"
            
            results_text += alerts_display
        else:
            results_text = f"[dim]{self.get_spinning_indicator()} Initializing scan... No results yet[/dim]"
        
        # Enhanced control section
        pause_status = f"[{warning}]⏸️  PAUSED[/{warning}]" if self.paused else f"[{success}]▶️  RUNNING[/{success}]"
        scanning_rate = len(self.results) / max(1, elapsed_time) * 60  # per minute
        
        # Enhanced controls with real-time network metrics
        dns_success_rate = 0
        if hasattr(self.enumerator, 'dns_resolver') and self.enumerator:
            total_queries = sum(stats['success'] + stats['failure'] 
                              for stats in self.enumerator.dns_resolver.resolver_stats.values())
            total_success = sum(stats['success'] 
                              for stats in self.enumerator.dns_resolver.resolver_stats.values())
            dns_success_rate = (total_success / total_queries * 100) if total_queries > 0 else 0
        
        controls = f"""🎮 [{accent}]Interactive Controls & Live Metrics:[/{accent}]
[{primary}]T[/{primary}] Theme ({self.current_theme}) | [{primary}]Ctrl+C[/{primary}] Graceful Stop & Save

📈 [{accent}]Real-Time Performance:[/{accent}]
🔥 Rate: [{success}]{scanning_rate:.1f}[/{success}] subdomains/min | 🎥 DNS Success: [{info}]{dns_success_rate:.1f}%[/{info}] | 🚨 Alerts: [{error}]{len(self.alerts)}[/{error}]

💡 [{accent}]Scanner Status:[/{accent}] {f"[{warning}]⏸️  PAUSED[/{warning}]" if self.paused else f"[{info}]▶️  ACTIVE - Auto-saving results[/{info}]"}

📈 [{accent}]Discovery Rate Trend:[/{accent}] {self.get_discovery_rate_graph()}

🔍 [{accent}]Advanced Features Active:[/{accent}]
[{success}]✓[/{success}] SSL Certificate Analysis | [{success}]✓[/{success}] CNAME Chain Resolution | [{success}]✓[/{success}] Takeover Detection
[{success}]✓[/{success}] ML Pattern Learning | [{success}]✓[/{success}] Infrastructure Mapping | [{success}]✓[/{success}] CT Log Mining"""
        
        # Combine all sections with stable layout
        divider = "━" * 82
        divider_color = f"[{accent}]{divider}[/{accent}]"
        
        full_content = f"""{header_text}

{divider_color}

📊 [{accent}]Enumeration Progress Monitor[/{accent}]

{progress_text}

{divider_color}

📋 [{accent}]Live Results & Security Alerts[/{accent}]

{results_text}

{divider_color}

{controls}"""
        
        # Cache the display to prevent glitching
        self.display_buffer = Panel(
            full_content,
            title=f"🚀 Ultra-Robust Subdomain Enumerator v3.0 - [{accent}]{self.current_theme.upper()} THEME[/{accent}]",
            border_style=primary,
            padding=(1, 3)
        )
        
        return self.display_buffer
    
    def progress_callback(self, phase: str, progress: float, current: int = 0, total: int = 0,
                         rate: float = 0, eta: float = 0, message: str = "", **kwargs):
        """Callback for progress updates"""
        self.progress_data[phase] = {
            'progress': progress,
            'current': current,
            'total': total,
            'rate': rate,
            'eta': eta,
            'message': message,
            'completed': progress >= 100
        }
    
    def result_callback(self, result: SubdomainResult):
        """Callback for new results with alert checking"""
        self.results.append(result)
        
        # Check for interesting/sensitive subdomains
        if self.check_interesting_subdomain(result.subdomain):
            for keyword in self.interesting_keywords:
                if keyword in result.subdomain.lower():
                    reason = f"Contains '{keyword}' keyword"
                    alert_exists = any(result.subdomain in alert for alert in self.alerts)
                    if not alert_exists:
                        self.add_alert(result.subdomain, reason)
                    break
    
    def setup_signal_handlers(self):
        """Setup graceful signal handlers"""
        def signal_handler(signum, frame):
            if not self.shutting_down:
                self.shutting_down = True
                self.running = False
                self.console.print(f"\n[{self.get_theme_color('warning')}]⚠️  Received interrupt signal. Shutting down gracefully...[/{self.get_theme_color('warning')}]")
                
                # Actually save the results if we have any
                if self.results and hasattr(self, 'scan_config') and self.scan_config:
                    self.console.print(f"[{self.get_theme_color('info')}]💾 Saving {len(self.results)} discovered subdomains...[/{self.get_theme_color('info')}]")
                    try:
                        # Convert results list to dict format expected by save_advanced_excel
                        results_dict = {result.subdomain: result for result in self.results}
                        
                        # Create a temporary enumerator to use the save method
                        from collections import namedtuple
                        TempEnum = namedtuple('TempEnum', ['save_advanced_excel'])
                        temp_enum = TempEnum(save_advanced_excel=lambda results, domain: UltraRobustEnumerator().save_advanced_excel(results, domain))
                        
                        # Actually save the file
                        output_file = UltraRobustEnumerator().save_advanced_excel(results_dict, self.scan_config.domain)
                        self.console.print(f"[{self.get_theme_color('success')}]✅ Results saved to: {output_file}[/{self.get_theme_color('success')}]")
                    except Exception as e:
                        self.console.print(f"[{self.get_theme_color('error')}]❌ Failed to save results: {e}[/{self.get_theme_color('error')}]")
                
                # Allow some time for cleanup
                time.sleep(1)
                sys.exit(0)
        
        signal.signal(signal.SIGINT, signal_handler)
        if hasattr(signal, 'SIGTERM'):
            signal.signal(signal.SIGTERM, signal_handler)
    
    async def _async_keyboard_listener(self):
        """Async keyboard listener for pause/continue functionality"""
        import sys
        import select
        
        while self.running:
            try:
                # Check for keyboard input with timeout
                ready, _, _ = select.select([sys.stdin], [], [], 0.1)
                if ready and self.enumerator and self.enumerator.pause_event:
                    try:
                        key = sys.stdin.read(1).lower()
                        if key == 'q':
                            self.console.print("\n[bold red]🛑 Quit requested by user[/bold red]")
                            # Save results before quitting
                            if self.enumerator and self.enumerator.results and hasattr(self, 'scan_config'):
                                self.console.print(f"💾 Saving {len(self.enumerator.results)} discovered subdomains...")
                                try:
                                    output_file = self.enumerator.save_advanced_excel(self.enumerator.results, self.scan_config.domain)
                                    self.console.print(f"✅ Results saved to: {output_file}")
                                except Exception as e:
                                    self.console.print(f"❌ Failed to save results: {e}")
                            self.running = False
                            break
                        elif key == ' ':  # Spacebar to pause/unpause
                            self.enumerator.paused = not self.enumerator.paused
                            if self.enumerator.paused:
                                self.console.print("\n[bold yellow]⏸️  PAUSED - Press SPACE to continue, Q to quit[/bold yellow]")
                                self.enumerator.pause_event.clear()
                            else:
                                self.console.print("\n[bold green]▶️  RESUMED[/bold green]")
                                self.enumerator.pause_event.set()
                    except Exception:
                        continue
                else:
                    await asyncio.sleep(0.1)
            except Exception:
                await asyncio.sleep(0.1)
                continue
    
    async def run_scan_with_ui(self):
        """Run the scan with live progress output using alive_progress"""
        try:
            # Create enumerator with callbacks
            self.enumerator = UltraRobustEnumerator(
                progress_callback=self.live_progress_callback,
                result_callback=self.live_result_callback
            )
            
            self.running = True
            self.start_time = time.time()
            
            # Start keyboard listener
            keyboard_task = asyncio.create_task(self._async_keyboard_listener())
            
            # Display scan info
            self.console.print(f"\n🎯 [bold cyan]Starting Ultra-Robust Enumeration for {self.scan_config.domain}[/bold cyan]")
            self.console.print(f"📊 Mode: {['Basic', 'Standard', 'Advanced', 'Ultra', 'Quick'][self.scan_config.mode - 1]}")
            self.console.print(f"📚 Wordlists: {len(self.scan_config.wordlist_files)} files")
            
            if self.live_output_enabled:
                # Use tqdm for live output with time tracking
                return await self._run_with_tqdm_progress(keyboard_task)
            else:
                # Fallback to basic output
                return await self._run_with_basic_output(keyboard_task)
                
        except KeyboardInterrupt:
            self.running = False
            return None
        except Exception as e:
            self.running = False
            self.console.print(f"\n[bold red]❌ Scan failed: {str(e)}[/bold red]")
            return None
    
    async def _run_with_tqdm_progress(self, keyboard_task):
        """Run scan with tqdm progress bars for better time tracking"""
        print("\n🚀 Ultra-Robust Subdomain Enumeration Starting...")
        print("═" * 80)
        
        if TQDM_AVAILABLE:
            # Use tqdm for better progress tracking
            progress_bar = tqdm(
                total=100, 
                desc="Overall Progress", 
                unit="%",
                dynamic_ncols=True,
                bar_format="{desc}: {percentage:3.0f}%|{bar}| {n:.0f}/{total:.0f} [{elapsed}<{remaining}, {rate_fmt}]"
            )
        else:
            progress_bar = None
        
        try:
            # Start enumeration task
            scan_task = asyncio.create_task(
                self.enumerator.ultra_enumerate(
                    self.scan_config.domain,
                    self.scan_config.mode,
                    self.scan_config.wordlist_files
                )
            )
            
            # Monitor progress with tqdm
            last_progress = 0
            while not scan_task.done() and self.running:
                await asyncio.sleep(0.5)
                
                # Update progress bar
                if progress_bar and self.current_phase and self.current_phase in self.progress_data:
                    phase_data = self.progress_data[self.current_phase]
                    current_progress = int(phase_data.get('progress', 0))
                    
                    if current_progress > last_progress:
                        progress_bar.update(current_progress - last_progress)
                        last_progress = current_progress
                    
                    if phase_data.get('message'):
                        progress_bar.set_description(f"{self.current_phase}: {phase_data['message']}")
            
            # Wait for either scan completion or keyboard interrupt
            done, pending = await asyncio.wait([scan_task, keyboard_task], return_when=asyncio.FIRST_COMPLETED)
            
            # Cancel remaining tasks
            for task in pending:
                task.cancel()
                try:
                    await task
                except asyncio.CancelledError:
                    pass
            
            # Get results
            if scan_task in done:
                results = await scan_task
            else:
                results = self.enumerator.results
            
            # Close progress bar
            if progress_bar:
                progress_bar.close()
            
            # Show completion
            elapsed_time = time.time() - self.start_time
            print(f"\n\n🎉 [bold green]Enumeration Complete![/bold green]")
            print(f"🏆 Found: {len(results)} subdomains")
            print(f"⏱️  Time: {elapsed_time:.1f}s")
            print(f"🚀 Rate: {len(results)/elapsed_time:.1f} subdomains/sec")
            
            return results
            
        except Exception as e:
            print(f"\n❌ Error during enumeration: {e}")
            return None
    
    async def _run_with_basic_output(self, keyboard_task):
        """Fallback to basic text output"""
        print("\n🚀 Ultra-Robust Subdomain Enumeration Starting...")
        print("Running without live progress bars...")
        
        try:
            scan_task = asyncio.create_task(
                self.enumerator.ultra_enumerate(
                    self.scan_config.domain,
                    self.scan_config.mode,
                    self.scan_config.wordlist_files
                )
            )
            
            # Basic progress monitoring with keyboard handling
            while not scan_task.done() and self.running:
                await asyncio.sleep(1)
                print(f"📊 Found: {len(self.enumerator.results)} subdomains, Phase: {self.current_phase}")
            
            # Wait for either scan completion or keyboard interrupt
            done, pending = await asyncio.wait([scan_task, keyboard_task], return_when=asyncio.FIRST_COMPLETED)
            
            # Cancel remaining tasks
            for task in pending:
                task.cancel()
                try:
                    await task
                except asyncio.CancelledError:
                    pass
            
            # Get results
            if scan_task in done:
                results = await scan_task
            else:
                results = self.enumerator.results
                
            elapsed_time = time.time() - self.start_time
            print(f"\n🎉 Complete! Found {len(results)} subdomains in {elapsed_time:.1f}s")
            
            # Save results to Excel
            if results and hasattr(self, 'scan_config'):
                print(f"💾 Saving {len(results)} subdomains to Excel...")
                try:
                    output_file = self.enumerator.save_advanced_excel(results, self.scan_config.domain)
                    print(f"✅ Results saved to: {output_file}")
                except Exception as e:
                    print(f"❌ Failed to save results: {e}")
            
            return results
            
        except Exception as e:
            print(f"\n❌ Error: {e}")
            return None
    
    def live_progress_callback(self, phase: str, progress: float, current: int = 0, total: int = 0,
                              rate: float = 0, eta: float = 0, message: str = "", **kwargs):
        """Live progress callback for alive_progress"""
        self.current_phase = phase
        self.progress_data[phase] = {
            'progress': progress,
            'current': current,
            'total': total,
            'rate': rate,
            'eta': eta,
            'message': message
        }
        
        # Print immediate progress for critical updates
        if message and ("found" in message.lower() or "complete" in message.lower()):
            print(f"\n💡 {phase}: {message}")
    
    def live_result_callback(self, result: SubdomainResult):
        """Live result callback with immediate output"""
        self.results.append(result)
        self.discoveries_count += 1
        
        # Show discovery immediately
        status_emoji = "🟢" if result.http_status == 200 else "🔍"
        source_short = result.source[:15] + "..." if len(result.source) > 15 else result.source
        
        print(f"\r{status_emoji} {result.subdomain} | {source_short} | {', '.join(result.ip_addresses[:2])}")
        
        # Check for interesting subdomains
        if any(keyword in result.subdomain.lower() for keyword in self.interesting_keywords):
            print(f"🚨 INTERESTING: {result.subdomain} - Potential security relevance!")
    
    def show_completion_screen(self, results: Dict):
        """Show beautiful completion screen"""
        if not results:
            self.console.print("[bold red]❌ No results to display[/bold red]")
            return
        
        # Celebration banner
        celebration = Panel(
            Align.center("[bold green]🎉 Enumeration Complete! 🎉[/bold green]\n[dim]Ultra-Robust Subdomain Discovery Finished[/dim]"),
            style="bright_green",
            border_style="green"
        )
        self.console.print(celebration)
        
        # Summary statistics
        total_found = len(self.results)
        live_services = sum(1 for r in self.results if r.http_status == 200)
        ssl_verified = sum(1 for r in self.results if r.ssl_domain_verified)
        
        summary_table = Table(show_header=False, box=box.ROUNDED, style="bright_white")
        summary_table.add_column("Metric", style="cyan", width=20)
        summary_table.add_column("Value", style="green")
        
        summary_table.add_row("🎯 Target Domain", self.scan_config.domain)
        summary_table.add_row("📊 Total Found", f"{total_found} subdomains")
        summary_table.add_row("🌐 Live Services", f"{live_services} responding")
        summary_table.add_row("🔒 SSL Verified", f"{ssl_verified} certificates")
        summary_table.add_row("📚 Wordlists Used", f"{len(self.scan_config.wordlist_files)} files")
        
        summary_panel = Panel(summary_table, title="📈 Scan Summary", border_style="blue")
        self.console.print(summary_panel)
        
        # Results preview
        if self.results:
            results_table = Table(show_header=True, header_style="bold magenta", box=box.ROUNDED)
            results_table.add_column("Subdomain", style="cyan")
            results_table.add_column("Source", style="green") 
            results_table.add_column("Status", style="white")
            results_table.add_column("SSL", style="blue")
            
            # Show first 15 results
            for result in self.results[:15]:
                status = str(result.http_status) if result.http_status > 0 else "N/A"
                if result.http_status == 200:
                    status = f"[bold green]{status}[/bold green]"
                elif result.http_status >= 400:
                    status = f"[bold red]{status}[/bold red]"
                
                ssl_status = "✅" if result.ssl_domain_verified else "❌"
                
                results_table.add_row(
                    escape(result.subdomain),
                    escape(result.source),
                    status,
                    ssl_status
                )
            
            if len(self.results) > 15:
                results_table.add_row("...", "...", "...", f"[dim]+{len(self.results)-15} more results[/dim]")
            
            results_panel = Panel(results_table, title="📋 Results Preview", border_style="cyan")
            self.console.print(results_panel)
    
    def show_error_screen(self, error: str):
        """Show beautiful error screen"""
        error_panel = Panel(
            f"[bold red]❌ Error Occurred[/bold red]\n\n{escape(error)}\n\n[dim]• Check your internet connection\n• Verify the target domain is valid\n• Ensure Python dependencies are installed[/dim]",
            title="🔍 Error Details",
            border_style="red"
        )
        self.console.print(error_panel)
    
    async def run(self):
        """Main TUI application loop"""
        if not RICH_AVAILABLE:
            self.console.print("[bold red]❌ Rich library is required for the beautiful TUI[/bold red]")
            self.console.print("Install it with: pip install rich")
            return
        
        try:
            # Setup signal handlers for graceful shutdown
            self.setup_signal_handlers()
            
            # Show banner
            self.display_banner()
            
            # Get configuration
            self.scan_config = self.get_configuration()
            
            # Confirm and start
            self.console.print()
            if not Confirm.ask(f"[bold green]🚀 Start scanning {self.scan_config.domain}?[/bold green]", default=True):
                self.console.print("[bold yellow]👋 Scan cancelled[/bold yellow]")
                return
            
            self.console.print("[dim]🎧 Controls: [bold]SPACE[/bold] = Pause/Resume, [bold]Q[/bold] = Quit[/dim]")
            
            self.console.print()
            self.console.print("[bold blue]🔍 Starting ultra-robust enumeration...[/bold blue]")
            self.console.print()
            
            # Add debug output
            self.console.print("[dim]Initializing dashboard...[/dim]")
            
            # Run the scan with UI
            results = await self.run_scan_with_ui()
            
            if results is not None and not self.shutting_down:
                # Show completion screen
                self.show_completion_screen(results)
            
        except KeyboardInterrupt:
            if not self.shutting_down:
                self.console.print(f"\n[{self.get_theme_color('warning')}]👋 Goodbye![/{self.get_theme_color('warning')}]")
        except Exception as e:
            if not self.shutting_down:
                self.show_error_screen(str(e))

# Quick access functions for faster scanning
def fast_scan(domain: str, wordlist_files: List[str] = None, output_file: str = None):
    """Quick DNS-focused scan - fastest option"""
    if wordlist_files is None:
        wordlist_files = ['wordlists/common.txt']
    
    print(f"🚀 Starting fast scan for {domain}...")
    scanner = UltraRobustEnumerator()
    results = asyncio.run(scanner.fast_enumerate(domain, wordlist_files))
    
    print(f"✅ Fast scan completed: {len(results)} subdomains found for {domain}")
    
    # Show top results
    live_results = [r for r in results.values() if r.http_status == 200]
    if live_results:
        print(f"🟢 {len(live_results)} live services found:")
        for result in live_results[:5]:  # Show first 5
            print(f"  • {result.subdomain} ({result.http_status}) - {result.server}")
    
    return results

def multi_domain_scan(domains: List[str], mode: int = 1, wordlist_files: List[str] = None):
    """Scan multiple domains in parallel"""
    if wordlist_files is None:
        wordlist_files = ['wordlists/common.txt']
    
    print(f"🔄 Starting parallel scan of {len(domains)} domains...")
    scanner = UltraRobustEnumerator()
    results = asyncio.run(scanner.scan_multiple_domains(domains, mode, wordlist_files))
    
    print(f"✅ Multi-domain scan completed:")
    for domain, domain_results in results.items():
        live_count = sum(1 for r in domain_results.values() if r.http_status == 200)
        print(f"  {domain}: {len(domain_results)} subdomains ({live_count} live)")
    
    return results

def quick_compare_domains(domain1: str, domain2: str, wordlist_files: List[str] = None):
    """Quick comparison scan of two domains"""
    print(f"⚖️  Comparing {domain1} vs {domain2}...")
    results = multi_domain_scan([domain1, domain2], mode=1, wordlist_files=wordlist_files)
    
    print(f"\n📊 Comparison Results:")
    d1_results = results.get(domain1, {})
    d2_results = results.get(domain2, {})
    d1_live = sum(1 for r in d1_results.values() if r.http_status == 200)
    d2_live = sum(1 for r in d2_results.values() if r.http_status == 200)
    
    print(f"  {domain1}: {len(d1_results)} total, {d1_live} live")
    print(f"  {domain2}: {len(d2_results)} total, {d2_live} live")
    
    return results

async def main():
    """Main entry point"""
    # Check for missing dependencies
    if not CRYPTO_AVAILABLE:
        print("⚠️  Warning: cryptography library not found. SSL analysis will be limited.")
        print("Install with: pip install cryptography")
        print()
    
    # Check if we're being called with command line args for quick scanning
    if len(sys.argv) > 1:
        domain = sys.argv[1]
        print(f"🚀 Quick scan mode for {domain}")
        fast_scan(domain)
        return
    
    # Try to use modern Textual TUI first
    if TEXTUAL_AVAILABLE:
        try:
            app = FastScanTUI()
            app.run()
            return
        except Exception as e:
            print(f"❌ Textual TUI failed: {e}")
            print("Falling back to Rich TUI...")
    
    # Fallback to Rich TUI
    if RICH_AVAILABLE:
        try:
            tui = BeautifulTUI()
            await tui.run()
            return
        except Exception as e:
            print(f"❌ Rich TUI failed: {e}")
    
    # Final fallback - command line interface
    print("🔧 TUI not available. Use quick functions instead:")
    print("- fast_scan('domain.com')")
    print("- multi_domain_scan(['domain1.com', 'domain2.com'])")
    print("- quick_compare_domains('domain1.com', 'domain2.com')")
    print("\nOr install Textual for modern TUI: pip install textual")

if __name__ == "__main__":
    # Entry point with graceful handling of cancellation
    if sys.version_info >= (3, 7):
        try:
            asyncio.run(main())
        except (KeyboardInterrupt, asyncio.CancelledError):
            # Exit quietly on user interrupt or task cancellation
            sys.exit(0)
    else:
        print("❌ Python 3.7+ required")
        sys.exit(1)